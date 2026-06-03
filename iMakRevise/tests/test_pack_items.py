"""pack_items.py + V8 計算への pack 反映 + xlsx 列追加 の unit test."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

THIS = Path(__file__).resolve().parent
PROJECT = THIS.parent
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from revise import pack_items, price_revise, review_xlsx
from revise.price_revise import ReviseCandidate, compute_new_usd


# ============================================================================
# pack_items load / get / detect
# ============================================================================
class TestLoadPackItems:
    def test_load_valid(self, tmp_path):
        p = tmp_path / "pack.json"
        p.write_text(json.dumps({
            "_comment": "meta",
            "356796294045": 3,
            "358390003755": 5,
        }), encoding="utf-8")
        result = pack_items.load_pack_items(p)
        # meta key 除外
        assert "_comment" not in result
        assert result["356796294045"] == 3
        assert result["358390003755"] == 5

    def test_load_missing_file(self, tmp_path):
        """ファイル不在 → 空 dict (fail-safe)"""
        p = tmp_path / "nope.json"
        assert pack_items.load_pack_items(p) == {}

    def test_load_invalid_json(self, tmp_path):
        p = tmp_path / "bad.json"
        p.write_text("invalid{json", encoding="utf-8")
        assert pack_items.load_pack_items(p) == {}


class TestGetPackCount:
    def test_registered_returns_pack(self):
        cache = {"356796294045": 3}
        assert pack_items.get_pack_count("356796294045", cache) == 3

    def test_unregistered_returns_1(self):
        cache = {"356796294045": 3}
        assert pack_items.get_pack_count("999", cache) == 1


class TestDetectPackSuspicion:
    def test_set_keyword(self):
        assert pack_items.detect_pack_suspicion("Pokemon カード セット")
        assert pack_items.detect_pack_suspicion("3枚セット 限定")

    def test_count_unit(self):
        assert pack_items.detect_pack_suspicion("UNIQLO エアリズム 3 枚入り")
        assert pack_items.detect_pack_suspicion("Mascot 5pcs Sanrio")
        assert pack_items.detect_pack_suspicion("ガシャポン 10 個")

    def test_multiplier(self):
        assert pack_items.detect_pack_suspicion("プラレール ×3 セット")

    def test_normal_title_not_flagged(self):
        """型番に数字あっても誤検出しない"""
        assert not pack_items.detect_pack_suspicion("G-Shock GA-100-1A1JF メンズ")
        assert not pack_items.detect_pack_suspicion("PSA10 リザードン")
        assert not pack_items.detect_pack_suspicion("Pokemon TCG カード")

    def test_empty(self):
        assert not pack_items.detect_pack_suspicion("")
        assert not pack_items.detect_pack_suspicion(None)


# ============================================================================
# compute_new_usd: pack 反映
# ============================================================================
class TestComputeNewUsdPack:
    @staticmethod
    def _candidate(item_id, new_jpy, pack_count=1, category="Tシャツ"):
        return ReviseCandidate(
            row_index=2, item_id=item_id, category=category,
            new_jpy=new_jpy, ah_jpy=None, f_jpy=None,
            delta_pct=0.0, basis="pending", title="",
            pack_count=pack_count,
        )

    def test_pack_5_uses_effective_cost(self):
        """pack=5 → V8 に N×5 が渡される"""
        c = self._candidate("358390003755", new_jpy=990, pack_count=5)
        captured_cost = []
        def v7_mock(cost_jpy, median_usd, category, country, title):
            captured_cost.append(cost_jpy)
            return {"price": 146.98, "shipping_usd": 20.0,
                    "shipping_profile_name": "DDP-C-P12",
                    "buyer_total_usd": 166.98, "profit_jpy": 500}
        compute_new_usd(c, v7_mock)
        assert captured_cost[0] == 990 * 5  # = 4950
        assert c.effective_cost == 4950
        assert c.new_usd == 146.98

    def test_pack_1_default(self):
        """pack=1 (= 通常) → N そのまま"""
        c = self._candidate("123", new_jpy=2000, pack_count=1)
        captured = []
        def v7_mock(cost_jpy, median_usd, category, country, title):
            captured.append(cost_jpy)
            return {"price": 50.98}
        compute_new_usd(c, v7_mock)
        assert captured[0] == 2000
        assert c.effective_cost == 2000


# ============================================================================
# review xlsx: pack 列
# ============================================================================
class TestReviewXlsxPack:
    @staticmethod
    def _candidate(item_id, pack_count=1, pack_suspect=False, new_usd=77.98):
        return ReviseCandidate(
            row_index=2, item_id=item_id, category="Tシャツ",
            new_jpy=990, ah_jpy=None, f_jpy=None,
            delta_pct=0.0, basis="decision", title=f"Item {item_id}",
            source_sheet="HIGH",
            new_usd=new_usd, shipping_profile_name="DDP-C-P09",
            pack_count=pack_count, pack_suspect=pack_suspect,
            revise_content="USD のみ",
        )

    def test_pack_columns_present(self, tmp_path):
        c = self._candidate("358390003755", pack_count=5)
        path = review_xlsx.write_review_xlsx(
            revisable=[c], output_dir=tmp_path,
            snapshot_map={"358390003755": {"price_usd": 146.98}},
            old_policy_map={"358390003755": {"shipping_profile_name": "DDP-C-P09"}},
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        assert "pack数" in headers
        assert "pack疑い" in headers
        pack_col = headers.index("pack数") + 1
        assert ws.cell(row=2, column=pack_col).value == 5

    def test_pack_suspect_flagged(self, tmp_path):
        c = self._candidate("999", pack_count=1, pack_suspect=True)
        path = review_xlsx.write_review_xlsx(
            revisable=[c], output_dir=tmp_path,
            snapshot_map={"999": {"price_usd": 50.0}},
            old_policy_map={},
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        suspect_col = headers.index("pack疑い") + 1
        assert "⚠️" in (ws.cell(row=2, column=suspect_col).value or "")


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
