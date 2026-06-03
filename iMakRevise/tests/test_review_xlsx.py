"""review_xlsx / snapshot_reader / ebay_trading_api の unit test."""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

THIS = Path(__file__).resolve().parent
PROJECT = THIS.parent
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from revise import snapshot_reader, ebay_trading_api, review_xlsx
from revise.price_revise import ReviseCandidate


# ============================================================================
# snapshot_reader
# ============================================================================
class TestSnapshotReader:
    def test_load_snapshot_basic(self, tmp_path):
        csv_path = tmp_path / "test_snapshot.csv"
        # BOM 付き UTF-8 (= eBay seller hub 出力 schema)
        csv_path.write_text(
            "﻿Item number,Title,Currency,Current price,Listing site,Available quantity\n"
            '"356641487278","Test item","USD",31.98,"US",1\n'
            '"358517889790","Another","USD",345.98,"US",0\n',
            encoding="utf-8",
        )
        result = snapshot_reader.load_snapshot(csv_path)
        assert "356641487278" in result
        assert result["356641487278"]["price_usd"] == 31.98
        assert result["356641487278"]["currency"] == "USD"
        assert result["356641487278"]["available_qty"] == 1
        assert result["358517889790"]["price_usd"] == 345.98
        assert result["358517889790"]["available_qty"] == 0  # 在庫切れ

    def test_load_snapshot_skips_empty_price(self, tmp_path):
        csv_path = tmp_path / "test_snapshot.csv"
        csv_path.write_text(
            "﻿Item number,Title,Currency,Current price\n"
            '"111","item1","USD",""\n'
            '"222","item2","USD",10\n',
            encoding="utf-8",
        )
        result = snapshot_reader.load_snapshot(csv_path)
        assert "111" not in result
        assert result["222"]["price_usd"] == 10.0

    def test_load_snapshot_missing_file(self):
        with pytest.raises(FileNotFoundError):
            snapshot_reader.load_snapshot(Path("/nonexistent.csv"))

    def test_get_old_price(self):
        sm = {"123": {"price_usd": 42.0}}
        assert snapshot_reader.get_old_price("123", sm) == 42.0
        assert snapshot_reader.get_old_price("999", sm) is None


# ============================================================================
# ebay_trading_api (mock)
# ============================================================================
class TestEbayTradingApi:
    def test_parse_shipping_profile_name(self):
        xml = (
            "<GetItemResponse>"
            "<Item><SellerProfiles><SellerShippingProfile>"
            "<ShippingProfileName>DDP-A-P09</ShippingProfileName>"
            "</SellerShippingProfile></SellerProfiles></Item>"
            "</GetItemResponse>"
        )
        assert ebay_trading_api._parse_shipping_profile_name(xml) == "DDP-A-P09"

    def test_parse_shipping_profile_name_missing(self):
        assert ebay_trading_api._parse_shipping_profile_name("<noop/>") is None

    def test_parse_current_price(self):
        xml = '<ConvertedCurrentPrice currencyID="USD">42.98</ConvertedCurrentPrice>'
        assert ebay_trading_api._parse_current_price(xml) == 42.98

    def test_parse_current_price_missing(self):
        assert ebay_trading_api._parse_current_price("<noop/>") is None

    def test_build_getitem_xml_contains_id(self):
        xml = ebay_trading_api._build_getitem_xml("358517889790")
        assert "<ItemID>358517889790</ItemID>" in xml
        assert "GetItemRequest" in xml

    def test_get_item_http_error(self, monkeypatch):
        """非 200 → error dict 返却"""
        import requests
        class _Resp:
            status_code = 401
            text = ""
        def fake_post(*a, **kw):
            return _Resp()
        monkeypatch.setattr(requests, "post", fake_post)
        result = ebay_trading_api.get_item("123", access_token="dummy")
        assert result["shipping_profile_name"] is None
        assert "HTTP 401" in result["error"]

    def test_get_item_network_error(self, monkeypatch):
        import requests
        def fake_post(*a, **kw):
            raise requests.ConnectionError("network down")
        monkeypatch.setattr(requests, "post", fake_post)
        result = ebay_trading_api.get_item("123", access_token="dummy")
        assert result["shipping_profile_name"] is None
        assert "ConnectionError" in result["error"]


# ============================================================================
# GetSellerList + snapshot save + rotation (2026-05-22 自動 DL)
# ============================================================================
class TestGetSellerList:
    def test_parse_seller_list_page(self):
        xml = '''
        <GetSellerListResponse>
          <Item>
            <ItemID>356700921169</ItemID>
            <Title>カウズ Tシャツ XL</Title>
            <Currency>USD</Currency>
            <ConvertedCurrentPrice currencyID="USD">61.98</ConvertedCurrentPrice>
            <QuantityAvailable>1</QuantityAvailable>
            <Site>US</Site>
            <SellerProfiles>
              <SellerShippingProfile>
                <ShippingProfileName>DDP-C-P07</ShippingProfileName>
              </SellerShippingProfile>
            </SellerProfiles>
          </Item>
          <Item>
            <ItemID>358517889790</ItemID>
            <Title>Pokemon TCG</Title>
            <Currency>USD</Currency>
            <ConvertedCurrentPrice currencyID="USD">357.98</ConvertedCurrentPrice>
            <QuantityAvailable>0</QuantityAvailable>
            <Site>US</Site>
          </Item>
          <HasMoreItems>false</HasMoreItems>
          <TotalNumberOfPages>1</TotalNumberOfPages>
        </GetSellerListResponse>
        '''
        items, has_more, total = ebay_trading_api._parse_seller_list_page(xml)
        assert len(items) == 2
        assert items[0]["item_id"] == "356700921169"
        assert items[0]["current_price"] == 61.98
        assert items[0]["available_qty"] == 1
        assert items[0]["shipping_profile_name"] == "DDP-C-P07"
        assert items[1]["available_qty"] == 0
        assert items[1]["shipping_profile_name"] == ""
        assert has_more is False
        assert total == 1

    def test_save_snapshot_csv_format_compatible(self, tmp_path):
        """save → load round-trip で snapshot_reader が読める (= format 互換)"""
        items = [
            {"item_id": "356700921169", "title": "T-shirt", "currency": "USD",
             "current_price": 61.98, "site": "US", "available_qty": 1,
             "shipping_profile_name": "DDP-C-P07"},
            {"item_id": "358517889790", "title": "TCG", "currency": "USD",
             "current_price": 357.98, "site": "US", "available_qty": 0,
             "shipping_profile_name": "Free"},
        ]
        path = ebay_trading_api.save_snapshot_csv(items, tmp_path)
        assert path.exists()
        assert path.name.startswith("ebay_active_")
        result = snapshot_reader.load_snapshot(path)
        assert "356700921169" in result
        assert result["356700921169"]["price_usd"] == 61.98
        assert result["356700921169"]["available_qty"] == 1
        assert result["358517889790"]["available_qty"] == 0

    def test_rotate_snapshots_keeps_5(self, tmp_path):
        import time as _time
        import os
        for i in range(7):
            p = tmp_path / f"ebay_active_2026-05-{15+i:02d}_060000.csv"
            p.write_text("dummy", encoding="utf-8")
            t = _time.time() - (7 - i) * 86400
            os.utime(p, (t, t))
        deleted = ebay_trading_api.rotate_snapshots(tmp_path, keep_count=5)
        assert len(deleted) == 2
        remaining = list(tmp_path.glob("ebay_active_*.csv"))
        assert len(remaining) == 5

    def test_rotate_snapshots_no_op_when_under_keep(self, tmp_path):
        for i in range(3):
            p = tmp_path / f"ebay_active_2026-05-{15+i:02d}_060000.csv"
            p.write_text("dummy", encoding="utf-8")
        deleted = ebay_trading_api.rotate_snapshots(tmp_path, keep_count=5)
        assert deleted == []
        remaining = list(tmp_path.glob("ebay_active_*.csv"))
        assert len(remaining) == 3

    def test_find_latest_snapshot_prefers_shared_dir(self, tmp_path, monkeypatch):
        """共有 dir に new file あれば優先、なければ fallback (= 旧デスクトップ)"""
        shared = tmp_path / "shared"
        fallback = tmp_path / "desktop"
        shared.mkdir()
        fallback.mkdir()
        (fallback / "eBay-all-active-listings-report-2026-05-21-X.csv").write_text(
            "d", encoding="utf-8")
        monkeypatch.setattr(snapshot_reader, "FALLBACK_SNAPSHOT_DIR", fallback)
        # 共有 dir 空 → fallback 拾う
        result = snapshot_reader.find_latest_snapshot(snapshot_dir=shared)
        assert result is not None
        assert "eBay-all-active" in result.name
        # 共有 dir に new file → そちらが優先
        (shared / "ebay_active_2026-05-22_060000.csv").write_text("d", encoding="utf-8")
        result = snapshot_reader.find_latest_snapshot(snapshot_dir=shared)
        assert result is not None
        assert result.name.startswith("ebay_active_")


# ============================================================================
# review_xlsx
# ============================================================================
class TestReviewXlsx:
    @staticmethod
    def _candidate(item_id, new_jpy=2000, ah_jpy=1500, new_usd=77.98,
                    policy="DDP-C-P07", is_abnormal=False, delta_pct=33.3,
                    profit_jpy=400):
        return ReviseCandidate(
            row_index=2, item_id=item_id, category="Tシャツ",
            new_jpy=new_jpy, ah_jpy=ah_jpy, f_jpy=None,
            delta_pct=delta_pct, basis="AH", title=f"Item {item_id}",
            source_sheet="HIGH",
            new_usd=new_usd, shipping_profile_name=policy,
            buyer_total_usd=(new_usd + 16.87) if new_usd else None,
            profit_jpy=profit_jpy,
            is_abnormal=is_abnormal,
        )

    def test_basic_write(self, tmp_path):
        candidates = [self._candidate("357111565952")]
        snapshot_map = {"357111565952": {"price_usd": 65.98}}
        old_policy_map = {"357111565952": {"shipping_profile_name": "DDP-C-P06"}}
        path = review_xlsx.write_review_xlsx(
            revisable=candidates,
            output_dir=tmp_path,
            snapshot_map=snapshot_map,
            old_policy_map=old_policy_map,
        )
        assert path.exists()
        assert path.suffix == ".xlsx"
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        assert "ItemID" in headers
        assert "旧USD" in headers
        assert "新USD" in headers
        assert "旧Policy" in headers
        assert "新Policy" in headers
        # data row
        row = [c.value for c in ws[2]]
        item_col = headers.index("ItemID")
        assert row[item_col] == "357111565952"
        old_usd_col = headers.index("旧USD")
        new_usd_col = headers.index("新USD")
        assert row[old_usd_col] == 65.98
        assert row[new_usd_col] == 77.98

    def test_abnormal_appears_with_flag(self, tmp_path):
        normal = [self._candidate("111")]
        abnormal = [self._candidate("999", new_jpy=363000, ah_jpy=16000,
                                     delta_pct=2169.0, is_abnormal=True,
                                     new_usd=None, profit_jpy=None)]
        path = review_xlsx.write_review_xlsx(
            revisable=normal,
            output_dir=tmp_path,
            snapshot_map={"111": {"price_usd": 50.0}},
            old_policy_map={},
            abnormal=abnormal,
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        anomaly_col = headers.index("異常検出")
        # 行2 (normal): 異常検出 空 (openpyxl は空文字を None として保存)
        normal_val = ws.cell(row=2, column=anomaly_col + 1).value
        assert normal_val in (None, "")
        # 行3 (abnormal): ABNORMAL_DELTA 入る
        assert "ABNORMAL_DELTA" in ws.cell(row=3, column=anomaly_col + 1).value

    def test_policy_change_flag(self, tmp_path):
        c = self._candidate("111", policy="DDP-A-P09")
        path = review_xlsx.write_review_xlsx(
            revisable=[c],
            output_dir=tmp_path,
            snapshot_map={"111": {"price_usd": 50.0}},
            old_policy_map={"111": {"shipping_profile_name": "DDP-A-P05"}},  # 旧 != 新
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        change_col = headers.index("Policy変更?")
        assert "変更" in ws.cell(row=2, column=change_col + 1).value

    def test_missing_old_policy(self, tmp_path):
        """Trading API 失敗 → 旧Policy = (取得失敗)"""
        c = self._candidate("111")
        path = review_xlsx.write_review_xlsx(
            revisable=[c],
            output_dir=tmp_path,
            snapshot_map={},
            old_policy_map={},
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        old_policy_col = headers.index("旧Policy")
        assert "取得失敗" in ws.cell(row=2, column=old_policy_col + 1).value

    def test_summary_sheet_present(self, tmp_path):
        """1 枚目に summary sheet が来る"""
        c1 = self._candidate("111")
        c1.revise_content = "USD のみ"
        c2 = self._candidate("222", policy="DDP-A-P10")
        c2.revise_content = "USD+Policy"
        path = review_xlsx.write_review_xlsx(
            revisable=[c1, c2],
            output_dir=tmp_path,
            snapshot_map={"111": {"price_usd": 50}},
            old_policy_map={"111": {"shipping_profile_name": "Free"},
                            "222": {"shipping_profile_name": "DDP-A-P05"}},
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        # 1 枚目 = summary
        assert wb.sheetnames[0] == "summary"
        assert "review" in wb.sheetnames
        # summary に集計値あり
        ws_sum = wb["summary"]
        all_text = " ".join(str(c.value) for row in ws_sum.iter_rows() for c in row if c.value)
        assert "リバイスくん" in all_text
        assert "revise 対象" in all_text or "合計" in all_text

    def test_diff_cells_highlighted(self, tmp_path):
        """旧 ≠ 新 のセルが黄色ハイライトされる"""
        c = self._candidate("111", policy="DDP-A-P10")
        c.revise_content = "USD+Policy"
        path = review_xlsx.write_review_xlsx(
            revisable=[c],
            output_dir=tmp_path,
            snapshot_map={"111": {"price_usd": 50.0}},  # vs new_usd=77.98
            old_policy_map={"111": {"shipping_profile_name": "DDP-A-P05"}},  # vs new DDP-A-P10
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        headers = [c.value for c in ws[1]]
        usd_col = headers.index("旧USD") + 1
        # 旧USD セルに黄色塗りつぶしがある
        fill = ws.cell(row=2, column=usd_col).fill
        assert fill.start_color.rgb in ("00FFEB9C", "FFFFEB9C")  # 黄色

    def test_bold_for_big_diff(self, tmp_path):
        """USD 差 |≥ $10| → 太字"""
        c = self._candidate("111", new_usd=77.98)
        c.revise_content = "USD のみ"
        path = review_xlsx.write_review_xlsx(
            revisable=[c],
            output_dir=tmp_path,
            snapshot_map={"111": {"price_usd": 50.0}},  # diff = $27.98 → ≥ $10
            old_policy_map={"111": {"shipping_profile_name": "DDP-C-P07"}},  # 同 Policy
        )
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["review"]
        # ItemID セルが太字
        assert ws.cell(row=2, column=2).font.bold is True


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
