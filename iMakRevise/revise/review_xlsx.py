"""review_xlsx.py - 人間レビュー用 ビフォーアフター xlsx (2 sheet 構成).

sheet 構成:
  1. summary  : 全体俯瞰 (= 件数集計、最初に開いた時に見える)
  2. review   : 1 行 1 listing の詳細比較

旧価格 = HQ snapshot CSV (`eBay-all-active-listings-report-*.csv`)
旧 Policy = eBay Trading API GetItem
新価格 / 新 Policy / 旧仕入¥ / 新仕入¥ = Revise 内部 (ReviseCandidate)

color/style:
  - 異常検出行: 行全体 赤背景
  - 旧 ≠ 新 のセル: 黄色 highlight
  - USD 差 |≥ $10| or Policy 変更ある行: 太字
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional


def _summary_counts(revisable: list, abnormal: list,
                     all_skipped: Optional[list] = None) -> dict:
    """summary sheet 用集計."""
    counts = {
        "price_diff_only": 0,
        "policy_change_only": 0,
        "both": 0,
        "skipped_reasons": {},
        "pack_registered": 0,
        "pack_suspect": 0,
        "variation_rows": 0,
        "variation_listings": 0,
    }
    for c in revisable:
        rc = c.revise_content or ""
        if rc == "USD のみ":
            counts["price_diff_only"] += 1
        elif rc == "Policy のみ":
            counts["policy_change_only"] += 1
        elif rc == "USD+Policy":
            counts["both"] += 1
    counts["total_revise"] = (
        counts["price_diff_only"] + counts["policy_change_only"] + counts["both"]
    )
    counts["abnormal"] = len(abnormal)
    # pack 集計 (= revisable + abnormal + skipped 全候補で集計、漏れ防止)
    var_listing_ids = set()
    for c in list(revisable) + list(abnormal) + list(all_skipped or []):
        if getattr(c, "pack_count", 1) > 1:
            counts["pack_registered"] += 1
        if getattr(c, "pack_suspect", False):
            counts["pack_suspect"] += 1
        if getattr(c, "is_variation", False):
            counts["variation_rows"] += 1
            var_listing_ids.add(c.item_id)
    counts["variation_listings"] = len(var_listing_ids)
    for c in (all_skipped or []):
        r = c.decision_reason or "unknown"
        counts["skipped_reasons"][r] = counts["skipped_reasons"].get(r, 0) + 1
    return counts


def write_review_xlsx(
    revisable: list,
    output_dir: Path,
    snapshot_map: Optional[dict] = None,
    old_policy_map: Optional[dict] = None,
    abnormal: Optional[list] = None,
    all_skipped: Optional[list] = None,
    filename_prefix: str = "revise_review",
) -> Path:
    """ビフォーアフター 比較 xlsx を生成 (user 目視承認用).

    snapshot_map: {ItemID: {"price_usd": ..., "available_qty": ..., ...}}
    old_policy_map: {ItemID: {"shipping_profile_name": ..., ...}}
    abnormal: 異常 delta で skip した candidates
    all_skipped: revise しなかった全 candidates (= summary 集計用、abnormal も含む)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError(f"openpyxl 未インストール: {e}")

    snapshot_map = snapshot_map or {}
    old_policy_map = old_policy_map or {}
    abnormal = abnormal or []

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"{filename_prefix}_{ts}.xlsx"

    wb = Workbook()

    # ----- style 共通 -----
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    SUMMARY_LABEL_FONT = Font(bold=True)
    ABNORMAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    DIFF_CELL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BOLD_FONT = Font(bold=True)

    # ============================================================
    # Sheet 1: summary (全体俯瞰)
    # ============================================================
    ws_sum = wb.active
    ws_sum.title = "summary"
    counts = _summary_counts(revisable, abnormal, all_skipped)

    ws_sum.append(["リバイスくん dry-run summary", ""])
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.append(["生成時刻", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_sum.append([])

    ws_sum.append(["■ revise 対象 (= eBay UP 件数)", ""])
    ws_sum["A4"].font = SUMMARY_LABEL_FONT
    ws_sum["A4"].fill = HEADER_FILL
    ws_sum["A4"].font = HEADER_FONT
    ws_sum["B4"].fill = HEADER_FILL

    rows_revise = [
        ["price_diff のみ (= 価格だけ変更)", counts["price_diff_only"]],
        ["policy_change のみ (= Policy だけ変更)", counts["policy_change_only"]],
        ["price_diff + policy_change (= 両方変更)", counts["both"]],
        ["合計 (= revise 対象)", counts["total_revise"]],
    ]
    for label, n in rows_revise:
        ws_sum.append([label, n])
    # 合計行 太字
    ws_sum.cell(row=ws_sum.max_row, column=1).font = BOLD_FONT
    ws_sum.cell(row=ws_sum.max_row, column=2).font = BOLD_FONT

    ws_sum.append([])
    ws_sum.append(["■ skip 内訳 (参考)", ""])
    skip_header_row = ws_sum.max_row
    ws_sum.cell(row=skip_header_row, column=1).font = HEADER_FONT
    ws_sum.cell(row=skip_header_row, column=1).fill = HEADER_FILL
    ws_sum.cell(row=skip_header_row, column=2).fill = HEADER_FILL

    skip_order = [
        ("aligned", "aligned (= 完全一致、変更不要)"),
        ("policy_change", "policy_change (revise 側に集計済)"),
        ("price_diff", "price_diff (revise 側に集計済)"),
        ("not_in_snapshot", "not_in_snapshot (= snapshot 不在、取下げ済 listing 想定)"),
        ("out_of_stock", "out_of_stock (= Available qty < 1)"),
        ("no_snapshot", "no_snapshot (= 部分取得失敗)"),
        ("abnormal_delta", "⚠️ abnormal_delta (= scrape miss 疑い)"),
        ("variation_no_sku_in_spreadsheet", "⚠️ variation_no_sku_in_spreadsheet (= snapshot側 variation だがスプシ SKU 未登録、 監視くん巡回不足)"),
        ("v8_calc_failed", "v8_calc_failed (= Unknown category 等)"),
        ("loss", "loss (= 利益 < 0)"),
        ("sold", "sold (= D 列 SOLD マーカー)"),
        ("no_item_id", "no_item_id (= ItemID 空欄)"),
        ("no_cost", "no_cost (= N 空欄/range外)"),
    ]
    for key, label in skip_order:
        n = counts["skipped_reasons"].get(key, 0)
        if n > 0 and key not in ("policy_change", "price_diff"):
            ws_sum.append([f"  - {label}", n])

    # pack 関連集計 (= 2026-05-22 pack 対応)
    if counts.get("pack_registered", 0) > 0 or counts.get("pack_suspect", 0) > 0:
        ws_sum.append([])
        ws_sum.append(["■ pack 関連", ""])
        pack_header_row = ws_sum.max_row
        ws_sum.cell(row=pack_header_row, column=1).font = HEADER_FONT
        ws_sum.cell(row=pack_header_row, column=1).fill = HEADER_FILL
        ws_sum.cell(row=pack_header_row, column=2).fill = HEADER_FILL
        ws_sum.append([
            "  - pack 商品 (mapping 登録済)", counts.get("pack_registered", 0)
        ])
        ws_sum.append([
            "  - ⚠️ pack 疑い (未登録、要確認)", counts.get("pack_suspect", 0)
        ])

    # variation listing 集計 (= 2026-05-24)
    if counts.get("variation_listings", 0) > 0:
        ws_sum.append([])
        ws_sum.append(["■ variation listing 関連", ""])
        var_header_row = ws_sum.max_row
        ws_sum.cell(row=var_header_row, column=1).font = HEADER_FONT
        ws_sum.cell(row=var_header_row, column=1).fill = HEADER_FILL
        ws_sum.cell(row=var_header_row, column=2).fill = HEADER_FILL
        ws_sum.append([
            "  - variation listing 数 (ItemID 単位)", counts.get("variation_listings", 0)
        ])
        ws_sum.append([
            "  - variation 行数 (= SKU 単位)", counts.get("variation_rows", 0)
        ])

    ws_sum.column_dimensions["A"].width = 55
    ws_sum.column_dimensions["B"].width = 12

    # ============================================================
    # Sheet 2: review (1 行 1 listing 詳細)
    # ============================================================
    ws = wb.create_sheet("review")

    headers = [
        "sheet", "ItemID", "SKU/Size",
        "Title", "Category",
        "旧仕入¥", "新仕入¥", "仕入差%",
        "旧USD", "新USD", "USD差",
        "旧Policy", "新Policy", "Policy変更?",
        "revise内容", "判定理由",
        "在庫", "利益¥", "異常検出",
        "pack数", "pack疑い",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    all_rows = list(revisable) + list(abnormal)
    for c in all_rows:
        old_price = snapshot_map.get(c.item_id, {}).get("price_usd")
        snap_qty = snapshot_map.get(c.item_id, {}).get("available_qty")
        old_policy_entry = old_policy_map.get(c.item_id) or {}
        old_policy = old_policy_entry.get("shipping_profile_name")
        if old_price is None and old_policy_entry.get("current_price_usd") is not None:
            old_price = old_policy_entry.get("current_price_usd")

        new_price = c.new_usd
        new_policy = c.shipping_profile_name

        usd_diff = None
        if old_price is not None and new_price is not None:
            usd_diff = round(new_price - old_price, 2)

        policy_changed = ""
        if old_policy and new_policy and old_policy != new_policy:
            policy_changed = "⚠️ 変更"
        elif old_policy and new_policy and old_policy == new_policy:
            policy_changed = "維持"
        elif not old_policy:
            policy_changed = "(旧不明)"

        anomaly_flag = ""
        if c.is_abnormal:
            anomaly_flag = f"⚠️ ABNORMAL_DELTA (+{c.delta_pct:.0f}%)"

        qty_display = c.available_qty if c.available_qty is not None else (snap_qty or "")

        pack_count = getattr(c, "pack_count", 1) or 1
        pack_suspect = getattr(c, "pack_suspect", False)
        # variation 表示 (= 2026-05-24)
        is_var = getattr(c, "is_variation", False)
        if is_var:
            size = getattr(c, "variation_size", "")
            color = getattr(c, "variation_color", "")
            spec_parts = [s for s in [size, color] if s]
            sku_display = "/".join(spec_parts) if spec_parts else (c.sku[:8] if c.sku else "")
        else:
            sku_display = ""
        row_values = [
            c.source_sheet,
            c.item_id,
            sku_display,
            (c.title or "")[:80],
            c.category,
            int(c.ah_jpy) if c.ah_jpy else "",
            int(c.new_jpy) if c.new_jpy else "",
            f"{c.delta_pct:+.1f}%" if c.delta_pct is not None else "",
            old_price if old_price is not None else "",
            new_price if new_price is not None else "",
            usd_diff if usd_diff is not None else "",
            old_policy or "(取得失敗)",
            new_policy or "",
            policy_changed,
            c.revise_content or "",
            c.decision_reason or "",
            qty_display,
            int(c.profit_jpy) if c.profit_jpy is not None else "",
            anomaly_flag,
            pack_count if pack_count > 1 else "",
            "⚠️" if pack_suspect else "",
        ]
        ws.append(row_values)
        row_num = ws.max_row

        # 異常: 行全体 赤背景
        if c.is_abnormal:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = ABNORMAL_FILL
            continue

        # 旧 ≠ 新 セルを黄色ハイライト
        if old_price is not None and new_price is not None and old_price != new_price:
            ws.cell(row=row_num, column=headers.index("旧USD") + 1).fill = DIFF_CELL_FILL
            ws.cell(row=row_num, column=headers.index("新USD") + 1).fill = DIFF_CELL_FILL
        if old_policy and new_policy and old_policy != new_policy:
            ws.cell(row=row_num, column=headers.index("旧Policy") + 1).fill = DIFF_CELL_FILL
            ws.cell(row=row_num, column=headers.index("新Policy") + 1).fill = DIFF_CELL_FILL
            ws.cell(row=row_num, column=headers.index("Policy変更?") + 1).fill = DIFF_CELL_FILL

        # 影響大 (USD 差 ≥$10 or Policy 変更ある) 行は太字
        if (usd_diff is not None and abs(usd_diff) >= 10) or (
            old_policy and new_policy and old_policy != new_policy
        ):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = BOLD_FONT

    # 列幅 (= 21 列、2026-05-24 SKU/Size 追加で 1 列増)
    col_widths = {
        "A": 8, "B": 14, "C": 14,  # sheet / ItemID / SKU/Size
        "D": 50, "E": 12,           # Title / Category
        "F": 10, "G": 10, "H": 10,  # 旧仕入¥ / 新仕入¥ / 仕入差%
        "I": 9, "J": 9, "K": 9,     # 旧USD / 新USD / USD差
        "L": 14, "M": 14, "N": 12,  # 旧Policy / 新Policy / Policy変更?
        "O": 14, "P": 14,           # revise内容 / 判定理由
        "Q": 6, "R": 10, "S": 30,   # 在庫 / 利益¥ / 異常検出
        "T": 7, "U": 8,             # pack数 / pack疑い
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    # summary を最初に開かれる sheet に
    wb.active = wb.index(ws_sum)

    wb.save(path)
    return path
