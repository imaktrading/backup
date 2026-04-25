#!/usr/bin/env python3
"""
iMak Trading Japan - メルカリスカウト
保存した検索条件を自動巡回 → eBay GATE判定 → 仕入GOリスト出力

使い方:
  初回ログイン:  python mercari_scout.py --login
  通常巡回:      python mercari_scout.py
"""

import csv
import sys
import os
import re
import json
import time
import base64
import requests
from datetime import datetime
from urllib.parse import quote, urlparse, parse_qs

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By

# ===== 設定 =====
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CHROME_PROFILE_DIR = r"C:\Users\imax2\local_data\iMakMercari\chrome_profile"
SEARCH_URLS_FILE = os.path.join(SCRIPT_DIR, "search_urls.txt")
SCOUT_LOG_FILE = os.path.join(SCRIPT_DIR, "data", "scout_log.csv")
PSA_CERT_CACHE_FILE = os.path.join(SCRIPT_DIR, "data", "mercari_psa_cache.json")
GO_LIST_FILE = os.path.join(SCRIPT_DIR, "data", "go_list.csv")

# Google Sheets（在庫管理スプシ）
INVENTORY_SHEET_IDS = [
    "1RbGaiQxhYDd7s8nqT0jHeh7sQ6FJNCVnVxkEJLFmz9s",  # TCG在庫管理
    "1QI0-L1A1DfTEi8Hl1-EFuRl9oTw9QPFe3X85stnaOD4",   # Tシャツ管理
]

# eBay API
EBAY_KEYS_FILE = os.path.join(SCRIPT_DIR, "..", "iMakeBayAPI", "ebay keys.txt")

# 利益計算パラメータ（SSOT 抽象化: profit_params.get_check_csv_params 経由）
# 2026-04-25 Step 7 拡張: ハードコード撲滅、yaml(SSOT) 注入型に統一。
#   旧: EBAY_FEE=0.185 / SHIPPING_JPY=2000 等を直接記述
#   新: Mercari は Tシャツ(UT) カテゴリで FVF=15.3% (yaml と一致)
import sys as _sys_pp
_sys_pp.path.insert(0, os.path.join(SCRIPT_DIR, "..", "iMakeBayAPI"))
from profit_params import get_check_csv_params as _gccp_pp
_pp = _gccp_pp("Tシャツ(UT)")
EXCHANGE_RATE = _pp["exchange_rate"]
EBAY_FEE      = _pp["ebay_fee_rate"]
PROMO_RATE    = _pp["promo_rate"]
PAYO_RATE     = _pp["payo_rate"]
SHIPPING_JPY  = _pp["shipping_jpy"]  # デフォルト。カテゴリで変える場合は検索URLに紐付ける
NET_RATIO = 1 - EBAY_FEE - PROMO_RATE - PAYO_RATE

# 価格帯別パラメータ: SSOT 抽象化 (profit_params.get_tier_params 経由)
# 旧: 本ファイル内に TIER_PARAMS リスト定義 (6ファイル重複)
# 新: yaml(global.yaml) の pricing_tiers を SSOT、共通 API で取得
from profit_params import get_tier_params  # noqa: F401  (re-export for back-compat)


# ===== Chromeプロファイル管理 =====
def create_driver(headless=False):
    """Chromeプロファイルを保持するドライバーを作成。
    ログイン情報（Cookie + localStorage + セッション）が全て保持される。"""
    os.makedirs(CHROME_PROFILE_DIR, exist_ok=True)
    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument(f"--user-data-dir={CHROME_PROFILE_DIR}")
    driver = uc.Chrome(options=options, version_main=146)
    return driver


def is_logged_in(driver):
    """ログイン状態を確認"""
    driver.get("https://jp.mercari.com")
    time.sleep(5)
    try:
        source = driver.page_source
        # ログイン済みなら「ログイン」ボタンが消えて「マイページ」系のリンクがある
        has_login_button = ">ログイン<" in source
        has_mypage = "mypage" in source.lower() or "マイページ" in source
        return has_mypage and not has_login_button
    except Exception:
        return False


# ===== ログイン =====
def do_login(driver):
    """手動ログイン（Chromeプロファイルに自動保存される）"""
    print("\n=== メルカリ手動ログイン ===")
    print("ブラウザでメルカリが開きます。")
    print("手順:")
    print("  1. ブラウザでメルカリにログインしてください")
    print("  2. マイページが表示されたことを確認")
    print("  3. このウィンドウに戻ってEnterを押す")
    print("  ※ ログイン情報はChromeプロファイルに自動保存されます\n")

    driver.get("https://jp.mercari.com")
    time.sleep(3)

    input(">>> ログイン完了後、Enterを押してください...")

    # ログイン確認
    if is_logged_in(driver):
        print("✅ ログイン成功！プロファイルに保存されました。")
        return True
    else:
        print("❌ ログインが確認できませんでした。再度お試しください。")
        return False


# ===== 検索URL管理 =====
def load_search_urls():
    """search_urls.txt から検索URLリストを読み込む。カテゴリタグも取得。
    Returns: list of (url, category) tuples
    """
    if not os.path.exists(SEARCH_URLS_FILE):
        print(f"⚠️ {SEARCH_URLS_FILE} が見つかりません。作成してください。")
        return []
    entries = []
    current_category = "tcg"  # デフォルト
    current_collab = ""
    with open(SEARCH_URLS_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("#"):
                tag = extract_category_tag(line)
                if tag:
                    current_category = tag
                collab = extract_collab_tag(line)
                if collab:
                    current_collab = collab
            elif line:
                entries.append((line, current_category, current_collab))
                current_collab = ""  # リセット
    return entries


def extract_search_label(url):
    """検索URLからラベル（表示用）を抽出"""
    parsed = urlparse(url)
    params = parse_qs(parsed.query)
    parts = []
    if "keyword" in params:
        parts.append(params["keyword"][0])
    if "category_id" in params:
        parts.append(f"cat:{params['category_id'][0]}")
    if "brand_id" in params:
        parts.append(f"brand:{params['brand_id'][0]}")
    price_min = params.get("price_min", [""])[0]
    price_max = params.get("price_max", [""])[0]
    if price_min or price_max:
        parts.append(f"¥{price_min}-{price_max}")
    return " | ".join(parts) if parts else url[:60]


# Tシャツ: 対象コラボのホワイトリスト（これ以外はスキップ）
TSHIRT_COLLAB_WHITELIST = [
    "dragon ball", "one piece", "naruto", "jujutsu kaisen",
    "demon slayer", "pokemon", "chainsaw man", "attack on titan",
    "gundam", "evangelion", "ghost in the shell", "star wars",
    "kaws", "dandadan", "spy x family", "my hero academia",
    "hunter x hunter", "oshi no ko", "kaiju no. 8", "kaiju no 8",
    "zelda", "final fantasy", "disney", "marvel",
    "berserk", "jojo", "death note", "bleach", "slam dunk",
    "doraemon", "sailor moon", "golden kamuy",
    "urusei yatsura", "lupin", "cowboy bebop", "akira",
]


def _clean_json_response(text):
    """Claude APIのJSON応答から```json```マーカーを除去"""
    if not text:
        return text
    text = text.strip()
    text = re.sub(r'^```json\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    return text.strip()


def is_tshirt_collab_target(product_info):
    """Tシャツの画像解析結果がホワイトリストのコラボかチェック"""
    if not product_info:
        return False
    try:
        cleaned = _clean_json_response(product_info)
        data = json.loads(cleaned)
        collab = data.get("collab", "").lower()
        character = data.get("character", "").lower()
        combined = collab + " " + character
        return any(wl in combined for wl in TSHIRT_COLLAB_WHITELIST)
    except json.JSONDecodeError:
        return False


def extract_category_tag(comment_line):
    """コメント行から [category:xxx] タグを抽出"""
    m = re.search(r'\[category:(\w+)\]', comment_line)
    return m.group(1) if m else "tcg"  # デフォルトはTCG


def extract_collab_tag(comment_line):
    """コメント行から [collab:xxx] タグを抽出"""
    m = re.search(r'\[collab:([^\]]+)\]', comment_line)
    return m.group(1).strip() if m else ""


def extract_product_info_from_image(api_key, image_b64, category):
    """カテゴリに応じた画像解析プロンプトで商品情報を抽出"""
    import anthropic

    prompts = {
        "tcg": "Read the PSA certification number from this PSA graded card label. Return ONLY the numeric certification number (e.g., 142490884). If no PSA label is visible, return NONE.",
        "tshirt": "This is a T-shirt listing image. Extract the following and return as JSON only: {\"brand\": \"UNIQLO/GU/other\", \"collab\": \"collaboration name (e.g. Dragon Ball, ONE PIECE)\", \"character\": \"character name if visible\", \"size\": \"size if visible (S/M/L/XL/XXL)\", \"color\": \"color\"}. If not identifiable, return {\"brand\": \"\", \"collab\": \"\", \"character\": \"\", \"size\": \"\", \"color\": \"\"}.",
        "apparel": "This is a clothing item. Extract: {\"brand\": \"brand name\", \"product\": \"product name/type\", \"model\": \"model number if visible on tag\", \"size\": \"size\", \"color\": \"color\"}. Return JSON only.",
        "ichiban": "This is a prize figure from Ichiban Kuji. Extract: {\"series\": \"series/anime name\", \"character\": \"character name\", \"prize\": \"prize tier (A/B/Last One etc)\", \"product\": \"product description\"}. Return JSON only.",
        "gashapon": "This is a capsule toy/gashapon item. Extract: {\"series\": \"series name\", \"character\": \"character name\", \"product\": \"product type\"}. Return JSON only.",
    }

    prompt = prompts.get(category, prompts["tcg"])

    try:
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=200,
            system="You extract product information from images. Return ONLY the requested format, nothing else.",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_b64}},
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        return message.content[0].text.strip()
    except Exception as e:
        print(f"       ⚠️ 画像解析エラー: {e}")
        return None


def extract_product_info_multi_image(api_key, image_b64_list, category, mercari_title="", mercari_desc=""):
    """画像+タイトル+説明文の3本柱でClaude APIに商品特定させる"""
    import anthropic

    prompts = {
        "tshirt": "Identify this T-shirt's collaboration/brand. Use the images AND the listing title/description below to determine the correct collaboration name. Return JSON only: {\"brand\": \"UNIQLO/GU/other\", \"collab\": \"collaboration name in English (e.g. Dragon Ball, ONE PIECE, Naruto)\", \"character\": \"character name if visible\", \"size\": \"size if visible (S/M/L/XL/XXL/3XL)\", \"color\": \"color in English\"}.",
        "apparel": "Identify this clothing item. Use images AND listing info. Return JSON only: {\"brand\": \"brand name\", \"product\": \"product name/type\", \"model\": \"model number if visible\", \"size\": \"size\", \"color\": \"color\"}.",
        "ichiban": "Identify this Ichiban Kuji prize. Use images AND listing info. Return JSON only: {\"series\": \"anime name in English\", \"character\": \"character name in English\", \"prize\": \"prize tier (A/B/Last One etc)\", \"product\": \"description\"}.",
        "gashapon": "Identify this capsule toy. Use images AND listing info. Return JSON only: {\"series\": \"series name in English\", \"character\": \"character name in English\", \"product\": \"product type\"}.",
    }

    prompt = prompts.get(category, prompts.get("tshirt"))

    # タイトルと説明文を追加
    context = ""
    if mercari_title:
        context += f"\n\nListing title (Japanese): {mercari_title}"
    if mercari_desc:
        context += f"\nListing description (Japanese): {mercari_desc[:300]}"

    content = []
    for b64 in image_b64_list:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}
        })
    content.append({"type": "text", "text": prompt + context})

    try:
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=200,
            system="You extract product information from images and listing text. Return ONLY valid JSON, no markdown formatting.",
            messages=[{"role": "user", "content": content}],
        )
        return message.content[0].text.strip()
    except Exception as e:
        print(f"       ⚠️ 画像解析エラー: {e}")
        return None


def check_uniqlo_official(collab_name):
    """UNIQLO公式サイトでコラボ名を検索。販売中ならTrue（スキップ対象）、なければFalse（仕入れ価値あり）。
    公式画像URLも返す。"""
    try:
        url = "https://www.uniqlo.com/jp/api/commerce/v5/ja/products"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        resp = requests.get(url, params={"q": collab_name, "offset": 0, "limit": 5}, headers=headers, timeout=15)
        if resp.status_code != 200:
            return False, None  # API失敗時はスキップしない

        items = resp.json().get("result", {}).get("items", [])

        # コラボ名が商品名に含まれているか厳密チェック
        collab_lower = collab_name.lower()
        matching = [item for item in items if collab_lower in item.get("name", "").lower()]

        if matching:
            # 公式で販売中 → 画像URLも取得
            product_id = matching[0].get("productId", "")
            # 画像URL構築
            pid_num = product_id.replace("E", "").split("-")[0] if product_id else ""
            official_image = f"https://image.uniqlo.com/UQ/ST3/jp/imagesgoods/{pid_num}/item/jpgoods_09_{pid_num}_3x4.jpg" if pid_num else None
            return True, official_image
        return False, None
    except Exception:
        return False, None


def build_ebay_query_from_product(product_info, category, mercari_title=""):
    """カテゴリ別にeBay検索クエリを生成"""
    if category == "tcg":
        # PSA番号で検索
        cert = re.search(r'(\d{8,})', product_info or "")
        if cert:
            return f"PSA 10 {cert.group(1)}", cert.group(1)
        return None, None

    elif category == "tshirt":
        try:
            cleaned = _clean_json_response(product_info)
            data = json.loads(cleaned)
            parts = []
            brand = data.get("brand", "")
            if brand:
                parts.append(brand)
            collab = data.get("collab", "")
            if collab:
                parts.append(collab)
            character = data.get("character", "")
            if character:
                parts.append(character)
            parts.append("T-Shirt")
            size = data.get("size", "")
            if size:
                parts.append(size)
            query = " ".join(parts)
            return query, json.dumps(data, ensure_ascii=False)
        except json.JSONDecodeError:
            return f"UNIQLO UT T-Shirt {mercari_title[:20]}", None

    elif category in ("apparel", "ichiban", "gashapon"):
        try:
            cleaned = _clean_json_response(product_info)
            data = json.loads(cleaned)
            parts = [v for v in data.values() if v]
            query = " ".join(parts[:4])
            return query, json.dumps(data, ensure_ascii=False)
        except json.JSONDecodeError:
            return None, None

    return None, None


# ===== スプシ在庫チェック =====
def load_inventory_urls():
    """全スプシからメルカリURLリストを取得。出品済み商品の重複除外用。"""
    all_urls = set()
    for sheet_id in INVENTORY_SHEET_IDS:
        try:
            url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0"
            resp = requests.get(url, timeout=30)
            if resp.status_code != 200:
                continue
            for line in resp.text.split("\n"):
                for match in re.findall(r'https://jp\.mercari\.com/(?:item|shops/product)/\S+', line):
                    clean = match.strip('",')
                    all_urls.add(clean)
        except Exception:
            pass
    return all_urls


# ===== GOリストCSV出力 =====
SCOUT_EXCEL_FILE = os.path.join(SCRIPT_DIR, "..", "scout_result.xlsx")


def _save_scout_excel(items):
    """スカウト全結果をExcel出力（FLG列付き）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("  ⚠️ openpyxl未インストール。Excel出力スキップ。")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "スカウト結果"

    headers = ["No.", "FLG", "判定", "理由", "タイトル", "メルカリURL",
               "メルカリ価格", "コラボ名", "eBay中央値", "eBay出品数",
               "出品価格", "利益(¥)", "利益率", "PSA番号", "公式URL"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    RED = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    status_fill = {
        "GO": GREEN, "HOLD": YELLOW, "NOGO": RED, "UNKNOWN": GRAY,
        "出品済み": GRAY, "対象外": GRAY, "公式販売中": GRAY,
    }

    for idx, item in enumerate(items):
        row = idx + 2
        status = item.get("status", "")
        fill = status_fill.get(status, GRAY)

        values = [
            idx + 1,
            "",  # FLG
            status,
            item.get("reason", ""),
            item.get("title", ""),
            item.get("mercari_url", ""),
            item.get("price_jpy", ""),
            item.get("collab", ""),
            item.get("median", ""),
            item.get("total", ""),
            item.get("ebay_price", ""),
            item.get("profit", ""),
            item.get("profit_rate", ""),
            item.get("psa_cert", ""),
            "",  # 公式URL（手動入力用）
        ]

        for j, val in enumerate(values):
            cell = ws.cell(row=row, column=j + 1, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # FLG列ドロップダウン
    dv = DataValidation(type="list", formula1='"GO,SKIP"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(items) + 2):
        dv.add(ws.cell(row=row, column=2))

    # 列幅
    widths = [5, 5, 8, 30, 40, 45, 10, 15, 10, 8, 10, 10, 8, 15, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:O{len(items) + 1}"

    wb.save(SCOUT_EXCEL_FILE)
    print(f"  📊 Excel: {SCOUT_EXCEL_FILE} ({len(items)}件)")


def save_go_list(go_items):
    """GO/保留の仕入候補をCSVに出力"""
    if not go_items:
        return
    os.makedirs(os.path.dirname(GO_LIST_FILE), exist_ok=True)
    with open(GO_LIST_FILE, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "判定", "PSA番号", "メルカリURL", "メルカリ価格",
            "eBay中央値", "推奨出品価格", "利益(¥)", "利益率", "タイトル",
        ])
        for item in go_items:
            writer.writerow([
                item["status"], item.get("psa_cert", ""),
                item["mercari_url"], item["price_jpy"],
                item.get("median", ""), item.get("ebay_price", ""),
                item.get("profit", ""), item.get("profit_rate", ""),
                item["title"],
            ])
    print(f"  📋 GOリスト: {GO_LIST_FILE} ({len(go_items)}件)")


# ===== PSA番号キャッシュ（メルカリ商品ID → PSA cert番号） =====
def _load_psa_cert_cache():
    if os.path.exists(PSA_CERT_CACHE_FILE):
        try:
            with open(PSA_CERT_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {}


def _save_psa_cert_cache(cache):
    os.makedirs(os.path.dirname(PSA_CERT_CACHE_FILE), exist_ok=True)
    with open(PSA_CERT_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ===== Claude API（PSA番号読み取り） =====
def load_anthropic_key():
    try:
        api_key_file = os.path.join(SCRIPT_DIR, "API key.txt")
        with open(api_key_file, "r", encoding="utf-8") as f:
            return f.read().strip()
    except FileNotFoundError:
        return None


def get_item_images(driver, item_url):
    """メルカリ商品詳細ページから画像URLリストを取得"""
    driver.get(item_url)
    time.sleep(5)
    source = driver.page_source
    imgs = re.findall(
        r'src="(https://static\.mercdn\.net/item/detail/orig/photos/[^"]+)"',
        source
    )
    # 重複除去
    return list(dict.fromkeys(imgs))


def get_item_description(driver):
    """現在開いている商品詳細ページから説明文を取得（get_item_imagesの後に呼ぶ）"""
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        # 「商品の説明」セクションを抽出
        desc_match = re.search(r'商品の説明\n(.*?)(?:\n商品の情報|\nカテゴリー)', body, re.DOTALL)
        if desc_match:
            return desc_match.group(1).strip()[:500]  # 最大500文字
        return ""
    except Exception:
        return ""


def download_image_via_selenium(driver, image_url):
    """Seleniumで画像をbase64として取得（403対策）"""
    try:
        b64 = driver.execute_script("""
            async function fetchImage(url) {
                const response = await fetch(url);
                const blob = await response.blob();
                return new Promise((resolve) => {
                    const reader = new FileReader();
                    reader.onloadend = () => resolve(reader.result.split(',')[1]);
                    reader.readAsDataURL(blob);
                });
            }
            return await fetchImage(arguments[0]);
        """, image_url)
        return b64
    except Exception:
        return None


def extract_psa_cert_from_image(api_key, image_b64):
    """Claude APIで画像（base64）からPSA認定番号を読み取る"""
    import anthropic
    try:

        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=100,
            system="You extract PSA certification numbers from images. Return ONLY the number, nothing else. If no PSA label is visible, return 'NONE'.",
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": image_b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": "Read the PSA certification number from this PSA graded card label. Return ONLY the numeric certification number (e.g., 142490884). If no PSA label is visible, return NONE.",
                    },
                ],
            }],
        )
        result = message.content[0].text.strip()
        # 数字だけ抽出
        cert_match = re.search(r'(\d{8,})', result)
        if cert_match:
            return cert_match.group(1)
        return None
    except Exception as e:
        print(f"       ⚠️ PSA番号読み取りエラー: {e}")
        return None


# ===== メルカリスクレイピング =====
def scrape_search_results(driver, url, max_items=30):
    """検索結果ページから商品リストを取得"""
    driver.get(url)
    time.sleep(10)

    source = driver.page_source

    # 商品データ抽出
    items = []

    # 商品URL抽出
    item_urls = re.findall(r'href="(/item/m\w+)"', source)
    item_urls = list(dict.fromkeys(item_urls))[:max_items]  # 重複除去

    # 商品名抽出（data-testid="thumbnail-item-name"）
    item_names = re.findall(
        r'data-testid="thumbnail-item-name"[^>]*>([^<]+)<',
        source
    )

    # 価格抽出（各商品ブロック内から）
    # 商品ブロックを分割して価格を取得
    blocks = re.split(r'data-testid="item-cell"', source)
    item_prices = []
    for block in blocks[1:]:  # 最初はヘッダー
        # number__XXXXX クラスのspan内の数字（カンマ区切り）
        price_match = re.search(r'class="number__\w+"[^>]*>([\d,]+)<', block)
        if not price_match:
            # fallback: ¥数字パターン
            price_match = re.search(r'[¥￥]([\d,]+)', block)
        if price_match:
            price_str = price_match.group(1).replace(",", "")
            try:
                item_prices.append(int(price_str))
            except ValueError:
                item_prices.append(0)
        else:
            item_prices.append(0)

    # 結合
    for i in range(min(len(item_urls), len(item_names), len(item_prices))):
        items.append({
            "url": f"https://jp.mercari.com{item_urls[i]}",
            "title": item_names[i].strip(),
            "price_jpy": item_prices[i],
        })

    return items


# ===== eBay API =====
def load_ebay_keys():
    keys = {}
    try:
        with open(EBAY_KEYS_FILE, "r") as f:
            for line in f:
                line = line.strip()
                if "=" in line:
                    k, v = line.split("=", 1)
                    keys[k.strip()] = v.strip()
    except FileNotFoundError:
        pass
    return keys


def get_ebay_oauth_token(app_id, app_secret):
    credentials = base64.b64encode(f"{app_id}:{app_secret}".encode()).decode()
    resp = requests.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Authorization": f"Basic {credentials}",
        },
        data={
            "grant_type": "client_credentials",
            "scope": "https://api.ebay.com/oauth/api_scope",
        },
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def search_ebay_price(token, query, limit=50):
    """eBay Browse APIで市場価格を取得"""
    url = "https://api.ebay.com/buy/browse/v1/item_summary/search"
    params = {
        "q": query,
        "filter": "buyingOptions:{FIXED_PRICE}",
        "sort": "price",
        "limit": limit,
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": "EBAY_US",
        "Content-Type": "application/json",
    }
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=15)
        if resp.status_code != 200:
            return None
        data = resp.json()
        items = data.get("itemSummaries", [])
        total = data.get("total", 0)
        if not items:
            return None

        prices = []
        for item in items:
            try:
                p = float(item.get("price", {}).get("value", 0))
                if p > 0:
                    prices.append(p)
            except (ValueError, TypeError):
                pass

        if not prices:
            return None

        s = sorted(prices)
        return {
            "median": s[len(s) // 2],
            "min": s[0],
            "max": s[-1],
            "count": len(s),
            "total": total,
        }
    except Exception as e:
        return None


# ===== GATE判定 =====
def gate_judgment(cost_jpy, market):
    """メルカリ仕入値 vs eBay市場価格でGATE判定"""
    if not market:
        return "UNKNOWN", "市場データなし", None

    median = market["median"]
    tier_profit, tier_gap_limit = get_tier_params(median)
    costs_jpy = cost_jpy + SHIPPING_JPY
    target_usd = costs_jpy / (EXCHANGE_RATE * (NET_RATIO - tier_profit))
    breakeven_usd = costs_jpy / (EXCHANGE_RATE * NET_RATIO)
    gap_pct = (target_usd - median) / median * 100 if median > 0 else 999
    gap_limit_pct = tier_gap_limit * 100

    revenue_jpy = median * EXCHANGE_RATE
    profit_jpy = revenue_jpy * NET_RATIO - costs_jpy
    profit_rate = profit_jpy / revenue_jpy if revenue_jpy > 0 else 0

    info = {
        "median": median,
        "target_usd": target_usd,
        "profit_jpy": profit_jpy,
        "profit_rate": profit_rate,
        "gap_pct": gap_pct,
        "total": market["total"],
    }

    if gap_pct <= 0:
        price = round(median * 0.95, 2)
        price = int(price) + 0.98 if price > 10 else price
        info["price"] = price
        return "GO", f"✅ GO ${price} 利益¥{profit_jpy:,.0f} ({profit_rate:.0%})", info
    elif gap_pct <= gap_limit_pct:
        price = round(target_usd, 2)
        price = int(price) + 0.98 if price > 10 else price
        info["price"] = price
        return "HOLD", f"🟡 保留 ${price} 乖離{gap_pct:.0f}%/許容{gap_limit_pct:.0f}%", info
    else:
        info["price"] = round(target_usd, 2)
        return "NOGO", f"❌ NO-GO ${target_usd:.0f} 乖離{gap_pct:.0f}% > 許容{gap_limit_pct:.0f}%", info


# ===== メイン =====
def main():
    print("=== iMak Trading Japan - メルカリスカウト ===\n")

    # ドライバー起動（Chromeプロファイル保持）
    driver = create_driver()

    # ログインモード
    if "--login" in sys.argv:
        do_login(driver)
        driver.quit()
        return

    # ログイン確認
    if is_logged_in(driver):
        print("✅ ログイン済み\n")
    else:
        print("⚠️ ログインされていません。--login で初回ログインしてください。")
        driver.quit()
        return

    # 検索URLリスト読み込み
    search_urls = load_search_urls()
    if not search_urls:
        driver.quit()
        return

    print(f"検索条件: {len(search_urls)}件\n")

    # スプシから出品済みURLを取得（重複除外用）
    inventory_urls = load_inventory_urls()
    if inventory_urls:
        print(f"✓ スプシ在庫: {len(inventory_urls)}件の既存URL読み込み済み\n")

    # eBay API準備
    ebay_keys = load_ebay_keys()
    ebay_token = None
    if ebay_keys.get("AppID") and ebay_keys.get("AppSecret"):
        try:
            ebay_token = get_ebay_oauth_token(ebay_keys["AppID"], ebay_keys["AppSecret"])
            print("✓ eBay API接続OK\n")
        except Exception as e:
            print(f"⚠️ eBay API接続失敗: {e}\n")

    # ログ蓄積用
    log_rows = []
    today = datetime.now().strftime("%Y-%m-%d")
    go_total = 0
    hold_total = 0
    go_items = []  # GOリスト蓄積
    all_scout_items = []  # 全結果（スキップ含む）Excel出力用

    # === 各検索条件を巡回 ===
    for url_idx, entry in enumerate(search_urls):
        search_url = entry[0]
        category = entry[1]
        known_collab = entry[2] if len(entry) > 2 else ""
        label = extract_search_label(search_url)
        print(f"{'═'*60}")
        print(f"  [{url_idx+1}/{len(search_urls)}] {label}")
        print(f"{'═'*60}")

        # メルカリ検索結果取得
        items = scrape_search_results(driver, search_url)
        print(f"  メルカリ: {len(items)}件取得\n")

        # Claude API準備
        anthropic_key = load_anthropic_key()
        if not anthropic_key:
            print("  ⚠️ Claude APIキーなし。PSA番号読み取りスキップ。")

        # PSA番号キャッシュ読み込み
        psa_cert_cache = _load_psa_cert_cache()

        for i, item in enumerate(items):
            title = item["title"]
            price_jpy = item["price_jpy"]
            mercari_url = item["url"]

            if price_jpy <= 0:
                continue

            # スプシ重複チェック
            if mercari_url in inventory_urls:
                print(f"  [{i+1}] ¥{price_jpy:,} {title[:50]}")
                print(f"       ⏭️ 出品済み（スキップ）")
                print()
                all_scout_items.append({
                    "title": title, "mercari_url": mercari_url, "price_jpy": price_jpy,
                    "status": "出品済み", "reason": "スプシに登録済み",
                    "psa_cert": "", "collab": "", "median": "", "total": "",
                    "ebay_price": "", "profit": "", "profit_rate": "",
                })
                continue

            print(f"  [{i+1}] ¥{price_jpy:,} {title[:50]}")
            print(f"       {mercari_url}")

            # メルカリ商品IDを抽出（キャッシュキー）
            item_id_match = re.search(r'/item/(m\w+)', mercari_url)
            item_id = item_id_match.group(1) if item_id_match else mercari_url

            # 1) 商品情報キャッシュチェック → なければ画像解析
            product_info = None
            psa_cert = None
            ebay_query = None

            if item_id in psa_cert_cache:
                cached = psa_cert_cache[item_id]
                if cached:
                    if category == "tcg":
                        psa_cert = cached
                        print(f"       🎯 PSA #{psa_cert} (キャッシュ)")
                    else:
                        product_info = cached
                        print(f"       🎯 商品情報 (キャッシュ)")
                else:
                    print(f"       ⚠️ 商品情報なし (キャッシュ)")
            elif anthropic_key:
                images = get_item_images(driver, mercari_url)
                if images:
                    img_b64 = download_image_via_selenium(driver, images[0])
                    if img_b64:
                        if category == "tcg":
                            psa_cert = extract_psa_cert_from_image(anthropic_key, img_b64)
                            if psa_cert:
                                print(f"       🎯 PSA #{psa_cert}")
                            elif len(images) > 1:
                                img_b64_2 = download_image_via_selenium(driver, images[1])
                                if img_b64_2:
                                    psa_cert = extract_psa_cert_from_image(anthropic_key, img_b64_2)
                                    if psa_cert:
                                        print(f"       🎯 PSA #{psa_cert}")
                            if not psa_cert:
                                print(f"       ⚠️ PSA番号が読み取れませんでした")
                        else:
                            # 非TCG: 画像+タイトル+説明文の3本柱
                            all_b64 = [img_b64]
                            for extra_url in images[1:3]:
                                extra_b64 = download_image_via_selenium(driver, extra_url)
                                if extra_b64:
                                    all_b64.append(extra_b64)
                            desc = get_item_description(driver)
                            product_info = extract_product_info_multi_image(
                                anthropic_key, all_b64, category,
                                mercari_title=title, mercari_desc=desc
                            )
                            if product_info:
                                print(f"       🎯 {product_info[:60]}")
                            else:
                                print(f"       ⚠️ 商品情報が読み取れませんでした")
                # キャッシュに保存
                cache_val = psa_cert if category == "tcg" else product_info
                psa_cert_cache[item_id] = cache_val
                _save_psa_cert_cache(psa_cert_cache)

            # 2) eBay検索クエリを生成してGATE判定
            market = None
            if category == "tcg":
                if psa_cert and ebay_token:
                    ebay_query = f"PSA 10 {psa_cert}"
                    market = search_ebay_price(ebay_token, ebay_query)
                    if not market:
                        card_num_match = re.search(r'((?:OP|ST|EB|PRB|FB|GD)\d+-\d+)', title)
                        if card_num_match:
                            ebay_query = f"PSA 10 {card_num_match.group(1)}"
                            market = search_ebay_price(ebay_token, ebay_query)
                    time.sleep(0.3)
            else:
                # コラボ名: 検索条件タグから既知 or 画像解析から取得
                collab_name = known_collab  # [collab:xxx] タグから
                if not collab_name and category == "tshirt" and product_info:
                    try:
                        cleaned = _clean_json_response(product_info)
                        collab_name = json.loads(cleaned).get("collab", "")
                    except (json.JSONDecodeError, Exception):
                        pass

                if category == "tshirt" and not known_collab and product_info and not is_tshirt_collab_target(product_info):
                    print(f"       ⏭️ 対象外コラボ → スキップ")
                    print()
                    all_scout_items.append({
                        "title": title, "mercari_url": mercari_url, "price_jpy": price_jpy,
                        "status": "対象外", "reason": f"コラボ「{collab_name}」がホワイトリスト外",
                        "psa_cert": "", "collab": collab_name, "median": "", "total": "",
                        "ebay_price": "", "profit": "", "profit_rate": "",
                    })
                    continue

                # Tシャツ: UNIQLO公式で販売中ならスキップ
                if category == "tshirt" and collab_name:
                    on_sale, official_img = check_uniqlo_official(collab_name)
                    if on_sale:
                        print(f"       ⏭️ UNIQLO公式で販売中 → スキップ")
                        print()
                        all_scout_items.append({
                            "title": title, "mercari_url": mercari_url, "price_jpy": price_jpy,
                            "status": "公式販売中", "reason": f"UNIQLO公式で「{collab_name}」販売中",
                            "psa_cert": "", "collab": collab_name, "median": "", "total": "",
                            "ebay_price": "", "profit": "", "profit_rate": "",
                        })
                        continue
                    else:
                        print(f"       ✅ 公式売り切れ（転売価値あり）")

                if ebay_token:
                    if known_collab and category == "tshirt":
                        # コラボ名が既知 → 直接eBay検索（画像解析不要）
                        ebay_query = f"UNIQLO UT {known_collab} T-Shirt"
                        market = search_ebay_price(ebay_token, ebay_query)
                    elif product_info or title:
                        ebay_query, extra_info = build_ebay_query_from_product(product_info, category, title)
                        if ebay_query:
                            market = search_ebay_price(ebay_token, ebay_query)
                    time.sleep(0.3)

            # 3) GATE判定
            status, gate_msg, info = gate_judgment(price_jpy, market)

            ebay_info = ""
            if info and info.get("median"):
                ebay_info = f" | eBay {info['total']}件 中央値${info['median']:.0f}"

            print(f"       {gate_msg}{ebay_info}")
            print()

            # 全結果をExcel用に蓄積
            scout_entry = {
                "title": title, "mercari_url": mercari_url, "price_jpy": price_jpy,
                "status": status, "reason": gate_msg,
                "psa_cert": psa_cert or "", "collab": collab_name if category != "tcg" else "",
                "median": info.get("median", "") if info else "",
                "total": info.get("total", "") if info else "",
                "ebay_price": info.get("price", "") if info else "",
                "profit": f"{info['profit_jpy']:,.0f}" if info and info.get("profit_jpy") else "",
                "profit_rate": f"{info['profit_rate']:.0%}" if info and info.get("profit_rate") else "",
            }
            all_scout_items.append(scout_entry)

            if status == "GO":
                go_total += 1
                go_items.append({
                    "status": "GO", "psa_cert": psa_cert, "mercari_url": mercari_url,
                    "price_jpy": price_jpy, "title": title,
                    "median": info.get("median", "") if info else "",
                    "ebay_price": info.get("price", "") if info else "",
                    "profit": f"{info['profit_jpy']:,.0f}" if info and info.get("profit_jpy") else "",
                    "profit_rate": f"{info['profit_rate']:.0%}" if info and info.get("profit_rate") else "",
                })
            elif status == "HOLD":
                hold_total += 1
                go_items.append({
                    "status": "HOLD", "psa_cert": psa_cert, "mercari_url": mercari_url,
                    "price_jpy": price_jpy, "title": title,
                    "median": info.get("median", "") if info else "",
                    "ebay_price": info.get("price", "") if info else "",
                    "profit": "", "profit_rate": "",
                })

            # ログ
            log_rows.append([
                today, label, title, price_jpy, mercari_url,
                psa_cert or "",
                info.get("median", "") if info else "",
                info.get("total", "") if info else "",
                status,
                info.get("price", "") if info else "",
            ])

        time.sleep(2)  # 検索間の待機

    driver.quit()

    # === サマリー ===
    print(f"\n{'═'*60}")
    print(f"  スカウト完了")
    print(f"{'═'*60}")
    print(f"  ✅ GO: {go_total}件  🟡 保留: {hold_total}件")

    # Excel全結果出力（FLG付き）
    if all_scout_items:
        _save_scout_excel(all_scout_items)

    # GOリストCSV出力
    if go_items:
        save_go_list(go_items)

    # psa_to_csv.py用のcerts.txt出力（GOのみ、PSA番号があるもの）
    tcg_dir = os.path.join(SCRIPT_DIR, "..", "iMakTCG")
    go_certs = [item for item in go_items if item.get("psa_cert") and item["status"] == "GO"]
    if go_certs:
        certs_path = os.path.join(tcg_dir, "certs_scout.txt")
        with open(certs_path, "w", encoding="utf-8") as f:
            for item in go_certs:
                # PSA番号,仕入値,メルカリURL,メルカリタイトル
                f.write(f"{item['psa_cert']},{item['price_jpy']},{item['mercari_url']},{item['title']}\n")
        print(f"  📝 certs_scout.txt: {certs_path} ({len(go_certs)}件)")
        print(f"     → certs.txt にコピーして psa_to_csv.py 実行 → スプシ自動追記")

    # ログ保存
    if log_rows:
        log_exists = os.path.exists(SCOUT_LOG_FILE)
        os.makedirs(os.path.dirname(SCOUT_LOG_FILE), exist_ok=True)
        with open(SCOUT_LOG_FILE, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if not log_exists:
                writer.writerow([
                    "日付", "検索条件", "タイトル", "メルカリ価格",
                    "メルカリURL", "PSA番号", "eBay中央値", "eBay出品数",
                    "判定", "推奨出品価格",
                ])
            writer.writerows(log_rows)
        print(f"  📊 ログ: {SCOUT_LOG_FILE} ({len(log_rows)}件追記)")


if __name__ == "__main__":
    main()
