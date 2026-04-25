#!/usr/bin/env python3
"""利益計算シート v2 の入力欄をクリアする。

動作:
  1. ファイルが開いていたら閉じるまで待機（5秒ごとにチェック）
  2. 閉じられたら F1/C4/E4/F4/H4 (US計算/UK計算/AU計算) をクリア
  3. 自動で再オープン
"""
import os
import sys
import time
import subprocess
from pathlib import Path

# WPS/Windows cmd の文字化け対策（スクリプト内部で処理）
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    input("Enterで終了...")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX = SCRIPT_DIR / "【NEW】利益計算シート_v2.xlsx"
TARGET_CELLS = ["F1", "C4", "E4", "F4", "H4"]
TARGET_SHEETS = ["US計算", "UK計算", "AU計算", "DE計算"]
MAX_WAIT_SECONDS = 120  # 最大2分待機


def is_locked(path):
    """ファイルが他プロセスに使用中かチェック"""
    if not path.exists():
        return False
    try:
        with open(path, "r+b"):
            return False
    except (PermissionError, OSError):
        return True


def main():
    if not XLSX.exists():
        print(f"❌ ファイルが見つかりません: {XLSX}")
        input("Enterで終了...")
        return

    # ファイルが開いている間は待機
    if is_locked(XLSX):
        print("⏳ ファイルが開かれています。WPS/Excel を閉じてください。")
        print("   （閉じたら自動で続行します。Ctrl+Cで中止）")
        waited = 0
        while is_locked(XLSX) and waited < MAX_WAIT_SECONDS:
            time.sleep(2)
            waited += 2
            if waited % 10 == 0:
                print(f"   ...待機中 {waited}秒")
        if is_locked(XLSX):
            print("❌ タイムアウト。ファイルが閉じられませんでした")
            input("Enterで終了...")
            return
        print("✅ 閉じられました、クリアを実行します")

    # クリア実行
    try:
        wb = openpyxl.load_workbook(XLSX)
    except Exception as e:
        print(f"❌ 読み込み失敗: {e}")
        input("Enterで終了...")
        return

    cleared = 0
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in TARGET_CELLS:
            ws[cell] = None
            cleared += 1

    try:
        wb.save(XLSX)
        print(f"✅ クリア完了: {cleared}セル（{len(TARGET_SHEETS)}シート × {len(TARGET_CELLS)}セル）")
    except Exception as e:
        print(f"❌ 保存失敗: {e}")
        input("Enterで終了...")
        return

    # 自動で再オープン
    try:
        os.startfile(str(XLSX))
        print(f"📂 ファイルを再度開きました")
    except Exception as e:
        print(f"⚠️ 自動オープン失敗（手動で開いてください）: {e}")

    # クイックに閉じる（3秒後自動終了）
    print("3秒後に終了します...")
    time.sleep(3)


if __name__ == "__main__":
    main()
