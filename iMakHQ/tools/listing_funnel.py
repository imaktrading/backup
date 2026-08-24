#!/usr/bin/env python3
"""
出品物フルファネル分析 (iMakHQ / 出品くんドメイン) — Seller Hub レポート版

eBay API を一切使わず、Seller Hub から DL する 4 レポートだけでファネルを組む。
(Analytics getTrafficReport の 100/日 クォータ問題を完全回避。bulk は eBay 公式も
 レポート DL を想定している = これが正規ルート)。

データ源 (--data-dir 配下、ファイル名 glob で自動検出):
  1. all-active   : eBay-all-active-listings-report-*.csv  … 全4サイト母集団 (qty/watchers/sold/price/site)
  2. quality      : Listing quality report*.xlsx            … US per-listing 深いファネル
                    (Daily impressions / CTR / Sales conversion / 適正価格 / item specifics 欠落 / 写真数 …)
  3. unsold       : eBay-unsold-listings-report-*.csv        … 売れ残り (Sold status / Relist status)
  4. orders       : *orders-report-*.csv                     … 実売 (Item Number 別に集計)

「今見る」snapshot より上位互換: impressions/CTR/転換率/適正価格 を per-listing で持つ。
これにより「views=0」の症状を「検索に出てない/クリックされない/買われない」の3原因に分解できる。

切り口 (在庫あり=qty!=0 に限定):
  - NO_SEARCH  : impressions ほぼ0 → 検索に出ていない (キーワード/カテゴリ)
  - NO_CLICK   : impr有るが CTR下位 → タイトル/サムネ/価格
  - NO_CONVERT : CTR有るが無販売   → 価格(適正価格比)/説明
  - OVERPRICED : 価格 > eBay trending price → 値付け
  (LQR 非対象=非US等は views/watchers ベースの簡易判定 DEAD_SIMPLE)

使い方:
  python listing_funnel.py                          # 既定 data-dir
  python listing_funnel.py --data-dir "C:/path/to/reports"
"""
import argparse
import csv
import datetime
import glob
import os
import re
import statistics
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

DEFAULT_DATA_DIR = r"C:\dev\iMak_data\seller_hub\reports"
FALLBACK_DATA_DIR = r"C:\Users\imax2\OneDrive\デスクトップ\新しいフォルダー (2)"
OUT_DIR = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "funnel_output"))
DESKTOP = r"C:\Users\imax2\OneDrive\デスクトップ"

# 出品管理スプシ (2枚で全カテゴリ。A列=仕入元URL=relistをまたぐ不変キー / B列=現ItemID)。
# 効果測定(funnel_diff)用に、生成時の ItemID→仕入元URL を CSV に焼き込む。relist で ItemID/title が
# 変わっても 仕入元URL は不変なので、これが世代をまたぐ商品キーになる。
SHEET_IDS = ["19kj8NqWHIGP1ptQDeGePw077hpdl6dNOO-v2J10HCjk",
             "1jF9vggbfUCddjneROMO2GGN-jTAPRbq6Qe2cbgr37B0"]
SHEET_GID = 851100680
CREDS_PATH = r"c:\dev\iMak\double-hold-421922-7c0d38d3f73d.json"


def load_supply_urls():
    """2スプシの {現ItemID(B列) → 仕入元URL(A列)}。
    失敗時は空 dict を返し supply_url 無しで継続 (funnel は診断用=非破壊なので fail-open。
    効果測定の relist 追跡は次回以降に持ち越すだけで、本体は壊さない)。"""
    out = {}
    if not os.path.isfile(CREDS_PATH):
        print("  [WARN] スプシ creds 無し → supply_url 列は空 (relist 追跡は無効)")
        return out
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                  "https://www.googleapis.com/auth/drive"]
        gc = gspread.authorize(Credentials.from_service_account_file(CREDS_PATH, scopes=scopes))
        for sid in SHEET_IDS:
            ws = gc.open_by_key(sid).get_worksheet_by_id(SHEET_GID)
            for r in ws.get_all_values()[1:]:
                url = r[0].strip() if len(r) > 0 else ""
                iid = r[1].strip() if len(r) > 1 else ""
                if iid.isdigit() and url:
                    out[iid] = url
    except Exception as e:
        print(f"  [WARN] スプシ読込失敗 → supply_url 無しで継続: {e}")
    return out

# 分類しきい値
TH_IMPR_NONE = 3        # 【fallback: PLレポート無し時】1日あたり organic impr がこれ以下 = 検索に出ていない
TH_IMPR_SHOWN = 8       # 【fallback】これ以上 = 露出あり
# 全件プロモ運用では露出の主成分は PL(広告) impressions。LQR の Daily impr は organic のみで
# PL を見落とす → NO_SEARCH 偽陽性 (実は広告で表示されてるがクリックされてない=NO_CLICK) を生む。
# PLレポート有り時は organic+PL の累計 impr で判定 (累計=CTRを判定できるサンプル数の意味)。
TH_PL_NONE = 30         # organic+PL 累計 impr がこれ以下 = ほぼ表示されてない (=真の NO_SEARCH)
TH_PL_SHOWN = 100       # 累計 impr がこれ以上 = CTR を判定できるサンプル十分
TH_AGE_MIN = 21         # 出品からこれ未満(日) = 新規 → impr低は時間不足なので NO_SEARCH 除外


def start_to_age(start):
    """active CSV の Start date ('Mar-06-25 ...') → 経過日数。不明は 0。"""
    if not start:
        return 0
    m = re.match(r"([A-Za-z]{3}-\d{2}-\d{2})", start.strip())
    if not m:
        return 0
    try:
        d = datetime.datetime.strptime(m.group(1), "%b-%d-%y").date()
        return max((datetime.date.today() - d).days, 0)
    except ValueError:
        return 0


def _f(v):
    try:
        return float(str(v).strip().replace(",", "").replace("$", ""))
    except (ValueError, AttributeError):
        return 0.0


def _i(v):
    try:
        return int(float(str(v).strip().replace(",", "")))
    except (ValueError, AttributeError):
        return 0


def archive_generation(data_dir, report_paths):
    """今回の生レポート一式を data_dir/<YYYYMMDD>/ に**コピー保管**(生データの永久資産化)。

    2026-07-24 制定: Seller Hub 生CSV/xlsx を世代ごとに貯めれば、後からどんなトレンド分析も
    遡って作れる(見られ続けない出品 / ずっと再仕入れ価値 / カテゴリ別 sell-through 推移 等)。
    派生分類を貯めるより生データ archive が完全・将来対応。日付は active レポートの内容日付
    (ファイル名 YYYY-MM-DD)。既に同名があれば skip(冪等)。
    ★2026-08-25: funnel は **日付フォルダの中も再帰で読む**ようになったので、
      「直下に loose で置く」必要は無くなった (find_file 参照)。直接 日付フォルダに
      入れる運用でも動く。戻り: (archive_dir, copied件数)。
    """
    import shutil
    active = next((p for p in report_paths if p and "active-listings" in os.path.basename(p)), None)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", os.path.basename(active or "")) if active else None
    stamp = f"{m[1]}{m[2]}{m[3]}" if m else None
    if not stamp:  # 内容日付が取れない時は archive しない(誤フォルダ名を作らない=fail-closed)
        return (None, 0)
    adir = os.path.join(data_dir, stamp)
    os.makedirs(adir, exist_ok=True)
    copied = 0
    for p in report_paths:
        if not p or not os.path.isfile(p):
            continue
        dst = os.path.join(adir, os.path.basename(p))
        if os.path.exists(dst):
            continue
        try:
            shutil.copy2(p, dst)
            copied += 1
        except Exception:
            pass
    return (adir, copied)


def _report_date_or_none(path):
    """ファイル名から「内容の日付」を読む。読めなければ None (純関数)。"""
    b = os.path.basename(path)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", b)            # YYYY-MM-DD
    if m:
        return datetime.date(int(m[1]), int(m[2]), int(m[3]))
    m = re.search(r"(\d{2})_(\d{2})_(\d{4})", b)            # MM_DD_YYYY (quality)
    if m:
        return datetime.date(int(m[3]), int(m[1]), int(m[2]))
    return None


def find_file(data_dir, pattern):
    """レポートを1本選ぶ。**日付フォルダの中も探す** (2026-08-25)。

    ★レポートは `reports/20260823/...` のように日付フォルダに入れて溜めていく運用。
      直下しか見ていなかったため **1件も見つからず**、ファネル分析が動かなかった
      (funnel CSV が 7/23 のまま古かったのはこれが原因)。今後フォルダが増えても、
      `**` で探すので何もしなくてよい。

    ★選ぶ基準は **ファイル名の日付**。mtime はフォルダに置き直すと更新され、
      古いレポートが最新に見える (この方針は _report_age_days と同じ)。
    """
    hits = glob.glob(os.path.join(data_dir, "**", pattern), recursive=True)
    if not hits:
        return None
    return max(hits, key=lambda p: (_report_date_or_none(p) or datetime.date.min,
                                    os.path.getmtime(p)))


STALE_REPORT_DAYS = 4   # 主要レポートがこれ以上古い → ファネル分析を中断 (古いと世代/効果測定が無意味)


def _report_age_days(path):
    """レポートの「内容の日付」(ファイル名から) → 経過日数。mtime は置き直しで更新され
    実態より新しく見えるので使わない (panel と同方針)。"""
    import datetime as _dt
    b = os.path.basename(path)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", b)              # YYYY-MM-DD (active/orders/unsold/promoted)
    if m:
        d = _dt.date(int(m[1]), int(m[2]), int(m[3]))
    else:
        m = re.search(r"(\d{2})_(\d{2})_(\d{4})", b)          # MM_DD_YYYY (quality)
        if m:
            d = _dt.date(int(m[3]), int(m[1]), int(m[2]))
        else:
            d = _dt.date.fromtimestamp(os.path.getmtime(path))
    return (_dt.date.today() - d).days


def worst_report_age(paths):
    """与えられたレポート群の最古経過日数 (None は無視)。"""
    ages = [_report_age_days(p) for p in paths if p]
    return max(ages) if ages else 999


def _norm_title(t):
    return re.sub(r"\s+", " ", (t or "").lower()).strip()


def build_us_title_map(active):
    """US出品の title→item_id。eBayリンクを常に US版(USD表示)にするための解決表。
    同カードは全サイト同一タイトルで出品されるため title 完全一致で US版に紐付く。"""
    return {_norm_title(a["title"]): iid for iid, a in active.items() if a.get("site") == "US"}


def us_ebay_url(row, us_by_title):
    """row の eBayリンクを US版優先で返す (US出品が無ければ自サイト)。"""
    iid = us_by_title.get(_norm_title(row.get("title", ""))) or row["item_id"]
    return f"https://www.ebay.com/itm/{iid}"


def load_active(path):
    """all-active CSV → {item_id: {...}}。全4サイト母集団。"""
    out = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            iid = (r.get("Item number") or "").strip().strip('"')
            if not iid:
                continue
            out[iid] = {
                "item_id": iid,
                "title": (r.get("Title") or "").strip(),
                "sku": (r.get("Custom label (SKU)") or "").strip(),
                "site": (r.get("Listing site") or "").strip(),
                "qty": _i(r.get("Available quantity")),
                "sold_qty": _i(r.get("Sold quantity")),
                "watch": _i(r.get("Watchers")),
                "price": _f(r.get("Current price") or r.get("Start price")),
                "category": (r.get("eBay category 1 name") or "").strip(),
                "start": (r.get("Start date") or "").strip(),
            }
    return out


def load_quality(path):
    """Listing quality report xlsx の全カテゴリシートを集約 → {item_id: {funnel...}} (US)。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    skip = {"Summary", "Guide", "Google Shopping Rejections"}
    want = {
        "Item Id": "item_id", "Daily impressions per listing": "impr",
        "Click-through rate": "ctr", "Sales conversion rate": "conv",
        "eBay trending price": "trend_price", "Number of photos": "photos",
        "Number of keywords in title": "keywords", "Sales count in last 90 days": "sales90",
        "Quantity available": "qty_q", "Item age in days": "age_days",
    }
    out = {}
    for sn in wb.sheetnames:
        if sn in skip:
            continue
        ws = wb[sn]
        hdr_row, cols = None, {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True)):
            if any(c == "Item title" for c in row if isinstance(c, str)):
                hdr_row = i + 1
                cols = {(c.strip() if isinstance(c, str) else c): j for j, c in enumerate(row) if c}
                break
        if not hdr_row or "Item Id" not in cols:
            continue
        # 'Sales count in last N days' は日数が可変 → 部分一致で拾う
        sales_col = next((j for k, j in cols.items() if isinstance(k, str) and k.startswith("Sales count")), None)
        for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row, values_only=True):
            iid = row[cols["Item Id"]] if cols["Item Id"] < len(row) else None
            if iid is None or not str(iid).strip():
                continue
            iid = str(iid).strip().strip('"')
            def g(key):
                j = cols.get(key)
                return row[j] if j is not None and j < len(row) else None
            out[iid] = {
                "impr": _f(g("Daily impressions per listing")),
                "ctr": _f(g("Click-through rate")),
                "conv_raw": g("Sales conversion rate"),
                "trend_price": _f(g("eBay trending price")),
                "photos": _f(g("Number of photos")),
                "keywords": _f(g("Number of keywords in title")),
                "sales90": _i(row[sales_col]) if sales_col is not None and sales_col < len(row) else 0,
                "category": sn,
            }
    return out


def load_unsold(path):
    """unsold CSV → {item_id: {sold_status, relist_status}}。"""
    out = {}
    if not path:
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            iid = (r.get("Item number") or "").strip().strip('"')
            if iid:
                out[iid] = {"sold_status": (r.get("Sold status") or "").strip(),
                            "relist_status": (r.get("Relist status") or "").strip()}
    return out


def load_promoted(path):
    """Promoted Listings general listing report → {item_id: (total_impr, total_clicks)}。
    total = PL(via eBay placements) + organic。全件プロモ運用では露出の主成分が PL なので、
    organic のみの LQR Daily impr を補正する。先頭に注記行があるためヘッダ行を探す。"""
    out = {}
    if not path:
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    hi = next((i for i, r in enumerate(rows) if "Item ID" in [c.strip() for c in r]), None)
    if hi is None:
        return out
    idx = {c.strip(): i for i, c in enumerate(rows[hi])}

    def g(r, c):
        j = idx.get(c)
        return r[j] if j is not None and j < len(r) else ""
    KI = "Promoted Listings Impressions (via eBay Placements)"
    KC = "Total Promoted Listings Clicks"
    for r in rows[hi + 1:]:
        iid = (g(r, "Item ID") or "").strip()
        if not iid:
            continue
        ti = _f(g(r, KI)) + _f(g(r, "Organic Impressions"))
        tc = _f(g(r, KC)) + _f(g(r, "Organic Clicks"))
        out[iid] = (ti, tc)
    return out


def classify(rows):
    """在庫(qty!=0)限定でファネル段階別に分類。LQR データ有り=詳細、無し=簡易。"""
    in_stock = [r for r in rows if r["qty"] != 0]
    oos = [r for r in rows if r["qty"] == 0]

    # 在庫切れも分析: 需要シグナルがあれば再仕入れ価値、皆無なら出品停止候補。
    # 2026-06-09: OOS品は買えない=販売もクリックも露出も抑制される(検索から隠れる)。よって
    # 「売れてない/クリック0/impr少」は不人気の証拠にならない。eBayが少しでも表示した(impr_total>0)
    # =関連性あり=再仕入れ候補(RESTOCK)。完全に表示ゼロ(impr_total==0)かつ販売/watch/90d=0 だけが
    # 真の死筋=CULL。End可否は最終的に再仕入れ可否ゲートで決める(eBay指標では死筋を確定できない)。
    def _demand(r):
        return r["sold_qty"] + r["watch"] + r.get("sales90", 0)   # 並び順用 (実需を上位に)
    # PSA10 の RESTOCK は再仕入れフロー(mercari_psa_resource)の strict 実需ゲート
    #   (実売/watch/organic impr ≥1) でしか拾われない。impr_total(広告込み)だけで RESTOCK に
    #   すると、再仕入れにも乗らず CULL にも落ちず宙ぶらりんになる(2026-06-28 実データ 211件)。
    #   → PSA は real_demand(=PSA再仕入れと同基準)で判定し、実需ゼロなら CULL(畳む)に落とす。
    #   非PSA は従来通り impr_total>0 を RESTOCK とみなす(各カテゴリの再仕入れ運用に委ねる)。
    try:
        from mercari_psa_resource import is_psa10
    except Exception:
        def is_psa10(_t): return False
    def _real_demand(r):
        return _demand(r) > 0 or r.get("impr", 0) >= 1   # organic のみ (impr_total=広告は除外)
    def _worth_restock(r):
        if is_psa10(r.get("title", "")):
            return _real_demand(r)
        return _demand(r) > 0 or r.get("impr_total", 0) > 0
    restock = [r for r in oos if _worth_restock(r)]
    cull = [r for r in oos if not _worth_restock(r)]
    restock.sort(key=lambda x: -_demand(x))

    # 露出/CTR の判定基盤を選ぶ: PLレポート有り → organic+PL 累計 impr/ctr + 累計閾値 (正しい)。
    # 無し → 旧 LQR organic-daily + 旧閾値 (後方互換)。
    use_pl = any(r.get("has_pl") for r in in_stock)
    if use_pl:
        def imp(r): return r.get("impr_total", 0.0)
        def ctr(r): return r.get("ctr_total", 0.0)
        def gate(r): return r.get("has_pl")
        th_none, th_shown = TH_PL_NONE, TH_PL_SHOWN
    else:
        def imp(r): return r["impr"]
        def ctr(r): return r["ctr"]
        def gate(r): return r.get("has_lqr")
        th_none, th_shown = TH_IMPR_NONE, TH_IMPR_SHOWN

    # CTR 下位四分位 (露出十分な listing で算出)
    ctrs = sorted([ctr(r) for r in in_stock if gate(r) and imp(r) >= th_shown])
    ctr_q1 = statistics.quantiles(ctrs, n=4)[0] if len(ctrs) >= 4 else (ctrs[0] if ctrs else 0)

    no_search, no_click, no_convert, overpriced, dead_simple, new_wait = [], [], [], [], [], []
    for r in in_stock:
        sold = r["sold_qty"] + r.get("sales90", 0)
        age = r.get("age_days", 0)
        if gate(r):
            if 0 < age < TH_AGE_MIN:
                # 出品<21日=時間不足 → 露出/CTR/転換とも判定しない (3バケツ共通の時間ガード。
                # age不明0は判定対象に残す)
                new_wait.append(r)
            elif imp(r) <= th_none:
                no_search.append(r)
            elif imp(r) >= th_shown and ctr(r) <= ctr_q1:
                no_click.append(r)
            elif sold == 0 and ctr(r) > ctr_q1:
                no_convert.append(r)
            if r["trend_price"] > 0 and r["price"] > r["trend_price"] * 1.05:
                overpriced.append(r)
        else:
            # 判定基盤なし (非US等): views 系が無いので watch/sold で簡易判定
            if sold == 0 and r["watch"] == 0:
                dead_simple.append(r)
    # 適正化: 各バケツを利益額(価格)優先で並べる → 直す価値の高い順
    for b in (no_search, no_click, no_convert):
        b.sort(key=lambda x: -x["price"])
    overpriced.sort(key=lambda x: -(x["price"] - x["trend_price"]))
    # 取下げ再出品(=新規ブースト) 候補 = NO_SEARCH + NO_CLICK 全件。
    #   relist は End+再Add で タイトル/価格/item specifics を現catalog/keyword で全項目再生成
    #   + 新規出品ブースト + 低CTR履歴の一掃 = 個別のtitle編集より包括的に効く。
    #   watcher有も含む (ユーザー判断: ブースト+全項目刷新 > 少数 watcher 保持)。
    relist = sorted(no_search + no_click, key=lambda x: -x["price"])
    return {"NO_SEARCH": no_search, "NO_CLICK": no_click, "NO_CONVERT": no_convert,
            "OVERPRICED": overpriced, "DEAD_SIMPLE": dead_simple, "NEW_WAIT": new_wait,
            "RELIST": relist, "OUT_OF_STOCK": oos, "RESTOCK": restock, "CULL": cull, "ctr_q1": ctr_q1}


# バケツ → 既にボタン化済の対策 (出品くん control_panel のボタン名)。空=対策ボタン無し
# (OVERPRICED/NEW_WAIT/DEAD_SIMPLE は意図的に専用ボタン無し)。Summary の E列に表示 (2026-06-30)。
BUCKET_REMEDY_BUTTON = {
    "NO_SEARCH":  "取下再出品①②③ (End→Add→書戻し)",
    "NO_CLICK":   "✏️ タイトル改修 / 取下再出品①②③",
    "NO_CONVERT": "💲 価格抵抗",
    "RELIST":     "取下再出品①②③ (End→Add→書戻し)",
    "RESTOCK":    "🛒 在庫切れ再仕入れ (PSA/一番くじは専用補充)",
    "CULL":       "🗑 取下げ (200件/回・月末までに)",   # 2026-08-23: 50→200 / 名前もパネルと一致させた
}


def write_xlsx(path, rows, c, summary_lines):
    """バケツ別シートの Excel を出力 (Summary + 各バケツ + eBayリンク)。"""
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "出品物フルファネル分析 (Seller Hub レポート版)"
    ws["A1"].font = Font(bold=True, size=13)
    for i, line in enumerate(summary_lines, start=3):
        ws.cell(row=i, column=1, value=line)
    buckets = [
        ("NO_SEARCH", "検索に出ていない (キーワード/カテゴリ)"),
        ("NO_CLICK", "クリックされない (サムネ/タイトル/価格)"),
        ("NO_CONVERT", "買われない (価格/説明)"),
        ("OVERPRICED", "適正価格より高い (値下げ余地)"),
        ("NEW_WAIT", "新規出品でimpr低 (時間不足=様子見・改修対象外)"),
        ("RELIST", "取下げ再出品候補 (在庫あり・検索露出ゼロ・watcher無)"),
        ("RESTOCK", "在庫切れだが需要実証済 (再仕入れ優先)"),
        ("CULL", "在庫切れ&需要皆無 (出品停止候補)"),
        ("DEAD_SIMPLE", "非US等・LQR無 (簡易判定)"),
    ]
    r0 = 3 + len(summary_lines) + 1
    ws.cell(row=r0, column=1, value="バケツ").font = Font(bold=True)
    ws.cell(row=r0, column=2, value="件数").font = Font(bold=True)
    ws.cell(row=r0, column=3, value="意味/アクション").font = Font(bold=True)
    ws.cell(row=r0, column=5, value="対策ボタン (済)").font = Font(bold=True)
    for k, (name, desc) in enumerate(buckets, start=r0 + 1):
        ws.cell(row=k, column=1, value=name)
        ws.cell(row=k, column=2, value=len(c.get(name, [])))
        ws.cell(row=k, column=3, value=desc)
        ws.cell(row=k, column=5, value=BUCKET_REMEDY_BUTTON.get(name, ""))  # E列: 対策済ボタン名
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["E"].width = 36

    cols = [("item_id", 13), ("title", 46), ("site", 6), ("category", 20), ("price", 8),
            ("trend_price", 10), ("qty", 5), ("sold_qty", 6), ("sales90", 8), ("watch", 7),
            ("impr", 7), ("ctr%", 7), ("photos", 7), ("keywords", 9), ("relist_status", 14), ("ebay_url", 32)]
    def write_sheet(sh_name, items):
        if not items:
            return
        sh = wb.create_sheet(sh_name[:31])
        for j, (h, w) in enumerate(cols, start=1):
            sh.cell(row=1, column=j, value=h).font = Font(bold=True)
            sh.column_dimensions[sh.cell(row=1, column=j).column_letter].width = w
        for i, r in enumerate(items, start=2):
            vals = [r["item_id"], r["title"], r["site"], r.get("category", ""), r["price"],
                    r["trend_price"], r["qty"], r["sold_qty"], r.get("sales90", 0), r["watch"],
                    round(r["impr"], 1), round(r["ctr"] * 100, 2), r.get("photos", 0), r.get("keywords", 0),
                    r.get("relist_status", ""), r.get("ebay_url") or f"https://www.ebay.com/itm/{r['item_id']}"]
            for j, v in enumerate(vals, start=1):
                sh.cell(row=i, column=j, value=v)

    # 在庫あり全件 / 在庫なし全件 (Summary の次に配置 = 一覧の起点)
    if rows:
        write_sheet("在庫あり", sorted([r for r in rows if r["qty"] != 0], key=lambda x: (-x["impr"], -x["watch"])))
        write_sheet("在庫なし", sorted([r for r in rows if r["qty"] == 0], key=lambda x: -(x["sold_qty"] + x["watch"])))
    for name, _ in buckets:
        write_sheet(name, c.get(name, []))
    wb.save(path)


# 「ファネル分析」スプシ (2026-06-07 集約: デスクトップxlsx置換)。eBayアップCSV以外はスプシに集約。
FUNNEL_SHEET_ID = "1UkaI4W6YCJgUbjgF7LLNN9_fHeVuz5qB4r9RqImElwg"
# (flag, 意味/アクション, 要件) — 要件は classify() の実判定条件を転記 (2026-06-09 追加)
FUNNEL_BUCKETS = [
    ("NO_SEARCH", "検索に出ていない (キーワード/カテゴリ)", "在庫>0 ∩ 露出ほぼ0 (impr≤無閾値)"),
    ("NO_CLICK", "クリックされない (サムネ/タイトル/価格)", "在庫>0 ∩ 露出十分 ∩ CTR下位25%以下"),
    ("NO_CONVERT", "買われない (価格/説明)", "在庫>0 ∩ CTR下位超 ∩ 販売0(sold+90d)"),
    ("OVERPRICED", "適正価格より高い (値下げ余地)", "在庫>0 ∩ 価格>適正価格×1.05"),
    ("NEW_WAIT", "新規出品でimpr低 (時間不足=様子見)", "在庫>0 ∩ 出品 0<age<21日"),
    ("RELIST", "取下げ再出品候補 (露出ゼロ)", "NO_SEARCH ∪ NO_CLICK"),
    ("RESTOCK", "在庫切れだが需要実証済 (再仕入れ)", "在庫=0 ∩ 需要>0 (生涯販売+watch+90d)"),
    ("CULL", "在庫切れ&需要皆無 (出品停止候補)", "在庫=0 ∩ 需要=0 (+End時 age≥21)"),
    ("DEAD_SIMPLE", "非US等・LQR無 (簡易判定)", "判定基盤無 ∩ 販売0 ∩ watch0"),
]
FUNNEL_COLS = ["item_id", "title", "site", "category", "price", "trend_price", "qty",
               "sold_qty", "sales90", "watch", "impr", "ctr%", "impr_total", "ctr_total%",
               "photos", "keywords", "relist_status", "ebay_url"]


def _funnel_vals(r):
    return [r["item_id"], r["title"], r["site"], r.get("category", ""), r["price"],
            r["trend_price"], r["qty"], r["sold_qty"], r.get("sales90", 0), r["watch"],
            round(r["impr"], 1), round(r["ctr"] * 100, 2),
            round(r.get("impr_total", 0), 1), round(r.get("ctr_total", 0) * 100, 2),
            r.get("photos", 0), r.get("keywords", 0), r.get("relist_status", ""),
            r.get("ebay_url") or f"https://www.ebay.com/itm/{r['item_id']}"]


def write_funnel_to_sheet(rows, c, summary_lines):
    """ファネル全結果を「ファネル分析」スプシのタブに書く (Summary + 在庫あり/なし + 全9バケツ)。

    デスクトップ xlsx の置換。失敗しても funnel_*.csv は別途出力済なのでデータは失われない。
    戻り: 書込タブ数 (失敗時は例外)。
    """
    import gspread
    from google.oauth2.service_account import Credentials
    creds = Credentials.from_service_account_file(
        CREDS_PATH, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(FUNNEL_SHEET_ID)

    def write_tab(name, data2d):
        try:
            ws = sh.worksheet(name)
            ws.clear()
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=name, rows=max(10, len(data2d) + 5),
                                  cols=max(8, len(data2d[0]) if data2d else 8))
        ws.update(range_name="A1", values=data2d, value_input_option="RAW")

    # Summary タブ
    summ = [["出品物フルファネル分析 (Seller Hub版)"]]
    summ += [[ln] for ln in summary_lines]
    summ += [[""], ["バケツ", "件数", "意味/アクション", "要件"]]
    for name, desc, req in FUNNEL_BUCKETS:
        summ.append([name, len(c.get(name, [])), desc, req])
    write_tab("Summary", summ)

    # 在庫あり / 在庫なし
    instock = sorted([r for r in rows if r["qty"] != 0], key=lambda x: (-x["impr"], -x["watch"]))
    oos = sorted([r for r in rows if r["qty"] == 0], key=lambda x: -(x["sold_qty"] + x["watch"]))
    write_tab("在庫あり", [FUNNEL_COLS] + [_funnel_vals(r) for r in instock])
    write_tab("在庫なし", [FUNNEL_COLS] + [_funnel_vals(r) for r in oos])

    # 9 バケツ (空でもヘッダのみ書いてタブを安定させる)
    for name, _, _ in FUNNEL_BUCKETS:
        write_tab(name, [FUNNEL_COLS] + [_funnel_vals(r) for r in c.get(name, [])])

    # 既定の空タブ「シート1」を掃除
    try:
        sh.del_worksheet(sh.worksheet("シート1"))
    except Exception:
        pass
    return 3 + len(FUNNEL_BUCKETS)  # Summary+在庫あり+在庫なし + 9


def _sec(title, note, items, limit=20):
    print(f"\n=== {title} ({len(items)}件) ===\n   {note}")
    if not items:
        print("   (該当なし)")
        return
    print(f"   {'item_id':<13}{'site':>4}{'impr/d':>7}{'CTR%':>6}{'sold':>5}{'$':>7}{'trend':>7}  title")
    for r in items[:limit]:
        print(f"   {r['item_id']:<13}{r['site']:>4}{r['impr']:>7.0f}{r['ctr']*100:>6.2f}"
              f"{r['sold_qty']+r.get('sales90',0):>5}{r['price']:>7.0f}{r['trend_price']:>7.0f}  {r['title'][:36]}")
    if len(items) > limit:
        print(f"   ... 他 {len(items) - limit} 件 (CSV 参照)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", default=None, help="4レポートを置いたフォルダ")
    ap.add_argument("--no-csv", action="store_true")
    ap.add_argument("--force", action="store_true",
                    help="レポートが古くても強制実行 (既定: 古いと中断)")
    args = ap.parse_args()

    # ★2026-08-25: ここも直下しか見ていなかったので、レポートが日付フォルダに入った
    #   時点で「既定の置き場は空」と判断し、存在しないデスクトップの旧フォルダに
    #   フォールバックして止まっていた。find_file と同じく再帰で見る。
    _has_reports = os.path.isdir(DEFAULT_DATA_DIR) and glob.glob(
        os.path.join(DEFAULT_DATA_DIR, "**", "*active*"), recursive=True)
    data_dir = args.data_dir or (DEFAULT_DATA_DIR if _has_reports else FALLBACK_DATA_DIR)
    if not os.path.isdir(data_dir):
        sys.exit(f"data-dir が見つかりません: {data_dir}\n--data-dir で4レポートのフォルダを指定してください。")

    f_active = find_file(data_dir, "*all-active-listings*.csv")
    f_quality = find_file(data_dir, "Listing quality report*.xlsx")
    f_unsold = find_file(data_dir, "*unsold-listings*.csv")
    f_promoted = find_file(data_dir, "*promoted-listing*report*.csv")
    f_orders = find_file(data_dir, "*orders-report*.csv")
    if not f_active:
        sys.exit(f"all-active CSV が {data_dir} に見つかりません。")

    # レポート鮮度ガード: 古いまま走らせると funnel世代/効果測定が無意味になる → 中断 (2026-06-07)
    # ★2026-08-23: **LQR を鮮度判定から外す**。Listing quality report は eBay が週次でしか
    #   作らないので、他4本を今日落としても必ず数日古く、毎回この gate に引っかかる
    #   (実際 8/23 に5本そろえても「最古4日前」で中断し、--force が要った)。
    #   毎回 --force を打たせると gate の意味が無くなるので、**日次で取れる4本だけ**で見る。
    #   LQR 自体の古さは下の表示で分かる。
    _worst = worst_report_age([f_active, f_unsold, f_promoted, f_orders])
    _lqr_age = worst_report_age([f_quality]) if f_quality else None
    if _worst >= STALE_REPORT_DAYS and not args.force:
        sys.exit(
            f"⛔ 中断: レポートが古いです (最古 {_worst}日前 ≥ {STALE_REPORT_DAYS}日)。\n"
            f"   古いレポートで走らせると funnel世代が更新されず効果測定も無意味になります。\n"
            f"   → Seller Hub で4-5レポートを再DL → {data_dir} に置き直してから再実行。\n"
            f"   (どうしても古いまま実行する場合は --force)")
    # 生レポートを世代フォルダに永久アーカイブ(以後 定期実行で自動蓄積=分析を後から遡れる)
    try:
        _adir, _ncopy = archive_generation(data_dir, [f_active, f_quality, f_unsold, f_promoted, f_orders])
        if _adir:
            print(f"🗄 生レポート archive: {os.path.basename(_adir)}/ に {_ncopy}件保管(世代蓄積=後から分析可)")
    except Exception as _ae:
        print(f"  ⚠ archive skip(非致命): {type(_ae).__name__}: {_ae}")
    print(f"data-dir: {data_dir}  (レポート最古 {_worst}日前)")
    print(f"  active  : {os.path.basename(f_active)}")
    print(f"  quality : {os.path.basename(f_quality) if f_quality else '(なし=簡易判定)'}"
          + (f"  ← {_lqr_age}日前 (eBayが週次でしか作らないので数日古いのが普通)"
             if _lqr_age else ""))
    print(f"  unsold  : {os.path.basename(f_unsold) if f_unsold else '(なし)'}")
    print(f"  promoted: {os.path.basename(f_promoted) if f_promoted else '(なし=organic限定の旧判定)'}")

    active = load_active(f_active)
    quality = load_quality(f_quality) if f_quality else {}
    unsold = load_unsold(f_unsold)
    promoted = load_promoted(f_promoted)
    us_by_title = build_us_title_map(active)  # eBayリンクを常に US版(USD)に解決
    supply = load_supply_urls()               # ItemID→仕入元URL (効果測定の世代またぎキー)
    print(f"  仕入元URL: スプシから {len(supply)} 件 (supply_url 列に焼き込み=relist追跡キー)")

    rows = []
    for iid, a in active.items():
        q = quality.get(iid)
        r = dict(a)
        r["ebay_url"] = us_ebay_url(a, us_by_title)
        r["has_lqr"] = bool(q)
        r["impr"] = q["impr"] if q else 0.0          # organic 日次 (LQR・CSV/demand_winners 互換用に温存)
        r["ctr"] = q["ctr"] if q else 0.0
        pl = promoted.get(iid)
        r["has_pl"] = bool(pl)
        r["impr_total"] = pl[0] if pl else 0.0        # organic+PL 累計 (分類の正しい露出指標)
        r["ctr_total"] = (pl[1] / pl[0]) if pl and pl[0] else 0.0
        r["trend_price"] = q["trend_price"] if q else 0.0
        r["sales90"] = q["sales90"] if q else 0
        r["age_days"] = start_to_age(a.get("start", "")) or (q.get("age_days", 0) if q else 0)
        r["photos"] = q["photos"] if q else 0
        r["keywords"] = q["keywords"] if q else 0
        u = unsold.get(iid, {})
        r["relist_status"] = u.get("relist_status", "")
        r["supply_url"] = supply.get(iid, "")     # relist で ItemID/title が変わっても不変の商品キー
        rows.append(r)

    # ★2026-08-25 ユーザー指示: **US だけを分析対象にする**。
    #   UK / AU / CA は eBaymag が US の親出品から作るミラーで、こちらから End も Revise も
    #   できない (グローバル CLAUDE.md「eBaymag のミラー — 直接 触るな」)。
    #   混ぜると、手を出せない行がバケツを埋めて件数が意味を持たなくなる。
    #   実測 2026-08-25: DEAD_SIMPLE 2,055件のうち **2,043件がミラー** で、
    #   CULL 1,518件のうち 867件、OUT_OF_STOCK 1,993件のうち 887件がミラーだった。
    from collections import Counter
    site_all = Counter(r["site"] for r in rows)
    _mirror = [r for r in rows if (r.get("site") or "").upper() != "US"]
    rows = [r for r in rows if (r.get("site") or "").upper() == "US"]
    if _mirror:
        print(f"  🪞 eBaymag のミラーを除外: {len(_mirror)}件 "
              f"({dict(Counter(r['site'] for r in _mirror))}) "
              f"— 親の US を操作すれば付いてくるので、こちらからは触らない")
    site_c = Counter(r["site"] for r in rows)
    oos = sum(1 for r in rows if r["qty"] == 0)
    in_stock = [r for r in rows if r["qty"] != 0]
    lqr_n = sum(1 for r in rows if r["has_lqr"])
    n = max(len(rows), 1)
    summary_lines = [
        f"対象レポート: {os.path.basename(f_active)} 他",
        f"分析対象は **US のみ** {len(rows)}件 "
        f"(eBaymag のミラー {len(_mirror)}件 を除外: "
        f"{dict((k, v) for k, v in site_all.items() if k != 'US')})",
        "  ※ミラーは親の US 出品から自動で作られるもので、こちらから取り下げも修正もできない。"
        "混ぜると手を出せない行がバケツを埋めて件数が意味を持たなくなる",
        f"在庫切れqty0={oos}件({oos*100//n}%)  在庫あり={len(in_stock)}件  LQR深掘り対象(US)={lqr_n}件",
    ]
    print(f"\n出品物フルファネル分析 (Seller Hub レポート版・API不使用)")
    for ln in summary_lines[1:]:
        print("   " + ln)

    c = classify(rows)
    _pl_on = any(r.get("has_pl") for r in rows)
    _none, _shown = (TH_PL_NONE, TH_PL_SHOWN) if _pl_on else (TH_IMPR_NONE, TH_IMPR_SHOWN)
    _basis = "organic+PL累計" if _pl_on else "organic日次"
    print(f"   (適正化) 露出基盤={_basis} (PLレポート{'有' if _pl_on else '無'}) / 新規<{TH_AGE_MIN}日は NEW_WAIT 隔離={len(c['NEW_WAIT'])}件 / 各バケツ価格順")
    _sec("① 検索に出ていない NO_SEARCH", f"在庫あり & 出品>={TH_AGE_MIN}日 & 露出({_basis})<={_none} → 真の検索不可。キーワード or 取下げ再出品 (価格高い順)", c["NO_SEARCH"])
    _sec("② クリックされない NO_CLICK", f"在庫あり & 出品>={TH_AGE_MIN}日 & 露出({_basis})>={_shown} & CTR下位25%({c['ctr_q1']*100:.2f}%) → タイトル/サムネ/価格", c["NO_CLICK"])
    print(f"   ▷ 取下げ再出品(=新規ブースト・全項目再生成)候補 = NO_SEARCH+NO_CLICK = {len(c['RELIST'])}件")
    _sec("③ 買われない NO_CONVERT", f"在庫あり & 出品>={TH_AGE_MIN}日 & CTR有 & 無販売 → 価格(適正価格比)/説明", c["NO_CONVERT"])
    _sec("④ 高すぎ OVERPRICED", "在庫あり & 価格 > eBay適正価格×1.05 → 値下げ余地 (差額大きい順)", c["OVERPRICED"])

    def _sec_oos(title, note, items, limit=20):
        print(f"\n=== {title} ({len(items)}件) ===\n   {note}")
        if not items:
            print("   (該当なし)"); return
        print(f"   {'item_id':<13}{'site':>4}{'sold':>5}{'watch':>6}{'$':>7}  title")
        for r in items[:limit]:
            print(f"   {r['item_id']:<13}{r['site']:>4}{r['sold_qty']+r.get('sales90',0):>5}{r['watch']:>6}{r['price']:>7.0f}  {r['title'][:38]}")
        if len(items) > limit:
            print(f"   ... 他 {len(items) - limit} 件 (CSV 参照)")

    print("\n--- 在庫切れ(qty0)の分析 ---")
    _sec_oos("⑤ 再仕入れ優先 RESTOCK", "在庫切れ & 過去販売 or watcher 有 → 需要実証済、仕入れ先を再確保 (需要大きい順)", c["RESTOCK"])
    print(f"\n(参考) CULL(在庫切れ&需要皆無=出品停止候補): {len(c['CULL'])}件 — 一度も売れず watcher も付かず。整理対象")
    print(f"(参考) DEAD_SIMPLE(非US等・LQR無): {len(c['DEAD_SIMPLE'])}件")

    if not args.no_csv:
        os.makedirs(OUT_DIR, exist_ok=True)
        path = os.path.join(OUT_DIR, f"funnel_{datetime.date.today():%Y%m%d}.csv")
        for r in rows:
            tags = []
            if r["qty"] == 0: tags.append("OUT_OF_STOCK")
            for k in ("NO_SEARCH", "NO_CLICK", "NO_CONVERT", "OVERPRICED", "DEAD_SIMPLE", "NEW_WAIT", "RELIST", "RESTOCK", "CULL"):
                if r in c[k]: tags.append(k)
            r["flags"] = "|".join(tags)
        fields = ["item_id", "title", "site", "category", "price", "trend_price", "qty", "sold_qty",
                  "sales90", "watch", "impr", "ctr", "impr_total", "ctr_total", "photos", "keywords",
                  "has_lqr", "relist_status", "age_days", "supply_url", "flags", "ebay_url"]
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            for r in sorted(rows, key=lambda x: (-x["impr"], -x["watch"])):
                w.writerow(r)
        print(f"CSV 出力: {path}")

    # ファネル全結果を「ファネル分析」スプシに集約 (Summary+在庫あり/なし+全9バケツ)。
    # デスクトップ xlsx は廃止 (2026-06-07 集約方針=eBayアップCSV以外はスプシ)。
    try:
        n = write_funnel_to_sheet(rows, c, summary_lines)
        print(f"📊 「ファネル分析」スプシ更新: {n}タブ "
              f"https://docs.google.com/spreadsheets/d/{FUNNEL_SHEET_ID}/edit")
    except Exception as _e:  # noqa: BLE001
        print(f"⚠ 「ファネル分析」スプシ更新失敗 (funnel_*.csv は出力済): {type(_e).__name__}: {_e}")

    # 「既存メンテ」スプシ 行動系タブ (取下再出品/需要・新規強化) を更新 (非致命)
    try:
        import existing_maint_dashboard as emd
        emd.main()
    except SystemExit:
        pass
    except Exception as _e:  # noqa: BLE001
        print(f"⚠ 「既存メンテ」スプシ更新スキップ: {type(_e).__name__}: {_e}")


if __name__ == "__main__":
    main()
