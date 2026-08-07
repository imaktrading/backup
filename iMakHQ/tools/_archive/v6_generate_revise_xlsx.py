# -*- coding: utf-8 -*-
"""V6 revise XLSX 生成 (前回 V5 xlsx と同じ項目 + V6 計算列追加)
- eBay 全カラム + V6_Category, Cost¥, Cost_Source, V6_Group, V6_Tier, V6_Policy,
  V6_Price$, V6_Shipping$, V6_BuyerTotal$, V6_Profit¥, V6_Rate%, V6_Status, Remap★, Pack
"""
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

CSV_PATH = Path(r"C:\Users\imax2\OneDrive\デスクトップ\eBay-all-active-listings-report-2026-05-20-13302285935.csv")
COSTS_JSON = Path(r"c:\tmp\v6_item_costs.json")
OUTPUT = Path(r"C:\Users\imax2\OneDrive\デスクトップ\active_with_v6_20260520.xlsx")

# パック商品
PACK_ITEMS = {
    "356796294045": 3, "358280566084": 3, "358280633511": 3,
    "358388897815": 3, "358390003755": 5, "358390109249": 5,
}

# B→A remap
B_TO_A_REMAP = {
    "DDP-B-P03": "DDP-A-P05", "DDP-B-P06": "DDP-A-P10", "DDP-B-P11": "DDP-A-P15",
    "DDP-B-P14": "DDP-A-P20", "DDP-B-P17": "DDP-A-P22", "DDP-B-P20": "DDP-A-P24",
    "DDP-B-P26": "DDP-A-P30", "DDP-B-P29": "DDP-A-P31",
}

# グループ別 設定 (yaml と同期予定)
GROUP_HTS = {"A": 0.18, "B": 0.30, "C": 0.43}  # +5% バッファ込み
GROUP_SPLIT = {"A": 1.0, "B": 1.0, "C": 0.5}    # C は案B (DDP 50% 送料外出し)
SHIPPING_JPY = 2000   # 国内送料 (= グループ共通)
FX_USD = 158.88
FVF_DEFAULT = 0.155
PROMO_RATE = 0.10
PAYO_RATE = 0.025
INSERTION_USD = 0.4
CLEARANCE_JPY = 245


def detect_group(title: str, ebay_cat: str) -> str:
    t = (title or "").lower()
    cat = (ebay_cat or "").lower()
    if any(k in t for k in ["porter", "anello", "アネロ"]):
        return "B"
    if any(k in cat for k in ["handbag", "bag"]):
        return "B"
    if "reel" in t or "リール" in t or "fishing" in cat:
        return "B"
    if any(k in t for k in ["uniqlo", "ut t-shirt", "ut tshirt"]):
        return "C"
    if "montbell" in t:
        return "C"
    if any(k in cat for k in ["t-shirt", "tshirt", "coat", "jacket", "vest", "shirts"]):
        return "C"
    return "A"


def price_tier_from_listing_price(price_usd: float) -> tuple:
    bins = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100,
            120, 140, 160, 180, 200, 220, 240, 260, 280, 300,
            350, 400, 450, 500, 550, 600,
            700, 800, 900, 1000, 1500]
    for i, upper in enumerate(bins, 1):
        if price_usd <= upper:
            return f"P{i:02d}", upper
    return "P31", 1500


def profit_rate_ifs(cost_jpy: float) -> float:
    """仕入¥ → 目標利益率 (V5 と同じ)"""
    tiers = [(2000,0.60),(3500,0.50),(6000,0.40),(10000,0.32),(15000,0.27),
             (20000,0.25),(25000,0.22),(30000,0.20),(35000,0.18),(40000,0.17),
             (45000,0.15),(50000,0.14),(55000,0.13),(60000,0.12)]
    for max_cost, rate in tiers:
        if cost_jpy < max_cost:
            return rate
    return 0.09


def compute_v6(cost_jpy: int, group: str):
    """V6 paid shipping 計算 → dict"""
    hts_rate = GROUP_HTS[group]
    split = GROUP_SPLIT[group]
    fvf = FVF_DEFAULT
    mr = fvf + PROMO_RATE + PAYO_RATE
    m = 1 + hts_rate * 1.021
    profit_rate = profit_rate_ifs(cost_jpy)
    P = cost_jpy * profit_rate
    H = cost_jpy
    J = SHIPPING_JPY
    insertion_jpy = INSERTION_USD * FX_USD

    # 案A (A/B): C×FX = (H + J + insertion + 245×mr + P) × m / (1 - mr × m)
    # G_buyer (= F + D) × FX = C × FX × m + 245
    denom = 1 - mr * m
    if denom <= 0:
        return None
    C_jpy = (H + J + insertion_jpy + CLEARANCE_JPY * mr + P) * m / denom
    G_buyer_jpy = C_jpy * m + CLEARANCE_JPY  # buyer 総額 ¥
    # 商品本体 USD
    F_usd = int(C_jpy / FX_USD) + 0.98  # listing price (案A)
    # 全 DDP USD
    DDP_total_usd = F_usd * hts_rate * 1.021 + CLEARANCE_JPY / FX_USD

    if split == 1.0:  # 案A
        listing_price_usd = F_usd
        shipping_usd = DDP_total_usd
    else:  # 案B (C グループ): DDP 半分を商品価格に
        listing_price_usd = round(F_usd + DDP_total_usd * (1 - split), 2)
        shipping_usd = round(DDP_total_usd * split, 2)
    buyer_total_usd = listing_price_usd + shipping_usd

    # 実利益再計算
    G_jpy = buyer_total_usd * FX_USD
    K_jpy = G_jpy * fvf + insertion_jpy
    L_jpy = G_jpy * PROMO_RATE
    M_jpy = G_jpy * PAYO_RATE
    N_jpy = shipping_usd * FX_USD + (listing_price_usd - F_usd) * FX_USD  # paid + 案B 上乗せ
    O_jpy = H + J + K_jpy + L_jpy + M_jpy + N_jpy
    profit_jpy = G_jpy - O_jpy
    rate_pct = profit_jpy / G_jpy * 100 if G_jpy > 0 else 0

    return {
        "listing_price_usd": round(listing_price_usd, 2),
        "shipping_usd": round(shipping_usd, 2),
        "buyer_total_usd": round(buyer_total_usd, 2),
        "profit_jpy": round(profit_jpy, 0),
        "rate_pct": round(rate_pct, 1),
    }


def main():
    # 仕入¥ JSON 読込
    costs_data = json.loads(COSTS_JSON.read_text(encoding="utf-8-sig"))
    costs_map = costs_data["costs"]
    source_map = costs_data["source"]

    # eBay CSV 読込 (US のみ、ItemID で集約)
    rows_by_id = {}
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames
        for r in reader:
            if r.get("Listing site", "").strip() != "US":
                continue
            itemid = r.get("Item number", "").strip().strip('"')
            if not itemid or itemid in rows_by_id:
                continue
            rows_by_id[itemid] = r

    # XLSX 作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active+V6"

    # ヘッダ: eBay original + V6 追加列
    v6_cols = ["V6_Category", "Cost¥", "Cost_Source", "Pack",
               "V6_Group", "V6_Tier", "V6_Policy", "Remap★",
               "V6_Price$", "V6_Shipping$", "V6_BuyerTotal$",
               "V6_Profit¥", "V6_Rate%", "V6_Status"]
    header = list(fields) + v6_cols
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    skipped_no_cost = 0
    for itemid, r in rows_by_id.items():
        title = r.get("Title", "")
        ebay_cat = r.get("eBay category 1 name", "")
        sku = r.get("Custom label (SKU)", "")
        try:
            cur_price = float(r.get("Current price", "") or r.get("Start price", "") or 0)
        except ValueError:
            cur_price = 0

        # 仕入¥
        cost_jpy = costs_map.get(itemid, 0)
        cost_src = source_map.get(itemid, "")
        pack_qty = PACK_ITEMS.get(itemid, 1)
        if cost_jpy > 0 and pack_qty > 1:
            cost_jpy = cost_jpy * pack_qty

        # カテゴリ判定 (= グループ + 商品カテゴリ名)
        group = detect_group(title, ebay_cat)
        v6_category = {
            "A": "ホビー系 (TCG/玩具/G-SHOCK)",
            "B": "バッグ/Porter/リール",
            "C": "アパレル (Tシャツ/Montbell)",
        }[group]

        # V6 計算
        if cost_jpy > 0:
            calc = compute_v6(cost_jpy, group)
        else:
            calc = None
            skipped_no_cost += 1

        if calc:
            v6_price = calc["listing_price_usd"]
            v6_ship = calc["shipping_usd"]
            v6_total = calc["buyer_total_usd"]
            v6_profit = calc["profit_jpy"]
            v6_rate = calc["rate_pct"]
            # Status 判定
            if v6_profit < 0:
                v6_status = "赤字"
            elif v6_rate < 5:
                v6_status = "薄利"
            elif v6_rate < 15:
                v6_status = "正常"
            else:
                v6_status = "厚利"
            # Tier / Policy (listing price ベース)
            tier, upper = price_tier_from_listing_price(v6_price)
        else:
            v6_price = v6_ship = v6_total = v6_profit = v6_rate = ""
            v6_status = "Cost不明" if cost_jpy == 0 else "ERR"
            tier, upper = price_tier_from_listing_price(cur_price)

        policy_raw = f"DDP-{group}-{tier}"
        policy = B_TO_A_REMAP.get(policy_raw, policy_raw)
        remapped = "★" if policy != policy_raw else ""
        pack_mark = f"×{pack_qty}" if pack_qty > 1 else ""

        # 行データ
        out = [r.get(f, "") for f in fields]  # eBay 元カラム
        out += [
            v6_category,
            cost_jpy if cost_jpy > 0 else "",
            cost_src,
            pack_mark,
            group,
            tier,
            policy,
            remapped,
            v6_price,
            v6_ship,
            v6_total,
            v6_profit,
            v6_rate,
            v6_status,
        ]
        ws.append(out)

    wb.save(OUTPUT)
    print(f"Output: {OUTPUT}")
    print(f"Total listings: {len(rows_by_id)}")
    print(f"Cost unknown (skipped): {skipped_no_cost}")


if __name__ == "__main__":
    main()
