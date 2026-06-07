# -*- coding: utf-8 -*-
"""93 Policy 一覧を Excel で出力
- A グループ (低関税): 案 A (全 DDP 送料)
- B グループ (中関税): 案 A (全 DDP 送料)
- C グループ (高関税): 案 B (DDP 50% 送料 + 50% 商品)
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# yaml SSOT 参照 (V8 FIX 2026-05-24: split 直書き廃止)
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "iMakeBayAPI"))
import config_loader

# 段階ピッチ価格帯 (上限基準)
bins = [
    # ($10 刻み, 0-100)
    (10, "<$10"), (20, "$10-20"), (30, "$20-30"), (40, "$30-40"), (50, "$40-50"),
    (60, "$50-60"), (70, "$60-70"), (80, "$70-80"), (90, "$80-90"), (100, "$90-100"),
    # ($20 刻み, 100-300)
    (120, "$100-120"), (140, "$120-140"), (160, "$140-160"), (180, "$160-180"), (200, "$180-200"),
    (220, "$200-220"), (240, "$220-240"), (260, "$240-260"), (280, "$260-280"), (300, "$280-300"),
    # ($50 刻み, 300-600)
    (350, "$300-350"), (400, "$350-400"), (450, "$400-450"), (500, "$450-500"),
    (550, "$500-550"), (600, "$550-600"),
    # ($100 刻み, 600-1000)
    (700, "$600-700"), (800, "$700-800"), (900, "$800-900"), (1000, "$900-1000"),
    # ($500 刻み)
    (1500, "$1000-1500"),
]

# yaml から split / rate を取得 (= SSOT 統一、 直書き廃止)
_v6 = config_loader.get_v6_pricing()
_yaml_groups = _v6.get("groups", {})

def _design_label(split: float) -> str:
    if split >= 1.0:
        return "案A (全DDP送料)"
    return f"案B (DDP{int(split*100)}%送料+{int((1-split)*100)}%商品)"

_cat_by_group = {}
for cat, gid in _v6.get("category_to_group", {}).items():
    _cat_by_group.setdefault(gid, []).append(cat)

groups = {
    gid: {
        "rate": float(g["hts_rate"]),
        "split": float(g["split"]),
        "name": {"A": "低関税", "B": "中関税", "C": "高関税"}.get(gid, gid),
        "design": _design_label(float(g["split"])),
        "categories": " / ".join(_cat_by_group.get(gid, [])),
    }
    for gid, g in _yaml_groups.items()
}

wb = openpyxl.Workbook()
ws_overview = wb.active
ws_overview.title = "概要"

# 概要シート
ws_overview.append(["V6 DDP対応 eBay Shipping Policy 一覧"])
ws_overview.append(["生成日", "2026-05-20"])
ws_overview.append(["バッファ", "+5% (グループ別 hts_rate に組込)"])
ws_overview.append(["通関固定", "$1.5 (= ¥245)"])
ws_overview.append(["関税処理", "× 1.021"])
ws_overview.append([])
ws_overview.append(["グループ", "対象カテゴリ", "hts_rate (+バッファ)", "設計案", "送料計算"])
for gid, g in groups.items():
    formula = f"F × {g['rate']:.2f} × 1.021"
    if g['split'] < 1.0:
        formula += f" × {g['split']:.1f} (DDP 50% 送料)"
    formula += " + $1.5"
    ws_overview.append([f"グループ {gid}", g['categories'], g['rate'], g['design'], formula])

# ヘッダ装飾
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
for cell in ws_overview[1]:
    cell.font = Font(bold=True, size=14)
for cell in ws_overview[7]:
    cell.font = header_font
    cell.fill = header_fill

# 列幅
ws_overview.column_dimensions["A"].width = 18
ws_overview.column_dimensions["B"].width = 50
ws_overview.column_dimensions["C"].width = 20
ws_overview.column_dimensions["D"].width = 30
ws_overview.column_dimensions["E"].width = 45

# 各グループのシート
for gid, g in groups.items():
    ws = wb.create_sheet(f"グループ{gid}_{g['name']}")
    ws.append([f"DDP-{gid}-XX グループ {gid} ({g['name']})"])
    ws.append(["対象", g['categories']])
    ws.append(["hts_rate", f"{g['rate']:.2f} (バッファ +5% 込み)"])
    ws.append(["設計", g['design']])
    ws.append(["送料計算", f"F × {g['rate']} × 1.021 × {g['split']} + $1.5"])
    ws.append([])
    ws.append(["Policy ID", "価格帯", "上限$ (計算基準)", "送料$", "送料率% (上限価格基準)"])

    for i, (upper, label) in enumerate(bins, 1):
        cost = upper * g['rate'] * 1.021 * g['split'] + 1.5
        rate_pct = (cost / upper * 100) if upper > 0 else 0
        ws.append([
            f"DDP-{gid}-P{i:02d}",
            label,
            upper,
            round(cost, 2),
            round(rate_pct, 1),
        ])

    # 装飾
    ws[1][0].font = Font(bold=True, size=14)
    for cell in ws[7]:
        cell.font = header_font
        cell.fill = header_fill
    # 列幅
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20

# 全 Policy 集約シート
ws_all = wb.create_sheet("全Policy_93個")
ws_all.append(["Policy ID", "グループ", "価格帯", "上限$", "rate", "送料$", "送料率% (上限基準)"])
for gid, g in groups.items():
    for i, (upper, label) in enumerate(bins, 1):
        cost = upper * g['rate'] * 1.021 * g['split'] + 1.5
        rate_pct = (cost / upper * 100) if upper > 0 else 0
        ws_all.append([
            f"DDP-{gid}-P{i:02d}",
            gid,
            label,
            upper,
            g['rate'],
            round(cost, 2),
            round(rate_pct, 1),
        ])
for cell in ws_all[1]:
    cell.font = header_font
    cell.fill = header_fill
ws_all.column_dimensions["A"].width = 16
ws_all.column_dimensions["C"].width = 16

# 出力
output = Path(r"C:/Users/imax2/OneDrive/デスクトップ/V6_Policy一覧_93個.xlsx")
wb.save(output)
print(f"Saved: {output}")
print(f"Total Policies: {len(bins) * len(groups)}")
