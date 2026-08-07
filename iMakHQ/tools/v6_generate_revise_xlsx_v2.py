# -*- coding: utf-8 -*-
"""V6 revise XLSX 生成 v2 (V5 xlsx の Cost を流用)
V5 xlsx の AF=Cost¥, AG=Cost_Source を ItemID で引いて V6 xlsx 生成
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

V5_XLSX = Path(r"C:\Users\imax2\OneDrive\デスクトップ\active_with_v5_20260519_130235.xlsx")
CSV_PATH = Path(r"C:\Users\imax2\OneDrive\デスクトップ\eBay-all-active-listings-report-2026-05-20-13302285935.csv")
COSTS_JSON = Path(r"c:\tmp\v6_item_costs.json")  # 4 ソース fallback
OUTPUT = Path(r"C:\Users\imax2\OneDrive\デスクトップ\active_with_v6_20260520.xlsx")

# パック商品 (= 仕入¥ × N)
PACK_ITEMS = {
    "356796294045": 3, "358280566084": 3, "358280633511": 3,
    "358388897815": 3, "358390003755": 5, "358390109249": 5,
}

# B→A remap
B_TO_A_REMAP = {
    "DDP-B-P03": "DDP-A-P05", "DDP-B-P06": "DDP-A-P10", "DDP-B-P11": "DDP-A-P15",
    "DDP-B-P14": "DDP-A-P20", "DDP-B-P17": "DDP-A-P22", "DDP-B-P20": "DDP-A-P24",
    "DDP-B-P26": "DDP-A-P30", "DDP-B-P29": "DDP-A-P31",
}

# V6 計算パラメータ (V5 v4_GS 設定 シートと完全一致 SSOT)
GROUP_HTS = {"A": 0.18, "B": 0.30, "C": 0.43}
GROUP_SPLIT = {"A": 1.0, "B": 1.0, "C": 0.5}
FX_USD = 159.005

# 国別 (US) — v4_GS 設定 r36
US_VAT = 0.085
US_INTL = 0.0165
US_REG = 0

# G-SHOCK 国別 override (v4_GS 設定 r48): US は 0.155
GSHOCK_FVF_US = 0.155

# カテゴリ別 FVF + 国内送料 (v4_GS 設定 r11-r28 SSOT)
CATEGORY_SETTINGS = {
    "TCG(PSA10)":         (0.133, 2000),
    "G-SHOCK":            (0.138, 2000),  # US override 0.155 で fvf_effective() 内で吸収
    "Tシャツ(UT)":         (0.14,  2000),
    "Montbell(軽)":        (0.14,  2000),
    "Montbell(重)":        (0.14,  3000),
    "一番くじ":            (0.14,  3000),
    "フィギュア":          (0.14,  3000),
    "ユニクロ(非UT)":       (0.14,  2000),
    "サンリオ文具":         (0.145, 2000),
    "ヴィンテージ玩具":      (0.14,  2000),
    "トミカ":              (0.14,  2000),
    "POPMart":            (0.14,  2000),
    "ガシャポン":           (0.14,  2000),
    "ダイソー":            (0.14,  2000),
    "バッグ(アネロ)":       (0.15,  3000),
    "サンリオぬいぐるみ":    (0.14,  2500),
    "Porter":             (0.15,  3500),
    "リール":              (0.14,  3000),
}
DEFAULT_FVF = 0.14
DEFAULT_SHIPPING_JPY = 2000

# HIGHT/LOW R列カテゴリ名 → v5_GS HTS_RATE カテゴリ名 への正規化
CATEGORY_NORMALIZE = {
    "TCG":             "TCG(PSA10)",
    "G-shock":         "G-SHOCK",
    "G-SHOCK":         "G-SHOCK",
    "Tシャツ":          "Tシャツ(UT)",
    "Tシャツ(UT)":      "Tシャツ(UT)",
    "ユニクロ":         "ユニクロ(非UT)",
    "ユニクロ(非UT)":    "ユニクロ(非UT)",
    "Montbell":        "Montbell(軽)",
    "Montbell(軽)":     "Montbell(軽)",
    "Montbell(重)":     "Montbell(重)",
    "アウトドア・ジャケット": "Montbell(重)",  # HIGHT/LOW R列
    "一番くじ":         "一番くじ",
    "フィギュア":        "フィギュア",
    "サンリオ文具":      "サンリオ文具",
    "ヴィンテージ玩具":   "ヴィンテージ玩具",
    "ヴィンテージ":      "ヴィンテージ玩具",  # HIGHT/LOW R列
    "トミカ":           "トミカ",
    "tomica":          "トミカ",  # HIGHT/LOW R列
    "POPMart":         "POPMart",
    "ガシャポン":        "ガシャポン",
    "カプセルトイ":      "ガシャポン",  # HIGHT/LOW R列
    "ダイソー":         "ダイソー",
    "バッグ":           "バッグ(アネロ)",  # title に PORTER あれば Porter に override (detect_group 内)
    "バッグ(アネロ)":    "バッグ(アネロ)",
    "アネロ":           "バッグ(アネロ)",
    "サンリオぬいぐるみ": "サンリオぬいぐるみ",
    "Porter":          "Porter",
    "リール":           "リール",
    # HIGHT/LOW R列のみ、v5_GS に未対応 (defaults: A group 0.14/2000)
    # "グッズ", "スニーカー", "ゴルフ", "その他" は title-based 後段判定
}
PROMO_RATE = 0.10
PAYO_RATE = 0.025
INSERTION_USD = 0.4
CLEARANCE_JPY = 245


def normalize_category(category_label: str) -> str:
    """HIGHT/LOW R列 略称 → v4_GS 設定 正式名"""
    cat = (category_label or "").strip()
    if not cat:
        return ""
    if cat in CATEGORY_NORMALIZE:
        return CATEGORY_NORMALIZE[cat]
    # 部分一致 fallback
    for short, full in CATEGORY_NORMALIZE.items():
        if short in cat or cat in short:
            return full
    return cat


def category_settings(category_label: str) -> tuple:
    """v4_GS 設定 r11-r28 から (fvf_base, 国内送料_JPY) を取得"""
    norm = normalize_category(category_label)
    if norm in CATEGORY_SETTINGS:
        return CATEGORY_SETTINGS[norm]
    return (DEFAULT_FVF, DEFAULT_SHIPPING_JPY)


def fvf_effective(category_label: str) -> float:
    """v4_GS US計算 G3 式: (FVF + Intl + Reg) × (1 + VAT). G-SHOCK のみ US override"""
    norm = normalize_category(category_label)
    if norm == "G-SHOCK":
        fvf_base = GSHOCK_FVF_US
    else:
        fvf_base, _ = category_settings(category_label)
    return (fvf_base + US_INTL + US_REG) * (1 + US_VAT)


def shipping_jpy(category_label: str) -> int:
    _, ship = category_settings(category_label)
    return ship


def detect_group(title: str, ebay_cat: str, v5_category: str = "", hl_category: str = "") -> tuple:
    """グループ判定: title pattern > HIGHT/LOW カテゴリ > V5_Category > ebay_cat
    Returns: (group, normalized_category_label)
    """
    t = (title or "").lower()
    hl = (hl_category or "").strip()
    # === Title-based override (R列がない or 曖昧な場合) ===
    # Porter: title に PORTER/YOSHIDA があれば Porter (バッグでも)
    if "porter" in t or "yoshida" in t:
        return "B", "Porter"
    # サンリオぬいぐるみ: title に plush/ぬいぐるみ + (kuromi/hello kitty/sanrio etc.)
    if ("plush" in t or "ぬいぐるみ" in t or "mascot" in t) and any(
        k in t for k in ["kuromi", "sanrio", "hello kitty", "cinnamoroll", "my melody", "pompompurin", "pochacco"]
    ):
        return "A", "サンリオぬいぐるみ"
    # アネロ: title に anello (Porter で先に return 済)
    if "anello" in t or "アネロ" in t:
        return "B", "バッグ(アネロ)"
    # アウトドア・ジャケット: title に montbell + jacket/coat 系
    if "montbell" in t and any(k in t for k in ["jacket", "coat", "parka", "wind", "rain"]):
        return "C", "Montbell(重)"

    if hl:
        # まず normalize_category で正式名取得試行
        norm = normalize_category(hl)
        if norm in CATEGORY_SETTINGS:
            # Group 判定 (apparel C / bags B / その他 A)
            if norm in ("Tシャツ(UT)", "Montbell(軽)", "Montbell(重)", "ユニクロ(非UT)"):
                return "C", norm
            if norm in ("バッグ(アネロ)", "Porter", "リール"):
                return "B", norm
            return "A", norm
        # HL は v5_GS HTS_RATE に未対応 (グッズ等) → 既定 A
        return "A", hl

    # V5_Category
    cat5 = (v5_category or "").strip()
    if cat5 in ("Tシャツ(UT)", "Montbell(軽)", "Montbell(重)", "ユニクロ(非UT)"):
        return "C", cat5
    if cat5 in ("バッグ(アネロ)", "Porter", "リール"):
        return "B", cat5
    if cat5:
        return "A", cat5

    # fallback: title / ebay_cat
    t = (title or "").lower()
    cat = (ebay_cat or "").lower()
    if any(k in t for k in ["porter", "anello", "アネロ"]) or "handbag" in cat or "bag" in cat:
        return "B", "bag (推測)"
    if "reel" in t or "リール" in t or "fishing" in cat:
        return "B", "reel (推測)"
    if any(k in t for k in ["uniqlo", "ut t-shirt", "ut tshirt"]) or "montbell" in t:
        return "C", "apparel (推測)"
    if any(k in cat for k in ["t-shirt", "tshirt", "coat", "jacket", "vest", "shirts"]):
        return "C", "apparel (推測)"
    return "A", "hobby (推測)"


PRICE_BINS = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100,
              120, 140, 160, 180, 200, 220, 240, 260, 280, 300,
              350, 400, 450, 500, 550, 600,
              700, 800, 900, 1000, 1500]


def price_tier_from_listing_price(price_usd: float) -> tuple:
    for i, upper in enumerate(PRICE_BINS, 1):
        if price_usd <= upper:
            return f"P{i:02d}", upper
    return "P31", 1500


def policy_shipping_usd(group: str, tier_upper: int) -> float:
    """Policy 送料 (= 価格帯上限ベース) = upper × hts_rate × 1.021 × split + 1.5"""
    hts_rate = GROUP_HTS[group]
    split = GROUP_SPLIT[group]
    return round(tier_upper * hts_rate * 1.021 * split + 1.5, 2)


def profit_rate_ifs(cost_jpy: float) -> float:
    tiers = [(2000,0.60),(3500,0.50),(6000,0.40),(10000,0.32),(15000,0.27),
             (20000,0.25),(25000,0.22),(30000,0.20),(35000,0.18),(40000,0.17),
             (45000,0.15),(50000,0.14),(55000,0.13),(60000,0.12)]
    for max_cost, rate in tiers:
        if cost_jpy < max_cost:
            return rate
    return 0.09


def compute_v6(cost_jpy: int, group: str, category_label: str = ""):
    """V5 v4_GS US計算 G9/N9 式を完全再現 (per-category 国内送料 + FVF SSOT)"""
    hts_rate = GROUP_HTS[group]
    split = GROUP_SPLIT[group]
    fvf_eff = fvf_effective(category_label)  # = G3 (= VAT 込み)
    mr = fvf_eff + PROMO_RATE + PAYO_RATE
    m = 1 + hts_rate * 1.021
    profit_rate = profit_rate_ifs(cost_jpy)
    P = cost_jpy * profit_rate
    H = cost_jpy
    J = shipping_jpy(category_label)  # per-category 国内送料 (v4_GS 設定 SSOT)
    insertion_jpy = INSERTION_USD * FX_USD
    denom = 1 - mr * m
    if denom <= 0:
        return None
    # G9 = (H-I+J+0.4×FX + 245×mr + P) × m / (1 - mr×m) + 245
    G_buyer_jpy = (H + J + insertion_jpy + CLEARANCE_JPY * mr + P) * m / denom + CLEARANCE_JPY
    # C9 = (G9 - 245) / m / FX = 商品本体 ¥売価 / FX = 商品本体 USD
    C_usd = (G_buyer_jpy - CLEARANCE_JPY) / m / FX_USD
    # V5 sheet F9 と一致: INT は最終結果に 1 回のみ (C は unrounded 利用)
    DDP_from_C = C_usd * hts_rate * 1.021 + CLEARANCE_JPY / FX_USD  # case A 換算 DDP USD
    if split == 1.0:
        listing_price_usd = int(C_usd) + 0.98
        # 案A の D_usd は F_usd ベース (= V5 D9 旧式と一致)
        F_usd = listing_price_usd
        D_usd = F_usd * hts_rate * 1.021 + CLEARANCE_JPY / FX_USD
        shipping_usd = round(D_usd, 2)
    else:  # C グループ 案B: V5 F9 = INT(C + DDP_from_C×(1-split)) + 0.98
        listing_price_usd = int(C_usd + DDP_from_C * (1 - split)) + 0.98
        D_usd = DDP_from_C  # V5 と同じ unrounded C ベース
        shipping_usd = round(D_usd * split, 2)
    buyer_total_usd = round(listing_price_usd + shipping_usd, 2)

    G_jpy = buyer_total_usd * FX_USD
    K_jpy = G_jpy * fvf_eff + insertion_jpy
    L_jpy = G_jpy * PROMO_RATE
    M_jpy = G_jpy * PAYO_RATE
    N_jpy = D_usd * FX_USD  # 実 DDP コスト ¥
    O_jpy = H + J + K_jpy + L_jpy + M_jpy + N_jpy
    profit_jpy = G_jpy - O_jpy
    rate_pct = profit_jpy / G_jpy * 100 if G_jpy > 0 else 0

    # V5 スプシ P9/Q9 と完全一致:
    #   P9 = 仕入 × IFS 利益率 (target)
    #   Q9 = P9 / G9 (= G_buyer_jpy theoretical equation)
    profit_target_jpy = cost_jpy * profit_rate
    rate_v5_q9 = profit_target_jpy / G_buyer_jpy * 100 if G_buyer_jpy > 0 else 0
    return {
        "listing_price_usd": round(listing_price_usd, 2),
        "shipping_usd": shipping_usd,
        "buyer_total_usd": buyer_total_usd,
        "profit_jpy": round(profit_target_jpy, 0),
        "rate_pct": round(rate_v5_q9, 1),
        "g_equation_jpy": round(G_buyer_jpy, 0),
    }


def main():
    # === V5 xlsx を読込、ItemID → (Cost, Source, V5_Category) マップ ===
    v5_wb = openpyxl.load_workbook(V5_XLSX, data_only=True)
    v5_ws = v5_wb.active
    v5_map = {}
    for r in range(2, v5_ws.max_row + 1):
        itemid = str(v5_ws.cell(r, 1).value or "").strip()
        if not itemid:
            continue
        cost = v5_ws.cell(r, 32).value  # AF
        source = v5_ws.cell(r, 33).value  # AG
        v5_cat = v5_ws.cell(r, 31).value  # AE
        v5_map[itemid] = {"cost": cost, "source": source, "v5_cat": v5_cat}
    print(f'V5 xlsx loaded: {len(v5_map)} items')

    # === 4 ソース fallback ===
    import json
    if COSTS_JSON.exists():
        cd = json.loads(COSTS_JSON.read_text(encoding="utf-8-sig"))
        fallback_costs = cd["costs"]
        fallback_src = cd["source"]
        hl_categories = cd.get("categories", {})
        print(f'Fallback costs (4 ソース): {len(fallback_costs)} items')
        print(f'HIGHT/LOW categories: {len(hl_categories)} items')
    else:
        fallback_costs = {}
        fallback_src = {}
        hl_categories = {}

    # === eBay CSV 読込 (US only, ItemID 集約) ===
    rows_by_id = {}
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames
        for r in reader:
            if r.get("Listing site", "").strip() != "US":
                continue
            itemid = r.get("Item number", "").strip().strip('"')
            if not itemid or itemid in rows_by_id:
                continue
            rows_by_id[itemid] = r
    print(f'CSV: {len(rows_by_id)} US listings')

    # === V6 xlsx 生成 ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active+V6"

    v6_cols = ["V6_Category", "Cost¥", "Cost_Source", "Pack",
               "V6_Group", "V6_Tier", "V6_Policy", "Remap★",
               "V6_Price$", "V6_Shipping$", "V6_BuyerTotal$",
               "V6_Profit¥", "V6_Rate%", "V6_Status"]
    header = list(fields) + v6_cols
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    cost_unknown = 0
    new_listings = 0
    for itemid, r in rows_by_id.items():
        title = r.get("Title", "")
        ebay_cat = r.get("eBay category 1 name", "")
        sku = r.get("Custom label (SKU)", "")
        try:
            cur_price = float(r.get("Current price", "") or r.get("Start price", "") or 0)
        except ValueError:
            cur_price = 0

        v5 = v5_map.get(itemid, {})
        cost_raw = v5.get("cost", "")
        try:
            cost_jpy = int(cost_raw) if cost_raw and str(cost_raw).strip() else 0
        except (ValueError, TypeError):
            cost_jpy = 0
        cost_src = v5.get("source", "") or ""
        v5_cat = v5.get("v5_cat", "") or ""

        # V5 になければ 4 ソース fallback
        if cost_jpy == 0 and itemid in fallback_costs:
            cost_jpy = int(fallback_costs[itemid])
            cost_src = fallback_src.get(itemid, "fallback")

        if itemid not in v5_map:
            new_listings += 1

        # パック × N
        pack_qty = PACK_ITEMS.get(itemid, 1)
        if cost_jpy > 0 and pack_qty > 1:
            cost_jpy = cost_jpy * pack_qty

        # グループ判定 (HIGHT/LOW カテゴリ優先 → V5_Category → fallback)
        hl_cat = hl_categories.get(itemid, "")
        group, v6_category = detect_group(title, ebay_cat, v5_cat, hl_cat)

        # V6 計算
        if cost_jpy > 0:
            calc = compute_v6(cost_jpy, group, v6_category)
        else:
            calc = None
            cost_unknown += 1

        if calc:
            v6_price = calc["listing_price_usd"]
            tier, upper = price_tier_from_listing_price(v6_price)
            v6_ship = policy_shipping_usd(group, upper)  # eBay Policy 送料
            v6_total = round(v6_price + v6_ship, 2)
            # 利益/利益率: V5 スプシ P9/Q9 と一致 (= 仕入 × IFS / G_equation)
            v6_profit = calc["profit_jpy"]
            v6_rate = calc["rate_pct"]
            if v6_profit < 0:
                v6_status = "赤字"
            elif v6_rate < 5:
                v6_status = "薄利"
            elif v6_rate < 15:
                v6_status = "正常"
            else:
                v6_status = "厚利"
        else:
            v6_price = v6_ship = v6_total = v6_profit = v6_rate = ""
            v6_status = "Cost不明"
            tier, upper = price_tier_from_listing_price(cur_price)

        policy_raw = f"DDP-{group}-{tier}"
        policy = B_TO_A_REMAP.get(policy_raw, policy_raw)
        remapped = "★" if policy != policy_raw else ""
        pack_mark = f"×{pack_qty}" if pack_qty > 1 else ""

        out = [r.get(f, "") for f in fields]
        out += [
            v6_category,
            cost_jpy if cost_jpy > 0 else "",
            cost_src,
            pack_mark,
            group,
            tier,
            policy,
            remapped,
            v6_price,
            v6_ship,
            v6_total,
            v6_profit,
            v6_rate,
            v6_status,
        ]
        ws.append(out)

    wb.save(OUTPUT)
    print(f'\nOutput: {OUTPUT}')
    print(f'Total: {len(rows_by_id)}')
    print(f'Cost 不明: {cost_unknown}')
    print(f'New listings (V5 にない): {new_listings}')


if __name__ == "__main__":
    main()
