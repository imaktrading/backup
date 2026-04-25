"""
利益計算シート_v2.xlsx を Google Sheets に移行する。
- 為替セル(B2/F2/H2/J2) を GOOGLEFINANCE に差し替え
- _xlfn. プレフィックス除去
- 全シートをコピー
- imax2303@gmail.com に編集権限で共有
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

EXCEL_PATH = Path(r"c:\dev\iMak\iMakHQ\sheets\【NEW】利益計算シート_v2.xlsx")
CREDS_PATH = Path(r"c:\dev\iMak\double-hold-421922-7c0d38d3f73d.json")
TARGET_URL = "https://docs.google.com/spreadsheets/d/1ft91iIsJjbMVw3Gx4GmeO-DQ0A47jp6O1TbiZeTslag/edit"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# 為替セルの差し替えマップ (シート名は移行後に同じ)
RATE_OVERRIDES = {
    "設定": {
        "B2": '=GOOGLEFINANCE("CURRENCY:USDJPY")',
        "F2": '=GOOGLEFINANCE("CURRENCY:EURJPY")',
        "H2": '=GOOGLEFINANCE("CURRENCY:GBPJPY")',
        "J2": '=GOOGLEFINANCE("CURRENCY:AUDJPY")',
    }
}


def clean_formula(value):
    """openpyxl が返す式の _xlfn. プレフィックスを除去"""
    if not isinstance(value, str):
        return value
    if value.startswith("="):
        # _xlfn.IFS → IFS, _xlfn.FILTER → FILTER 等
        return re.sub(r"_xlfn\.", "", value)
    return value


def excel_to_gsheet():
    print(f"Loading Excel: {EXCEL_PATH.name}")
    wb = load_workbook(EXCEL_PATH, data_only=False)

    print(f"Authenticating with {CREDS_PATH.name}")
    creds = Credentials.from_service_account_file(str(CREDS_PATH), scopes=SCOPES)
    gc = gspread.authorize(creds)

    print(f"Opening target: {TARGET_URL}")
    target_sh = gc.open_by_url(TARGET_URL)
    print(f"  Title: {target_sh.title}")

    # 既存のワークシート(Sheet1等)を最後まで保持、最後に削除
    existing_wss = list(target_sh.worksheets())

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        print(f"  Migrating: {sheet_name} ({max_row}x{max_col})")

        rows = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                v = clean_formula(v)
                coord = ws.cell(row=r, column=c).coordinate
                if sheet_name in RATE_OVERRIDES and coord in RATE_OVERRIDES[sheet_name]:
                    v = RATE_OVERRIDES[sheet_name][coord]
                row_vals.append(v if v is not None else "")
            rows.append(row_vals)

        # 同名タブが既にあれば削除してから作り直し
        for ew in existing_wss:
            if ew.title == sheet_name:
                target_sh.del_worksheet(ew)
                existing_wss.remove(ew)
                break

        new_ws = target_sh.add_worksheet(title=sheet_name, rows=max(max_row, 100), cols=max(max_col, 26))
        if rows:
            new_ws.update(values=rows, range_name="A1", value_input_option="USER_ENTERED")

    # 残っている既存タブ(Sheet1など移行対象外のもの)は削除
    for ew in list(target_sh.worksheets()):
        if ew.title not in wb.sheetnames:
            try:
                target_sh.del_worksheet(ew)
                print(f"  Deleted leftover: {ew.title}")
            except Exception as e:
                print(f"  Warn: could not delete {ew.title}: {e}")

    print(f"\n完了: {target_sh.url}")
    return target_sh.url


if __name__ == "__main__":
    try:
        url = excel_to_gsheet()
        print(f"\n>>> {url}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
