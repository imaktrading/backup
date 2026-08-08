# -*- coding: utf-8 -*-
"""invoice_duty_rate — Orange Connex 請求書から実効関税率を出す。

背景 (2026-07-31 → 2026-08-08 全面改訂):
    SpeedPAK の関税は「事前設定レート」で、**どの料金表にも率が書かれていない**
    (RATE_GUIDE L430-437 に「最終的な決済は請求金額に基づく」と明記)。
    HTS を引いても実請求は分からないため、**請求書が唯一の一次情報**。

★2026-08-08: 当初の「HSコードが違っても同率 (9.95%/9.98%)」という結論は **読み間違い**だった。
    パーサが `引落金額` 行 (= **推定**関税) だけを読み、後日の `料金調整` 行で入る
    **実額**を取り落としていた。実額で読み直すと:

        TCG    HS 9504.40 (MFN 0%)  → 実効  9.98%
        Tシャツ HS 6109.90 (MFN 32%) → 実効 41.82%   ← 推定は ¥937 = 9.9% だった

    = **HTS本則 + 約10%**。HS ごとに率はまったく違う。
    旧テスト `test_hs_code_does_not_change_the_rate` はこのバグを「正しい仕様」として
    固定してしまっていたので、**逆向きの固定に差し替えた** (下の
    `test_hs_code_does_change_the_rate`)。

守りたい性質:
  1. 数値パースが通貨記号・カンマに耐える (請求書は "1,289" 形式)
  2. 費目は **見出し名**で引く (列番号は export ごとに入れ替わる)
  3. `料金調整` を合算した **実額**を関税とする。推定のままなら status="estimate"
  4. シート末尾の **総合計行**を個別 tid に合算しない
  5. 請求書の書式が変わったら **黙って0件にせず例外**にする (silent drop 禁止)
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "iMakeBayAPI"))
import invoice_duty_rate as idr  # noqa: E402

REAL = Path(r"C:\dev\iMak_data\shipping\invoices")

FEE_COLS = ["運送料金", "燃料割増金", "推定関税及び税金料金",
            "輸入通関手数料", "推定関税処理手数料", "関税", "関税処理手数料"]


def _build(path, fee_cols, fee_rows, skus, orders):
    """合成請求書を組む。fee_cols の並びを変えて『列入れ替わり』を再現できる."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = idr.SH_ORDER
    ws.append(["作成時間", "OrangeConnex追跡番号", "ご請求金額（JPY）", "実支払額（JPY）",
               "割引（JPY）", "還元金額（JPY）", "輸送業者名", "追跡番号", "配送サービス",
               "目的の国", "サービスタイプ", "インコタームズ"])
    for o in orders:
        ws.append([o["date"], o["tid"], "0", "0", "0", "0", "ECO", "T",
                   "eBay SpeedPAK Economy", o["dest"], "eBay SpeedPAK Economy", "DDP"])

    ws = wb.create_sheet(idr.SH_SKU)
    ws.append(["OrangeConnex追跡番号", "輸送業者名", "輸送業者追跡番号", "仕向地", "申告単価",
               "通貨", "eBay取引ID", "SKUの詳細", "バッテリー有無", "原産国",
               "申請HSコード", "HS (HTS)コード(再分類)"])
    for s in skus:
        ws.append([s["tid"], "", "T", s["dest"], s["price"], s["cur"], "1",
                   s["title"], "NO", "JP", "", s["hs"]])

    ws = wb.create_sheet(idr.SH_FEE)
    ws.append(["OrangeConnex追跡番号", "支払いタイプ", "通貨", "還元金額", "小計",
               "料金詳細"] + [""] * (len(fee_cols) - 1))
    ws.append(["", "", "", "", ""] + list(fee_cols))
    for tid, typ, vals in fee_rows:
        ws.append([tid, typ, "JPY", "0", "0"] + [str(v) for v in vals])
    wb.save(path)
    return path


@pytest.mark.parametrize("raw,want", [
    ("1,289", 1289.0), ("¥4,443", 4443.0), (" 245 ", 245.0),
    (278.98, 278.98), ("", 0.0), (None, 0.0), ("abc", 0.0),
])
def test_num_parses_invoice_formats(raw, want):
    assert idr._num(raw) == want


def test_unknown_workbook_raises_not_silent(tmp_path):
    """書式が変わったら例外。0件で黙って返すと『関税ゼロ』と誤読される."""
    import openpyxl
    p = tmp_path / "wrong.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "別の形式"
    wb.save(p)
    with pytest.raises(ValueError, match="想定のシートが無い"):
        idr.parse(str(p))


def test_columns_are_read_by_header_not_position(tmp_path):
    """★2026-08-08: 8月版 export は 運送料金/燃料割増金 が5月版と**逆順**。

    列番号決め打ちだと燃油率が 833% になり、「燃油が下がった」と誤報告する。
    並びを入れ替えても同じ値が出ることを固定する。
    """
    got = []
    for cols in (FEE_COLS, ["燃料割増金", "運送料金"] + FEE_COLS[2:]):
        vals = {"運送料金": 1491, "燃料割増金": 231, "推定関税及び税金料金": 937,
                "輸入通関手数料": 245, "推定関税処理手数料": 20, "関税": 0, "関税処理手数料": 0}
        p = _build(tmp_path / f"{cols[0]}.xlsx", cols,
                   [("EE1", "引落金額", [vals[c] for c in cols]),
                    ("", "合計金額", [vals[c] for c in cols])],
                   [{"tid": "EE1", "dest": "US", "price": 58.98, "cur": "USD",
                     "title": "tee", "hs": "6109901007"}],
                   [{"tid": "EE1", "date": "2026-05-22", "dest": "US"}])
        got.append(idr.parse(str(p))[0])
    a, b = got
    assert a["ship"] == b["ship"] == 1491
    assert a["fuel"] == b["fuel"] == 231
    assert abs(a["fuel_rate"] - 0.155) < 0.001, a["fuel_rate"]
    assert a["fuel_rate"] == b["fuel_rate"]


def test_adjustment_row_replaces_estimate_with_actual_duty(tmp_path):
    """★関税は2段階請求。`引落金額`=推定 → 後日 `料金調整` で `関税` 列に実額。

    引落金額行だけ読むと 5/22 の Tシャツが 9.9% に見える (実際は 41.8%)。
    """
    p = _build(tmp_path / "adj.xlsx", FEE_COLS, [
        ("EE1", "引落金額", [1491, 231, 937, 245, 20, 0, 0]),
        ("", "料金調整", [0, 0, -937, 0, -20, 3938, 83]),
        ("", "合計金額", [1491, 231, 0, 245, 0, 3938, 83]),
    ], [{"tid": "EE1", "dest": "US", "price": 58.98, "cur": "USD",
         "title": "tee", "hs": "6109901007"}],
        [{"tid": "EE1", "date": "2026-05-22", "dest": "US"}])
    r = idr.parse(str(p), fx_usd=159.65)[0]
    assert r["duty"] == 3938, "実額でなく推定を拾っている"
    assert r["duty_estimated"] == 0
    assert r["status"] == "final"
    assert 0.41 < r["duty_rate"] < 0.43, r["duty_rate"]


def test_estimate_only_is_flagged_not_treated_as_final(tmp_path):
    """料金調整がまだ来ていない = 推定。これで採算判断すると過小評価する."""
    p = _build(tmp_path / "est.xlsx", ["燃料割増金", "運送料金", "推定関税及び税金料金",
                                       "輸入通関手数料", "推定関税処理手数料"], [
        ("EE1", "引落金額", [179, 1491, 1118, 245, 23]),
        ("", "合計金額", [179, 1491, 1118, 245, 23]),
    ], [{"tid": "EE1", "dest": "US", "price": 56.98, "cur": "USD",
         "title": "UT tee", "hs": "6109908010"}],
        [{"tid": "EE1", "date": "2026-07-31", "dest": "US"}])
    r = idr.parse(str(p), fx_usd=159.65)[0]
    assert r["status"] == "estimate", "推定なのに確定扱いしている"
    assert r["duty"] == 1118


def test_grand_total_row_is_not_merged_into_last_order(tmp_path):
    """★シート末尾の総合計行を最後の tid に合算しない.

    合算すると 5/22 の Tシャツが ¥3,938 → ¥8,381 (= 5/28 分込み) になり 89% と出る。
    """
    p = _build(tmp_path / "total.xlsx", FEE_COLS, [
        ("EE1", "引落金額", [1289, 200, 4443, 245, 93, 0, 0]),
        ("", "料金調整", [0, 0, -4443, 0, -93, 4443, 93]),
        ("", "合計金額", [1289, 200, 0, 245, 0, 4443, 93]),
        ("EE2", "引落金額", [1491, 231, 937, 245, 20, 0, 0]),
        ("", "料金調整", [0, 0, -937, 0, -20, 3938, 83]),
        ("", "合計金額", [1491, 231, 0, 245, 0, 3938, 83]),
        ("", "合計金額", [2780, 431, 0, 490, 0, 8381, 176]),   # ← 総合計
    ], [{"tid": "EE1", "dest": "US", "price": 278.98, "cur": "USD",
         "title": "psa", "hs": "9504400000"},
        {"tid": "EE2", "dest": "US", "price": 58.98, "cur": "USD",
         "title": "tee", "hs": "6109901007"}],
        [{"tid": "EE1", "date": "2026-05-28", "dest": "US"},
         {"tid": "EE2", "date": "2026-05-22", "dest": "US"}])
    by = {r["tid"]: r for r in idr.parse(str(p), fx_usd=159.65)}
    assert len(by) == 2, f"総合計行を1注文として数えている: {list(by)}"
    assert by["EE1"]["duty"] == 4443
    assert by["EE2"]["duty"] == 3938, "総合計が最後の注文に乗っている"


def test_multiple_skus_in_one_order_are_summed(tmp_path):
    """1注文に複数SKUなら申告額は合算。上書きすると分母が小さくなり率が跳ねる."""
    p = _build(tmp_path / "multi.xlsx", FEE_COLS, [
        ("EE1", "引落金額", [1491, 231, 0, 245, 0, 2000, 42]),
        ("", "合計金額", [1491, 231, 0, 245, 0, 2000, 42]),
    ], [{"tid": "EE1", "dest": "US", "price": 30.0, "cur": "USD",
         "title": "a", "hs": "6109901007"},
        {"tid": "EE1", "dest": "US", "price": 20.0, "cur": "USD",
         "title": "a", "hs": "6109901007"}],
        [{"tid": "EE1", "date": "2026-05-22", "dest": "US"}])
    r = idr.parse(str(p), fx_usd=160.0)[0]
    assert r["declared"] == 50.0 and r["skus"] == 2
    assert r["declared_jpy"] == 8000


@pytest.mark.skipif(not REAL.exists() or not list(REAL.glob("*.xlsx")),
                    reason="実請求書が未配置")
def test_real_invoice_effective_rate():
    """実請求書: 実効率が妥当域で、処理料が関税の約2.1%であること."""
    rows = []
    for f in REAL.glob("*.xlsx"):
        rows += idr.parse(str(f))
    assert rows, "明細が取れていない"
    for r in rows:
        # Tシャツ (MFN 32% + 約10%) で 42% まで出る。1.0 を超えたら読み間違いを疑う
        assert 0 <= r["duty_rate"] < 1.0, f"実効率が異常: {r}"
        assert r["declared_jpy"] > 0
        if r["duty"]:
            # PDF: 米国関税処理手数料 = 関税合計額の 2.1%
            assert 0.015 < r["proc_rate"] < 0.03, f"処理料率が2.1%から外れる: {r}"
        if r["ship"]:
            # 燃油サーチャージ 実測 15.5% (5月) → 12.0% (8月)
            assert 0.10 < r["fuel_rate"] < 0.25, f"燃油率が想定外: {r}"


@pytest.mark.skipif(not REAL.exists() or not list(REAL.glob("*.xlsx")),
                    reason="実請求書が未配置")
def test_hs_code_does_change_the_rate():
    """★HSコードで率が変わる = 関税は「HTS本則 + 約10%」(2026-08-08 実額で確認).

    旧テストは逆 (「HS が違っても同率」) を固定していたが、それは推定額を読んでいた
    ためのバグだった。カテゴリ別 hts_rate の設計は**意味を持つ**ので、
    group A/B/C の値は実額ベースで維持すること。

    この assert が落ちる = 実額の読み取りがまた推定に戻った可能性が高い。
    """
    rows = [r for f in REAL.glob("*.xlsx") for r in idr.parse(str(f))]
    us = [r for r in rows if r["dest"] == "US" and r["duty"] and r["status"] == "final"]
    if len({r["hs"][:6] for r in us}) < 2:
        pytest.skip("HSコードが2種類以上ある確定済 US 明細が足りない")
    by_hs = {r["hs"][:6]: r["duty_rate"] for r in us}
    assert max(by_hs.values()) - min(by_hs.values()) > 0.05, (
        f"HS別に率が変わっていない → 推定額を読んでいないか確認: {by_hs}")
    # MFN 0% の TCG (9504.40) は上乗せ分だけ = 約10%
    if "950440" in by_hs:
        assert 0.08 < by_hs["950440"] < 0.13, f"TCG が約10%から外れる: {by_hs}"
    # MFN 32% の Tシャツ (6109.90) は 32 + 約10 = 約42%
    if "610990" in by_hs:
        assert 0.38 < by_hs["610990"] < 0.46, f"Tシャツが約42%から外れる: {by_hs}"
