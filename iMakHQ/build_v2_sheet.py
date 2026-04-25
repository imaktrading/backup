import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = Workbook()
wb.remove(wb.active)

bold = Font(bold=True)
header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
section_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
note_font = Font(italic=True, color="808080", size=9)
pct = '0.00%'
yen = '¥#,##0'
usd = '"$"#,##0.00'

# ================= 設定シート =================
ws = wb.create_sheet("設定")
for col, w in [('A', 28), ('B', 14), ('C', 18), ('D', 48)]:
    ws.column_dimensions[col].width = w

ws['A1'] = "■ 共通設定"
ws['A1'].fill = section_fill
ws['A1'].font = bold

ws['A2'] = "為替レート (USD→JPY)"
ws['B2'] = '=IFERROR(GOOGLEFINANCE("CURRENCY:USDJPY"),159.245)'
ws['B2'].number_format = '#,##0.000'
ws['D2'] = "Google Sheetsで自動更新 / Excelではフォールバック159.245"
ws['D2'].font = note_font
FX_REF = '設定!$B$2'

ws['A3'] = "実効プロモ率"
ws['B3'] = 0.06
ws['B3'].number_format = pct
ws['D3'] = "販売実績実測5.82% → 6%(保守的round up)。旧11%から修正"
ws['D3'].font = note_font
PROMO_REF = '設定!$B$3'

ws['A4'] = "ペイオニア手数料率"
ws['B4'] = 0.025
ws['B4'].number_format = pct
ws['D4'] = "FXマークアップ+銀行出金の合算推定 (新規追加)"
ws['D4'].font = note_font
PAYO_REF = '設定!$B$4'

ws['A5'] = "目標最低利益率"
ws['B5'] = 0.10
ws['B5'].number_format = pct
ws['D5'] = "仕入GATE判定の閾値"
ws['D5'].font = note_font
TARGET_REF = '設定!$B$5'

ws['A7'] = "■ 国別eBay実効手数料率"
ws['A7'].fill = section_fill
ws['A7'].font = bold

ws['A8'] = "国"
ws['B8'] = "実効率"
ws['C8'] = "備考(実測値)"
for c in ['A8', 'B8', 'C8']:
    ws[c].font = bold
    ws[c].fill = header_fill

countries = [
    ("US", 0.185, "実測18.53%"),
    ("UK", 0.225, "実測22.55% (VAT関連)"),
    ("AU", 0.150, "実測14.88%"),
    ("EU", 0.185, "US相当想定"),
    ("その他", 0.180, "平均想定"),
]
for i, (c, r, m) in enumerate(countries, start=9):
    ws[f'A{i}'] = c
    ws[f'B{i}'] = r
    ws[f'B{i}'].number_format = pct
    ws[f'C{i}'] = m
    ws[f'C{i}'].font = note_font

FEE_US = '設定!$B$9'
FEE_UK = '設定!$B$10'
FEE_AU = '設定!$B$11'

# カテゴリ
ws['A15'] = "■ カテゴリ設定"
ws['A15'].fill = section_fill
ws['A15'].font = bold

ws['A16'] = "カテゴリ名"
ws['B16'] = "手数料参考"
ws['C16'] = "実送料(JPY)"
ws['D16'] = "備考"
for c in ['A16', 'B16', 'C16', 'D16']:
    ws[c].font = bold
    ws[c].fill = header_fill

categories = [
    ("TCG(PSA10)", 0.135, 2000, "新カテゴリ・データ蓄積中"),
    ("G-SHOCK", 0.175, 2000, "年11件・黒字率73%"),
    ("Tシャツ(UT)", 0.149, 2000, "★主力 年33件・黒字率91%"),
    ("Montbell(一般)", 0.149, 2000, "バッグ小物・非アウター"),
    ("Montbell(ジャケット)", 0.149, 4500, "嵩張り注意・送料厚め"),
    ("一番くじ", 0.149, 2500, "新・実測調整中"),
    ("フィギュア", 0.149, 3500, ""),
    ("ユニクロ(非UT)", 0.149, 2000, ""),
    ("サンリオ文具", 0.149, 2000, ""),
    ("ヴィンテージ玩具", 0.149, 2500, ""),
    ("トミカ", 0.149, 2000, ""),
    ("POPMart", 0.149, 2500, ""),
    ("ガシャポン", 0.149, 2000, ""),
    ("ダイソー", 0.149, 2000, "薄利注意"),
    ("バッグ(アネロ)", 0.149, 2500, ""),
    ("サンリオぬいぐるみ", 0.149, 2500, ""),
]
cat_start = 17
for i, (cat, fee, ship, memo) in enumerate(categories, start=cat_start):
    ws[f'A{i}'] = cat
    ws[f'B{i}'] = fee
    ws[f'B{i}'].number_format = pct
    ws[f'C{i}'] = ship
    ws[f'C{i}'].number_format = yen
    ws[f'D{i}'] = memo
    ws[f'D{i}'].font = note_font
cat_end = cat_start + len(categories) - 1
CAT_RANGE = f'設定!$A${cat_start}:$D${cat_end}'


# =============== 計算シート生成関数 ===============
def create_calc(wb, name, fee_ref):
    ws = wb.create_sheet(name)
    widths = {'A': 11, 'B': 13, 'C': 11, 'D': 13, 'E': 13, 'F': 12, 'G': 12, 'H': 11,
              'I': 14, 'J': 11, 'K': 10, 'L': 14, 'M': 10, 'N': 14, 'O': 14, 'P': 14,
              'Q': 10, 'R': 11, 'S': 10, 'T': 11, 'U': 10}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    # Row 1: 設定参照
    ws['A1'] = "国"
    ws['B1'] = name.replace("計算", "")
    ws['A1'].font = bold
    ws['C1'] = "為替"
    ws['D1'] = f'={FX_REF}'
    ws['C1'].font = bold
    ws['D1'].number_format = '#,##0.000'
    ws['H1'] = "手数料率"
    ws['I1'] = f'={fee_ref}'
    ws['H1'].font = bold
    ws['I1'].number_format = pct
    ws['R1'] = "プロモ"
    ws['S1'] = f'={PROMO_REF}'
    ws['R1'].font = bold
    ws['S1'].number_format = pct
    ws['T1'] = "Payo"
    ws['U1'] = f'={PAYO_REF}'
    ws['T1'].font = bold
    ws['U1'].number_format = pct

    # Row 2: ヘッダ
    headers_map = {
        'A': '送料込み', 'B': '割引率', 'D': '出品価格($)', 'E': '仕入(¥)', 'F': 'ポイント還元',
        'G': '売上(¥)', 'H': '実送料(¥)', 'I': 'eBay手数料(¥)', 'J': 'DDP送料($)', 'K': 'プロモ率',
        'L': '利益(¥)', 'M': '利益率', 'N': 'プロモ費(¥)', 'P': 'Payo費(¥)'
    }
    for col, h in headers_map.items():
        c = ws[f'{col}2']
        c.value = h
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    def make_row(r, is_input_row=False):
        pr = f'C{r}'
        ws[f'D{r}'] = f'=INT({pr})+0.98'
        if not is_input_row:
            ws[f'E{r}'] = '=$E$4'
            ws[f'E{r}'].number_format = yen
            ws[f'H{r}'] = '=$H$4'
            ws[f'H{r}'].number_format = yen
            ws[f'F{r}'] = '=$F$4'
        ws[f'G{r}'] = f'=D{r}*$D$1'
        ws[f'G{r}'].number_format = yen
        ws[f'J{r}'] = f'=IFS(D{r}>=800,260,D{r}>=600,210,D{r}>=500,155,D{r}>=400,130,D{r}>=300,105,D{r}>=200,75,D{r}>=100,50,D{r}>=60,30,D{r}>=40,20,D{r}<40,15)'
        ws[f'I{r}'] = f'=((D{r}+J{r})*$D$1)*$I$1'
        ws[f'I{r}'].number_format = yen
        ws[f'K{r}'] = '=$S$1'
        ws[f'K{r}'].number_format = pct
        ws[f'N{r}'] = f'=(D{r}+J{r})*K{r}*$D$1'
        ws[f'N{r}'].number_format = yen
        ws[f'P{r}'] = f'=G{r}*$U$1'
        ws[f'P{r}'].number_format = yen
        ws[f'L{r}'] = f'=G{r}-H{r}-I{r}-E{r}-N{r}-P{r}+F{r}'
        ws[f'L{r}'].number_format = yen
        ws[f'M{r}'] = f'=L{r}/G{r}'
        ws[f'M{r}'].number_format = pct
        ws[f'A{r}'] = f'=D{r}+J{r}'

    # Row 3: 推奨 (refers to row 18 tier A for Tシャツ)
    ws['B3'] = '推奨'
    ws['B3'].font = bold
    ws['C3'] = '=J18'
    make_row(3)

    # Row 4: 出品中 (MAIN INPUT)
    ws['B4'] = '出品中'
    ws['B4'].font = bold
    ws['B4'].fill = input_fill
    ws['C4'] = 432.0
    ws['C4'].fill = input_fill
    ws['E4'] = 37000
    ws['E4'].number_format = yen
    ws['E4'].fill = input_fill
    ws['F4'] = 0
    ws['H4'] = 2000
    ws['H4'].number_format = yen
    ws['H4'].fill = input_fill
    make_row(4, is_input_row=True)

    # Row 5-8: 割引ラダー
    for r, disc in [(5, 0.05), (6, 0.08), (7, 0.10), (8, 0.15)]:
        ws[f'B{r}'] = disc
        ws[f'B{r}'].number_format = pct
        ws[f'C{r}'] = f'=$C$4*(1-B{r})'
        make_row(r)

    # Row 11: オファー価格
    ws['B11'] = 'オファー価格'
    ws['B11'].font = bold
    ws['C11'] = '=J18'
    make_row(11)

    # Row 12: 試算用
    ws['B12'] = '試算用'
    ws['C12'] = 240
    ws['C12'].fill = input_fill
    make_row(12)

    # Row 15-21: カテゴリ参考価格
    ws['C15'] = "■ カテゴリ別参考価格 (/140 参考レート基準)"
    ws['C15'].fill = section_fill
    ws['C15'].font = bold

    ws['D16'] = "手数料"
    ws['E16'] = "実送料"
    ws['H16'] = "仕入"
    ws['I16'] = "SS"
    ws['J16'] = "A"
    ws['K16'] = "B"
    ws['L16'] = "C"
    ws['M16'] = "D"
    ws['N16'] = "E"
    ws['O16'] = "E2"
    for c in ['D16', 'E16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16']:
        ws[c].font = bold
        ws[c].fill = header_fill
        ws[c].alignment = Alignment(horizontal='center')

    cat_refs = [
        (17, 'TCG(PSA10)', 0.135, 2000),
        (18, 'Tシャツ/Montbell一般', 0.149, 2000),
        (19, 'フィギュア', 0.149, 3500),
        (20, 'G-SHOCK', 0.175, 2000),
        (21, 'Montbell(ジャケット)', 0.149, 4500),
    ]
    for r, cat, fee, ship in cat_refs:
        ws[f'C{r}'] = cat
        ws[f'D{r}'] = fee
        ws[f'D{r}'].number_format = pct
        ws[f'E{r}'] = ship
        ws[f'E{r}'].number_format = yen
        ws[f'H{r}'] = '=E4'
        ws[f'H{r}'].number_format = yen
        net_ratio = '(1-$I$1-$S$1-$U$1)'
        for col, buf in [('I', 5000), ('J', 4000), ('K', 3000), ('L', 2000),
                         ('M', 1000), ('N', 500), ('O', 0)]:
            ws[f'{col}{r}'] = f'=ROUNDDOWN(((H{r}+E{r}+{buf})/{net_ratio})/140,0)+0.98'

    ws['C23'] = "※ /140 は意図的な参考レート (C4で手動調整する運用)"
    ws['C23'].font = note_font
    ws['C24'] = "※ net_ratio = 1 - 手数料 - プロモ - Payo で国別自動補正済"
    ws['C24'].font = note_font

    ws.freeze_panes = 'A3'
    return ws


create_calc(wb, "US計算", FEE_US)
create_calc(wb, "UK計算", FEE_UK)
create_calc(wb, "AU計算", FEE_AU)

# ================= 仕入GATE =================
ws = wb.create_sheet("仕入GATE")
for c, w in [('A', 28), ('B', 16), ('C', 52)]:
    ws.column_dimensions[c].width = w

ws['A1'] = "■ 仕入GATE 判定"
ws['A1'].fill = section_fill
ws['A1'].font = bold

ws['A3'] = "【入力】"
ws['A3'].font = bold
ws['A4'] = "仕入値(JPY)"
ws['B4'] = 5000
ws['B4'].number_format = yen
ws['B4'].fill = input_fill

ws['A5'] = "カテゴリ名"
ws['B5'] = "Tシャツ(UT)"
ws['B5'].fill = input_fill
ws['C5'] = "設定シートのカテゴリ名を正確に入力 (大文字小文字/括弧まで一致)"
ws['C5'].font = note_font

ws['A6'] = "対象国"
ws['B6'] = "US"
ws['B6'].fill = input_fill
ws['C6'] = "US / UK / AU / EU / その他"
ws['C6'].font = note_font

ws['A7'] = "eBay SOLD中央値($)"
ws['B7'] = 0
ws['B7'].number_format = usd
ws['B7'].fill = input_fill
ws['C7'] = "直近30日・類似品のSOLD中央値を手動入力"
ws['C7'].font = note_font

ws['A9'] = "【参照値(自動)】"
ws['A9'].font = bold

ws['A10'] = "参照送料(¥)"
ws['B10'] = f'=VLOOKUP(B5,{CAT_RANGE},3,FALSE)'
ws['B10'].number_format = yen

ws['A11'] = "国別手数料率"
ws['B11'] = '=VLOOKUP(B6,設定!$A$9:$B$13,2,FALSE)'
ws['B11'].number_format = pct

ws['A12'] = "プロモ率"
ws['B12'] = f'={PROMO_REF}'
ws['B12'].number_format = pct

ws['A13'] = "ペイオニア率"
ws['B13'] = f'={PAYO_REF}'
ws['B13'].number_format = pct

ws['A14'] = "目標利益率"
ws['B14'] = f'={TARGET_REF}'
ws['B14'].number_format = pct

ws['A15'] = "為替レート"
ws['B15'] = f'={FX_REF}'
ws['B15'].number_format = '#,##0.000'

ws['A17'] = "【計算】"
ws['A17'].font = bold

ws['A18'] = "損益分岐 最低$"
ws['B18'] = '=(B4+B10)/((1-B11-B12-B13)*B15)'
ws['B18'].number_format = usd
ws['C18'] = "この価格で売れば利益ゼロ"
ws['C18'].font = note_font

ws['A19'] = "目標確保 下限$"
ws['B19'] = '=(B4+B10)/((1-B11-B12-B13-B14)*B15)'
ws['B19'].number_format = usd
ws['C19'] = "目標利益率を満たす最低eBay価格"
ws['C19'].font = note_font

ws['A21'] = "【GATE判定】"
ws['A21'].font = bold

ws['A22'] = "判定"
ws['B22'] = '=IF(B7=0,"SOLD中央値未入力",IF(B7*0.9>=B19,"✅ GO","❌ NOGO"))'
ws['B22'].font = Font(bold=True, size=14)
ws['C22'] = "条件: SOLD中央値×90% ≥ 目標確保下限$"
ws['C22'].font = note_font

ws['A23'] = "期待売価($)"
ws['B23'] = '=IF(B7=0,0,B7*0.95)'
ws['B23'].number_format = usd
ws['C23'] = "SOLD中央値×95% (市場より少し安く設定)"
ws['C23'].font = note_font

ws['A24'] = "期待利益(¥)"
ws['B24'] = '=IF(B7=0,0,B23*B15*(1-B11-B12-B13)-B4-B10)'
ws['B24'].number_format = yen

ws['A25'] = "期待利益率"
ws['B25'] = '=IF(OR(B7=0,B23=0),0,(B23*B15*(1-B11-B12-B13)-B4-B10)/(B23*B15))'
ws['B25'].number_format = pct

ws['A28'] = "■ 使い方"
ws['A28'].fill = section_fill
ws['A28'].font = bold
usage = [
    "1. メルカリで候補を見つけたら 仕入値(B4) を入力",
    "2. カテゴリ(B5)・対象国(B6)を選択 → 送料・手数料が自動入力",
    "3. 目標確保下限$(B19) が出る = この価格以上で売らないと目標利益が出ない",
    "4. eBay で同等商品のSOLD中央値を調べて B7 に入力",
    "5. GATE判定(B22) を確認:",
    "   ✅ GO → 出品する価値あり (市場価格が目標下限を満たす)",
    "   ❌ NOGO → 市場価格がコストプラス目標に届かない。見送り推奨",
    "",
    "※ 判定条件は 中央値×90% ≥ 目標下限$。90%の理由: 出品価格を市場より",
    "   少し安く設定して成約率を上げるため。攻めたいなら 0.95 等に調整可",
    "※ DDP送料・割引分は未考慮のシンプル版",
    "※ Phase2 で iMakeBayAPI 経由で SOLD中央値を自動取得予定",
]
for i, u in enumerate(usage, start=29):
    ws[f'A{i}'] = u
    ws.merge_cells(f'A{i}:C{i}')
    ws[f'A{i}'].font = note_font

# ================= 変更履歴 =================
ws = wb.create_sheet("変更履歴")
for c, w in [('A', 14), ('B', 58), ('C', 48)]:
    ws.column_dimensions[c].width = w

ws['A1'] = "■ 変更履歴 (旧【NEW】利益計算シート.xlsx からの差分)"
ws['A1'].fill = section_fill
ws['A1'].font = bold
ws.merge_cells('A1:C1')

ws['A3'] = "区分"
ws['B3'] = "内容"
ws['C3'] = "理由・備考"
for c in ['A3', 'B3', 'C3']:
    ws[c].font = bold
    ws[c].fill = header_fill

changes = [
    ('修正', 'プロモ率 0.11 → 0.06', '販売実績実測 5.82%(保守的にround up)'),
    ('修正', '国別手数料率を分離 (US18.5 / UK22.5 / AU15.0)', '実測: US18.53% / UK22.55% / AU14.88%'),
    ('新規', 'ペイオニア手数料 2.5% を独立費目に', 'FX+出金の実コストを顕在化 (隠れ費用の可視化)'),
    ('新規', '設定シート (中央制御盤)', '全パラメータを1箇所集約、全シートから参照'),
    ('新規', '仕入GATEシート', 'メルカリ仕入時のGO/NOGO判定 (シンプル版)'),
    ('新規', 'Montbell(ジャケット) を別カテゴリ化', '送料4500円で分離 (過去赤字3件が全てジャケット)'),
    ('改善', 'カテゴリ参考価格の後費用率を国別自動補正', '旧: /0.72 固定 → 新: 1 - 手数料 - プロモ - Payo'),
    ('保持', '140円/$ ハードコード (I-O 17-21行)', '意図的参考レート。C4で手動調整運用を尊重'),
    ('保持', 'eBay手数料 18.5% (US)', '実測18.37%と一致、TAX/miscバッファ込み構造'),
    ('保持', '割引ラダー 5/8/10/15%', 'オファー時・値下げ時の試算用'),
    ('保持', 'US/UK/AU の3シート分離', '地域別に手数料率が異なるため'),
    ('保持', '推奨行/オファー価格/試算行', '既存ワークフローを維持'),
    ('', '', ''),
    ('作成日', '2026-04-12', ''),
    ('作成者', 'Claude (iMakHQ 構想会話内で自動生成)', ''),
    ('旧ファイル', '【NEW】利益計算シート.xlsx (同じデスクトップ)', '戻したい時はそちらを参照'),
]
for i, (a, b, c) in enumerate(changes, start=4):
    ws[f'A{i}'] = a
    ws[f'B{i}'] = b
    ws[f'C{i}'] = c
    if a == '修正':
        ws[f'A{i}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    elif a == '新規':
        ws[f'A{i}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    elif a == '改善':
        ws[f'A{i}'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    elif a == '保持':
        ws[f'A{i}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# シート順
desired_order = ['設定', 'US計算', 'UK計算', 'AU計算', '仕入GATE', '変更履歴']
wb._sheets = [wb[name] for name in desired_order]

output = r'C:/Users/imax2/OneDrive/デスクトップ/【NEW】利益計算シート_v2.xlsx'
wb.save(output)
print("Saved:", output)
print("Sheets:", wb.sheetnames)
