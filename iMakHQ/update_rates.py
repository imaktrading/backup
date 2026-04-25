"""
為替レート自動更新スクリプト

利益計算シート v2 の設定シートに最新の為替レート(対JPY)を書き込む。
- B2: USD/JPY
- F2: EUR/JPY
- H2: GBP/JPY
- J2: AUD/JPY

ソース: Frankfurter API (ECB公式データ、無料・APIキー不要)
実行: python update_rates.py
"""

from __future__ import annotations

import json
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SHEET_PATH = Path(r"c:\dev\iMak\iMakHQ\sheets\【NEW】利益計算シート_v2.xlsx")
LOG_PATH = Path(r"c:\dev\iMak\iMakHQ\update_rates.log")

# 無料API (APIキー不要)
API_URL = "https://api.exchangerate-api.com/v4/latest/USD"


def fetch_all_rates() -> dict[str, float] | None:
    """USD基準の為替レートを取得し、{通貨: JPYに対するレート} で返す"""
    try:
        with urllib.request.urlopen(API_URL, timeout=10) as resp:
            data = json.loads(resp.read())
        rates = data["rates"]
        usd_jpy = float(rates["JPY"])
        # XXX/USD = 1 / (USD/XXX)、XXX/JPY = XXX/USD × USD/JPY
        return {
            "USD": usd_jpy,
            "EUR": usd_jpy / float(rates["EUR"]),
            "GBP": usd_jpy / float(rates["GBP"]),
            "AUD": usd_jpy / float(rates["AUD"]),
        }
    except Exception as e:
        print(f"  ERROR: {e}")
        return None


def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def main() -> int:
    raw = fetch_all_rates()
    if raw is None:
        log("FAIL: 為替レート取得失敗")
        return 1
    rates = {k: round(v, 3) for k, v in raw.items()}
    for code in ("USD", "EUR", "GBP", "AUD"):
        log(f"{code}/JPY = {rates[code]}")

    if not SHEET_PATH.exists():
        log(f"FAIL: {SHEET_PATH} が見つからない")
        return 2

    try:
        wb = load_workbook(SHEET_PATH)
    except PermissionError:
        log(f"FAIL: シートが開かれています。閉じてから再実行してください")
        return 3

    ws = wb["設定"]
    ws["B2"] = rates["USD"]
    ws["F2"] = rates["EUR"]
    ws["H2"] = rates["GBP"]
    ws["J2"] = rates["AUD"]

    try:
        wb.save(SHEET_PATH)
    except PermissionError:
        log(f"FAIL: シートが開かれていて保存不可")
        return 3

    log(f"更新完了: USD={rates['USD']}, EUR={rates['EUR']}, GBP={rates['GBP']}, AUD={rates['AUD']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
