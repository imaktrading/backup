# -*- coding: utf-8 -*-
"""Orange Connex 請求書 (xlsx) から **実効関税率** を算出する。

なぜ必要か (2026-07-31):
    SpeedPAK の関税は「事前設定レート」で請求され、**その率はどの資料にも書かれていない**。
    料金表 (RATE_GUIDE_SpeedPAK_Economy_JP_20260730.pdf L430-437) に明記:

      「貨物の申告価値及び平均関税率に基づき、特定割合(事前設定レート)で算出される。
        適用される事前設定レートは随時変更される場合があり、実際の課金時点の
        事前設定レートが優先される。**最終的な決済は請求金額に基づく**」

    = **請求書が唯一の一次情報**。HTS を引いても実際の請求額は分からない。
    実測 (2026-05 の2件) では HSコードが違っても両方 9.95〜9.98% で、
    Orange Connex は真の HTS を見ずに一律レートを当てていた。

    さらに米国側の制度が短期間で動いている:
      2025-09 日本に相互関税15%(IEEPA) → 2026-02-20 最高裁が無効化
      → 一律10%(Section 122) → **2026-07-24 失効**
    → **請求書を継続的に取り込んで実効率を追う**しかない。

使い方:
    python invoice_duty_rate.py <請求書.xlsx> [<請求書2.xlsx> ...]
    python invoice_duty_rate.py --dir C:/dev/iMak_data/shipping/invoices

出力: 明細ごとの実効率 + 仕向地/期間ごとの集計。
      `iMak_data/shipping/duty_rate_history.jsonl` に追記して推移を残す。
"""
from __future__ import annotations

import glob
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

HIST = Path(r"C:\dev\iMak_data\shipping\duty_rate_history.jsonl")

# 請求書 (新版) のシート名。書式が変わったらここだけ直す。
SH_ORDER, SH_SKU, SH_FEE = "オーダー", "SKUの詳細", "料金詳細（新版）"

# 料金詳細 の費目見出し。**列番号で読んではいけない** (2026-08-08):
#   5月版 = 運送料金, 燃料割増金 / 8月版 = 燃料割増金, 運送料金 と **並びが入れ替わる**。
#   番号決め打ちだと燃油率が 833% になる (実害: 「燃油が 15.5%→12.0% に下がった」と誤報告)。
C_SHIP, C_FUEL = "運送料金", "燃料割増金"
C_DUTY_EST, C_DUTY_ACT = "推定関税及び税金料金", "関税"
C_PROC_EST, C_PROC_ACT = "推定関税処理手数料", "関税処理手数料"
C_CLEAR = "輸入通関手数料"


def _num(v):
    try:
        return float(str(v).replace(",", "").replace("¥", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _fee_header(ws) -> dict:
    """料金詳細 の見出しは2行に割れている (1行目=グループ / 2行目=費目)。費目名→列index を返す。"""
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    top, sub = rows[0], rows[1] if len(rows) > 1 else ()
    idx = {}
    for i, name in enumerate(top):
        if name and not (i < len(sub) and sub[i]):
            idx[str(name).strip()] = i
    for i, name in enumerate(sub):
        if name:
            idx[str(name).strip()] = i
    return idx


def parse(path: str, fx_usd: float = 159.65) -> list[dict]:
    """請求書1ファイル → 明細 list。fx は申告通貨→円の換算に使う。

    ★関税は2段階で請求される (2026-08-08 判明)。
      `引落金額` 行 = **推定**関税。後日 `料金調整` 行で推定を打ち消し、`関税` 列に**実額**が入る。
      `合計金額` 行がその tid の最終状態なので、**必ず合計金額行を読む**。
      推定のまま (= `関税` 列が 0/不在) なら `status="estimate"` で、まだ確定していない。
      引落金額行だけ読むと実額を取り落とす (実害: 5/22 の Tシャツを 9.95% と記録。実際は 41.9%)。
    """
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if SH_SKU not in wb.sheetnames or SH_FEE not in wb.sheetnames:
        raise ValueError(f"想定のシートが無い ({wb.sheetnames}) — 請求書の書式変更を疑うこと")

    # 1注文に複数SKUが載りうるので申告額は合算する
    sku: dict[str, dict] = {}
    for r in wb[SH_SKU].iter_rows(min_row=2, values_only=True):
        if not (r and r[0]):
            continue
        s = sku.setdefault(str(r[0]), {"dest": r[3], "price": 0.0, "cur": r[5],
                                       "title": str(r[7] or "")[:60], "origin": r[9],
                                       "hs": str(r[11] or ""), "skus": 0})
        s["price"] += _num(r[4])
        s["skus"] += 1
    order = {}
    for r in wb[SH_ORDER].iter_rows(min_row=2, values_only=True):
        if r and r[1]:
            order[str(r[1])] = {"date": str(r[0])[:10], "service": r[8], "incoterm": r[11]}

    ws = wb[SH_FEE]
    col = _fee_header(ws)

    def g(row, name):
        i = col.get(name)
        return _num(row[i]) if i is not None and i < len(row) else 0.0

    # tid は最初の行だけに入り、以降の 料金調整 / 合計金額 行は空欄 → 直前の tid を引き継ぐ。
    # ★シート末尾に **全 tid を合算した総合計行** がぶら下がる (同じく tid 空欄・`合計金額`)。
    #   単純に「最後の合計金額行」を採ると、最後の tid に全注文の関税が合算されて乗る
    #   (実害: 5/22 の Tシャツが ¥3,938 → ¥8,381 = 5/28 分との合算で 89% と表示)。
    #   → tid ごとに **最初の** 合計金額行だけを採り、以降 (= 総合計) は捨てる。
    final: dict[str, tuple] = {}
    totaled: set[str] = set()
    cur_tid = ""
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r:
            continue
        if r[0] and str(r[0]).startswith("EE"):
            cur_tid = str(r[0])
        typ = str(r[1] or "")
        if not cur_tid or typ not in ("引落金額", "合計金額"):
            continue
        if typ == "合計金額":
            if cur_tid in totaled:      # 2本目以降の合計金額行 = 総合計。捨てる
                continue
            final[cur_tid] = r
            totaled.add(cur_tid)
        elif cur_tid not in final:      # 合計金額が無い書式なら引落金額で代用 (= 推定のみ)
            final[cur_tid] = r

    out = []
    for tid, r in final.items():
        s = sku.get(tid)
        if not s or not s["price"]:
            continue
        yen = s["price"] * (fx_usd if s["cur"] == "USD" else 1.0)
        duty_act, duty_est = g(r, C_DUTY_ACT), g(r, C_DUTY_EST)
        duty = duty_act or duty_est
        proc = g(r, C_PROC_ACT) or g(r, C_PROC_EST)
        ship, fuel = g(r, C_SHIP), g(r, C_FUEL)
        out.append({
            "tid": tid, "file": os.path.basename(path),
            "date": order.get(tid, {}).get("date", ""),
            "dest": s["dest"], "hs": s["hs"], "origin": s["origin"], "title": s["title"],
            "declared": s["price"], "cur": s["cur"], "declared_jpy": round(yen),
            "skus": s["skus"],
            "status": "final" if duty_act else "estimate",
            "duty": duty, "duty_rate": round(duty / yen, 4) if yen else 0,
            "duty_estimated": duty_est, "fx_usd": fx_usd,
            "clearance": g(r, C_CLEAR), "proc_fee": proc,
            "proc_rate": round(proc / duty, 4) if duty else 0,
            "ship": ship, "fuel": fuel,
            "fuel_rate": round(fuel / ship, 4) if ship else 0,
            "incoterm": order.get(tid, {}).get("incoterm", ""),
        })
    return out


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    if "--dir" in sys.argv:
        i = sys.argv.index("--dir")
        args = sorted(glob.glob(os.path.join(sys.argv[i + 1], "*.xlsx")))
    if not args:
        print(__doc__)
        return 1

    rows = []
    for p in args:
        try:
            rows += parse(p)
        except Exception as e:                                  # noqa: BLE001
            print(f"⚠️ {os.path.basename(p)}: {e}")
    if not rows:
        print("明細0件 — 請求書の書式が変わった可能性")
        return 1

    print(f"{'日付':<11}{'仕向':<5}{'HS':<12}{'申告':>10}{'関税¥':>8}{'実効率':>8}"
          f"{'処理':>7}{'燃油':>7}  状態")
    for r in sorted(rows, key=lambda x: x["date"]):
        mark = "確定" if r["status"] == "final" else "★推定(未確定)"
        print(f"{r['date']:<11}{str(r['dest']):<5}{r['hs']:<12}"
              f"{r['declared']:>9,.2f}{r['duty']:>8,.0f}{r['duty_rate']*100:>7.2f}%"
              f"{r['proc_rate']*100:>6.2f}%{r['fuel_rate']*100:>6.1f}%  {mark}")
        print(f"           {r['title'][:64]}")

    # HTS ごとに率が違う (TCG 9504.40 = 10% / Tシャツ 6109.90 = 42%) ので HS 別に集計する。
    # 仕向地だけで平均すると 10% と 42% が混ざって「26%」という存在しない値になる。
    print("\n=== 仕向地 × HS の実効関税率 (確定分のみ) ===")
    agg = defaultdict(list)
    for r in rows:
        if r["status"] == "final":
            agg[(r["dest"], r["hs"][:6])].append(r["duty_rate"])
    if not agg:
        print("  確定分なし — すべて推定のまま。後日の 料金調整 を待つこと")
    for (dest, hs), v in sorted(agg.items()):
        lo, hi = min(v), max(v)
        flag = "" if hi - lo < 0.01 else "  ★同HSでばらつきあり"
        print(f"  {dest} HS{hs}: n={len(v)} 平均 {sum(v)/len(v)*100:5.2f}% "
              f"(min {lo*100:.2f}% / max {hi*100:.2f}%){flag}")

    pend = [r for r in rows if r["status"] == "estimate"]
    if pend:
        print(f"\n⚠️ 推定のまま {len(pend)}件 — **後から実額に差し替わる**。この率で採算を判断しない")
        for r in pend:
            print(f"   {r['date']} {r['dest']} HS{r['hs']} 推定 ¥{r['duty']:,.0f} "
                  f"({r['duty_rate']*100:.2f}%) — {r['title'][:44]}")

    HIST.parent.mkdir(parents=True, exist_ok=True)
    # 同じ tid が estimate → final に化けるので、key は (tid, status)。
    # tid だけだと推定値が居座り、実額が永久に記録されない。
    known = set()
    if HIST.exists():
        for ln in HIST.read_text(encoding="utf-8").splitlines():
            try:
                d = json.loads(ln)
                known.add((d["tid"], d.get("status", "")))
            except (ValueError, KeyError):
                pass
    new = [r for r in rows if (r["tid"], r["status"]) not in known]
    with HIST.open("a", encoding="utf-8") as fh:
        for r in new:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"\n履歴に追記: {len(new)}件 (既知 {len(rows)-len(new)}件はスキップ) → {HIST}")
    print("※ レートは Orange Connex の『事前設定レート』で、料金表には記載されない。"
          "請求書が唯一の一次情報 (RATE_GUIDE L430-437)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
