#!/usr/bin/env python3
"""
iMak Trading Japan - 利益計算パラメータ SSOT (Single Source of Truth)

データソース優先順位:
  1. Google Sheets (PRIMARY): 利益計算シート v2_GS
     https://docs.google.com/spreadsheets/d/1ft91iIsJjbMVw3Gx4GmeO-DQ0A47jp6O1TbiZeTslag/
  2. ローカルキャッシュ (cache/profit_params_cache.json): GS取得結果を1時間保持
  3. Excel フォールバック: iMakHQ/sheets/【NEW】利益計算シート_v2.xlsx
  4. ハードコードフォールバック: 全部失敗時の最終値
"""
import json
import os
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    gspread = None

SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_ROOT = SCRIPT_DIR.parent

GSHEET_URL = "https://docs.google.com/spreadsheets/d/1ft91iIsJjbMVw3Gx4GmeO-DQ0A47jp6O1TbiZeTslag/edit"
CREDS_PATH = WORKSPACE_ROOT / "double-hold-421922-7c0d38d3f73d.json"
GSCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

CACHE_DIR = SCRIPT_DIR / "cache"
CACHE_FILE = CACHE_DIR / "profit_params_cache.json"
CACHE_TTL_SECONDS = 3600

SPREADSHEET = WORKSPACE_ROOT / "iMakHQ" / "sheets" / "【NEW】利益計算シート_v2.xlsx"

FALLBACK_EXCHANGE_RATE = 159.245
FALLBACK_AD_RATE = 0.10
FALLBACK_PAYO_FEE = 0.025
FALLBACK_TARGET_PROFIT = 0.10
INTL_FEE = 0.02

FALLBACK_CATEGORIES = {
    "TCG(PSA10)": (0.1325, 2000),
    "G-SHOCK": (0.1325, 2000),
    "Tシャツ(UT)": (0.153, 2000),
    "Montbell(一般)": (0.153, 2000),
    "Montbell(ジャケット)": (0.153, 4500),
    "一番くじ": (0.1325, 2500),
    "フィギュア": (0.1325, 3500),
    "ユニクロ(非UT)": (0.153, 2000),
    "ヴィンテージ玩具": (0.1325, 2500),
    "トミカ": (0.1325, 2000),
    "POPMart": (0.1325, 2500),
    "ガシャポン": (0.1325, 2000),
    "ダイソー": (0.1325, 2000),
    "バッグ(アネロ)": (0.153, 2500),
}

_cache = None


def _default_cache():
    return {
        "exchange_rate": FALLBACK_EXCHANGE_RATE,
        "ad_rate": FALLBACK_AD_RATE,
        "payo_fee": FALLBACK_PAYO_FEE,
        "target_profit": FALLBACK_TARGET_PROFIT,
        "categories": dict(FALLBACK_CATEGORIES),
        "source": "fallback",
    }


def _save_local_cache(cache_data):
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        with CACHE_FILE.open("w", encoding="utf-8") as f:
            json.dump({"ts": time.time(), "data": cache_data}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _load_local_cache(allow_stale=False):
    if not CACHE_FILE.exists():
        return None
    try:
        with CACHE_FILE.open("r", encoding="utf-8") as f:
            payload = json.load(f)
        age = time.time() - payload.get("ts", 0)
        if age > CACHE_TTL_SECONDS and not allow_stale:
            return None
        data = payload["data"]
        if "categories" in data:
            data["categories"] = {k: tuple(v) for k, v in data["categories"].items()}
        return data
    except Exception:
        return None


def _load_from_gsheet():
    if gspread is None or not CREDS_PATH.exists():
        return None
    try:
        creds = Credentials.from_service_account_file(str(CREDS_PATH), scopes=GSCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_url(GSHEET_URL)
        ranges = ['設定!B2:B5', '設定!F2', '設定!H2', '設定!J2', '設定!A17:C40']
        result = sh.values_batch_get(ranges, params={'valueRenderOption': 'UNFORMATTED_VALUE'})
        vals = result['valueRanges']
        b_vals = vals[0].get('values', [])
        usd = float(b_vals[0][0]) if len(b_vals) > 0 and b_vals[0] else FALLBACK_EXCHANGE_RATE
        ad = float(b_vals[1][0]) if len(b_vals) > 1 and b_vals[1] else FALLBACK_AD_RATE
        payo = float(b_vals[2][0]) if len(b_vals) > 2 and b_vals[2] else FALLBACK_PAYO_FEE
        tgt = float(b_vals[3][0]) if len(b_vals) > 3 and b_vals[3] else FALLBACK_TARGET_PROFIT
        eur = float(vals[1].get('values', [[None]])[0][0]) if vals[1].get('values') else None
        gbp = float(vals[2].get('values', [[None]])[0][0]) if vals[2].get('values') else None
        aud = float(vals[3].get('values', [[None]])[0][0]) if vals[3].get('values') else None
        cat_rows = vals[4].get('values', [])
        categories = {}
        for row in cat_rows:
            if len(row) < 3:
                continue
            name = str(row[0]).strip() if row[0] else ""
            try:
                fvf = float(row[1])
                ship = int(row[2])
                if name:
                    categories[name] = (fvf, ship)
            except (TypeError, ValueError):
                continue
        cache = {
            "exchange_rate": usd,
            "exchange_rate_eur": eur,
            "exchange_rate_gbp": gbp,
            "exchange_rate_aud": aud,
            "ad_rate": ad,
            "payo_fee": payo,
            "target_profit": tgt,
            "categories": categories if categories else dict(FALLBACK_CATEGORIES),
            "source": "gsheet",
        }
        _save_local_cache(cache)
        return cache
    except Exception:
        return None


def _load_from_excel():
    if openpyxl is None or not SPREADSHEET.exists():
        return None
    try:
        wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
        ws = wb["設定"] if "設定" in wb.sheetnames else wb.active
        cache = _default_cache()
        if ws["B2"].value is not None:
            cache["exchange_rate"] = float(ws["B2"].value)
        if ws["B3"].value is not None:
            cache["ad_rate"] = float(ws["B3"].value)
        if ws["B4"].value is not None:
            cache["payo_fee"] = float(ws["B4"].value)
        if ws["B5"].value is not None:
            cache["target_profit"] = float(ws["B5"].value)
        for src_cell, key in [("F2", "exchange_rate_eur"), ("H2", "exchange_rate_gbp"), ("J2", "exchange_rate_aud")]:
            try:
                v = ws[src_cell].value
                if v is not None:
                    cache[key] = float(v)
            except Exception:
                pass
        categories = {}
        for row in range(17, 60):
            name = ws[f"A{row}"].value
            fvf = ws[f"B{row}"].value
            ship = ws[f"C{row}"].value
            if not name or fvf is None or ship is None:
                continue
            try:
                categories[str(name).strip()] = (float(fvf), int(ship))
            except (TypeError, ValueError):
                continue
        if categories:
            cache["categories"] = categories
        cache["source"] = "excel"
        return cache
    except Exception:
        return None


def _load():
    global _cache
    if _cache is not None:
        return _cache
    cached = _load_local_cache(allow_stale=False)
    if cached:
        cached["source"] = cached.get("source", "gsheet") + "/cache"
        _cache = cached
        return cached
    gs_data = _load_from_gsheet()
    if gs_data:
        _cache = gs_data
        return gs_data
    stale = _load_local_cache(allow_stale=True)
    if stale:
        stale["source"] = stale.get("source", "gsheet") + "/stale"
        _cache = stale
        return stale
    excel_data = _load_from_excel()
    if excel_data:
        _cache = excel_data
        return excel_data
    _cache = _default_cache()
    return _cache


def get_exchange_rate(currency="USD"):
    cache = _load()
    key_map = {"USD": "exchange_rate", "EUR": "exchange_rate_eur",
               "GBP": "exchange_rate_gbp", "AUD": "exchange_rate_aud"}
    key = key_map.get(currency.upper(), "exchange_rate")
    val = cache.get(key)
    if val is None:
        val = cache.get("exchange_rate", FALLBACK_EXCHANGE_RATE)
    return val


def get_category_params(category):
    cache = _load()
    if category in cache["categories"]:
        fvf, ship = cache["categories"][category]
        return {"fvf": fvf, "shipping_jpy": ship}
    return None


def get_net_ratio(category):
    cache = _load()
    params = get_category_params(category)
    if params is None:
        return None
    return 1 - params["fvf"] - INTL_FEE - cache["ad_rate"] - cache["payo_fee"] - cache["target_profit"]


def get_effective_fvf(category):
    params = get_category_params(category)
    if params is None:
        return None
    return params["fvf"] + INTL_FEE


def compute_min_price_usd(cost_jpy, category):
    cache = _load()
    params = get_category_params(category)
    if params is None:
        raise ValueError(f"Unknown category: {category}")
    net_ratio = get_net_ratio(category)
    return (cost_jpy + params["shipping_jpy"]) / (cache["exchange_rate"] * net_ratio)


def get_source():
    return _load()["source"]


def force_refresh():
    global _cache
    _cache = None
    if CACHE_FILE.exists():
        try:
            CACHE_FILE.unlink()
        except Exception:
            pass
    return _load()


EXCHANGE_RATE = property(lambda self: get_exchange_rate())


if __name__ == "__main__":
    import sys, io
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    cache = _load()
    print(f"Source: {cache['source']}")
    print(f"USD/JPY: {cache.get('exchange_rate')}")
    print(f"EUR/JPY: {cache.get('exchange_rate_eur')}")
    print(f"GBP/JPY: {cache.get('exchange_rate_gbp')}")
    print(f"AUD/JPY: {cache.get('exchange_rate_aud')}")
    for name, (fvf, ship) in sorted(cache["categories"].items()):
        net = get_net_ratio(name)
        print(f"  {name:25s} FVF={fvf:.4f} Ship=¥{ship} NET={net:.4f}")
