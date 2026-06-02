#!/usr/bin/env python3
"""価格決定エンジン（全プロジェクト共通SSOT）

設計思想:
  価格 = コストプラス（仕入 + 送料 + 手数料 + 価格帯別利益率上限）
  乖離 ≤ 価格帯別乖離率上限 → 出品OK
  乖離 > 価格帯別乖離率上限 → 見送りアラート

参照SSOT:
  - 利益計算シートv2.xlsx: 為替/手数料/プロモ/Payo（profit_params.py 経由）
  - GATE判定パラメータ検討.xlsx: 価格帯別 利益率上限 / 乖離率上限

使い方:
  from pricing_engine import compute_listing_price
  result = compute_listing_price(cost_jpy=1500, median_usd=25, category="Tシャツ(UT)")
  result["price"]      → 出品価格
  result["status"]     → "GO" / "ALERT" / "NO_MEDIAN"
  result["alert_msg"]  → アラート時の人間向けメッセージ
"""
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# profit_params 経由で利益計算シートv2を参照
sys.path.insert(0, SCRIPT_DIR)
from profit_params import _load as _profit_load, get_category_params, INTL_FEE

# GATE.xlsx パス
GATE_XLSX_PATH = os.path.join(SCRIPT_DIR, "..", "iMakHQ", "sheets", "GATE判定パラメータ検討.xlsx")

# フォールバック値（GATE.xlsx読込失敗時）
_TIER_FALLBACK = [
    (39, 0.25, 0.50), (60, 0.25, 0.50), (100, 0.20, 0.50), (200, 0.15, 0.50),
    (300, 0.15, 0.40), (400, 0.10, 0.25), (500, 0.10, 0.20), (600, 0.10, 0.15),
    (800, 0.10, 0.10), (9999, 0.10, 0.10),
]


def _load_tier_params():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(GATE_XLSX_PATH, read_only=True, data_only=True)
        if "確定値" not in wb.sheetnames:
            return _TIER_FALLBACK
        ws = wb["確定値"]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                out.append((float(row[0]), float(row[1]), float(row[2])))
            except (TypeError, ValueError):
                continue
        return sorted(out, key=lambda x: x[0]) if out else _TIER_FALLBACK
    except Exception:
        return _TIER_FALLBACK


TIER_PARAMS = _load_tier_params()


def get_tier_params(price_usd):
    """価格(USD)から(profit_target, gap_limit)を取得。"""
    for threshold, profit_target, gap_limit in TIER_PARAMS:
        if price_usd <= threshold:
            return profit_target, gap_limit
    return 0.10, 0.10


def _ddp_usd_for_price(price_usd: float) -> float:
    """価格 tier から DDP 金額 ($) を逆引き (= global.yaml ddp_shipping_tiers)。

    Free Shipping mode 移行 (2026-05-18) 後、出品価格は bundled (= item + DDP)。
    DDP 実コストは seller が DHL に支払うため、利益計算に含める必要がある。
    """
    cache = _profit_load()
    tiers = cache.get("ddp_shipping_tiers", [])
    if not tiers:
        # fallback: 古い yaml 構造 or 読込失敗時
        defaults = [(39, 15), (60, 20), (100, 30), (200, 50), (300, 75),
                    (400, 105), (500, 130), (600, 155), (800, 210), (1000, 260)]
        for max_usd, ddp in defaults:
            if price_usd <= max_usd:
                return float(ddp)
        return 260.0
    for tier in tiers:
        if price_usd <= tier.get("max_usd", 0):
            return float(tier.get("ddp_usd", 0))
    return float(tiers[-1].get("ddp_usd", 260))


def _compute_target_usd(cost_jpy, category, profit_target):
    """コストプラスで最低出品価格を計算 (Free Shipping bundled price)。

    根本修正 (2026-05-18): DDP実コスト + 国内送料 + 仕入 を含めた total cost
    から target_profit を確保できる bundled price (= item + DDP 込) を出力。

    式: bundled_price × FX × (1 - fees - profit_target) = cost + 国内送料 + DDP実
    DDP は bundled_price tier に依存 → 5 回 iteration で収束。
    """
    cache = _profit_load()
    params = get_category_params(category)
    if params is None:
        raise ValueError(f"Unknown category: {category}")
    net_ratio = 1 - params["fvf"] - INTL_FEE - cache["ad_rate"] - cache["payo_fee"] - profit_target
    if net_ratio <= 0:
        raise ValueError(f"net_ratio<=0 (profit_target {profit_target} too high)")
    fx = cache["exchange_rate"]
    base_cost = cost_jpy + params["shipping_jpy"]

    # DDP iteration: 価格 ↔ DDP の循環依存を 5 回反復で収束
    price = base_cost / (fx * net_ratio)  # 初期推定 (= 旧式、DDP なし)
    for _ in range(5):
        ddp_usd = _ddp_usd_for_price(price)
        ddp_jpy = ddp_usd * fx
        new_price = (base_cost + ddp_jpy) / (fx * net_ratio)
        if abs(new_price - price) < 0.01:
            break
        price = new_price
    return price


def _round_98(price_usd):
    """$X.98 形式に丸め。$10以下はそのまま round。"""
    p = round(price_usd, 2)
    if p > 10:
        return int(p) + 0.98
    return p


# ============================================================================
# V5 価格決定 (2026-05-19 確定): 35% 一律 markup + 仕入ベース IFS 利益率
# 詳細: V5 spreadsheet 1dFU... row 22 / yaml v5_pricing section
# ============================================================================
def _profit_rate_from_cost_v5(cost_jpy: float) -> float:
    """yaml v5_pricing.profit_rate_ifs から仕入¥に応じた利益率を引く."""
    import config_loader
    ifs = config_loader.get_v5_pricing().get("profit_rate_ifs", [])
    for tier in ifs:
        if cost_jpy < tier["cost_max"]:
            return float(tier["rate"])
    return 0.09  # fallback


def _fvf_effective_v5(category: str, country: str = "US") -> float:
    """V5 G3 formula 完全再現: (FVF_base or G-SHOCK 上書き + Intl + Reg) × (1+VAT)."""
    import config_loader
    v5 = config_loader.get_v5_pricing()
    cache = _profit_load()
    # G-SHOCK 国別 上書き
    if category == "G-SHOCK":
        override = v5.get("gshock_country_fvf", {}).get(country)
        if override is not None:
            fvf_base = float(override)
        else:
            fvf_base = get_category_params(category)["fvf"]
    else:
        fvf_base = get_category_params(category)["fvf"]
    # 国別 fees
    country_fees = v5.get("country_fees", {}).get(country, {"vat": 0, "intl": INTL_FEE, "reg": 0})
    return (fvf_base + float(country_fees["intl"]) + float(country_fees["reg"])) * (1 + float(country_fees["vat"]))


def _compute_target_usd_v5(cost_jpy: float, category: str, country: str = "US") -> float:
    """V5 closed-form 価格計算 (= iteration なし、35% markup 一律).

    式:
      利益¥ = 仕入¥ × IFS(cost)
      C = (仕入 + 国内送料 + insertion + 利益¥) / (1 - markup × (FVF_eff + ad + payo))
      G = C × markup_factor  (= US は ×1.35、非US は ×1.0)
      F (USD) = G / FX
    """
    import config_loader
    v5 = config_loader.get_v5_pricing()
    cache = _profit_load()
    params = get_category_params(category)
    if params is None:
        raise ValueError(f"Unknown category: {category}")
    fx = cache["exchange_rate"]
    shipping_jpy = params["shipping_jpy"]
    fvf_eff = _fvf_effective_v5(category, country)
    insertion_jpy = 0.4 * fx
    rate = _profit_rate_from_cost_v5(cost_jpy)
    profit_jpy = cost_jpy * rate
    # markup_factor: US は 1.35、他は 1.0
    if country == "US":
        mf = 1 + float(v5.get("markup_rate_us", 0.35))
    else:
        mf = 1 + float(v5.get("markup_rate_other", 0.0))
    denom = 1 - mf * (fvf_eff + cache["ad_rate"] + cache["payo_fee"])
    if denom <= 0:
        raise ValueError(f"net_ratio<=0 (cat={category}, country={country})")
    base = cost_jpy + shipping_jpy + insertion_jpy + profit_jpy
    C = base / denom
    G = C * mf
    return G / fx


def compute_listing_price_v5(cost_jpy, median_usd, category, gap_limit_override=None, country="US"):
    """V5 ロジックで出品価格決定 (= 35% markup + IFS 利益率 closed-form).

    互換 API: compute_listing_price と同じ signature (= country 引数追加).
    Returns: 同 dict 形式.
    """
    target_usd = _compute_target_usd_v5(cost_jpy, category, country)
    price = _round_98(target_usd)
    # 利益再計算 (確定 F での実利益)
    cache = _profit_load()
    params = get_category_params(category)
    fx = cache["exchange_rate"]
    fvf_eff = _fvf_effective_v5(category, country)
    revenue_jpy = price * fx
    import config_loader
    v5 = config_loader.get_v5_pricing()
    mf = 1 + (float(v5.get("markup_rate_us", 0.35)) if country == "US" else float(v5.get("markup_rate_other", 0)))
    ddp_jpy = revenue_jpy - revenue_jpy / mf  # = revenue × (1 - 1/mf)
    insertion_jpy = 0.4 * fx
    cost_total = cost_jpy + params["shipping_jpy"] + insertion_jpy + revenue_jpy * (fvf_eff + cache["ad_rate"] + cache["payo_fee"]) + ddp_jpy
    profit_jpy = revenue_jpy - cost_total
    profit_target = _profit_rate_from_cost_v5(cost_jpy)
    # 中央値ゲート
    if not median_usd or median_usd <= 0:
        return {"price": price, "target_usd": target_usd, "profit_target": profit_target,
                "profit_jpy": profit_jpy, "gap_pct": None, "gap_limit_pct": None,
                "status": "NO_MEDIAN", "alert_msg": None}
    _, gap_limit = get_tier_params(price)
    if gap_limit_override is not None:
        gap_limit = gap_limit_override
    gap_pct = (target_usd - median_usd) / median_usd * 100
    gap_limit_pct = gap_limit * 100
    if gap_pct <= gap_limit_pct:
        return {"price": price, "target_usd": target_usd, "profit_target": profit_target,
                "profit_jpy": profit_jpy, "gap_pct": gap_pct, "gap_limit_pct": gap_limit_pct,
                "status": "GO", "alert_msg": None}
    return {"price": price, "target_usd": target_usd, "profit_target": profit_target,
            "profit_jpy": profit_jpy, "gap_pct": gap_pct, "gap_limit_pct": gap_limit_pct,
            "status": "ALERT",
            "alert_msg": f"乖離率超過: 当社${price:.2f} vs 中央値${median_usd:.2f} = +{gap_pct:.0f}% (上限+{gap_limit_pct:.0f}%) → 見送り推奨"}


# ============================================================================
# V6 価格決定 (2026-05-20 確定): paid shipping + Policy tier + group A/B/C HTS
# 詳細: V5 spreadsheet 1dFU-0Zl... 設定!HTS_RATE + yaml v6_pricing
# - 商品本体 C ($) を closed-form で算出 (V5 G9 equation 同式)
# - F ($) = case A: INT(C)+0.98 / case B: INT(C + DDP_from_C×(1-split))+0.98
# - D ($) = Policy = tier_upper(F) × hts × 1.021 × split + 1.5
# - 利益 = 仕入 × IFS (target、V5 P9 と一致)
# ============================================================================
def _v6_resolve_category(category: str, title: str = "") -> str:
    """title-based override を最優先で適用 (HIGHT/LOW R列 曖昧時)."""
    import re
    import config_loader
    v6 = config_loader.get_v6_pricing()
    overrides = v6.get("title_overrides", [])
    t = (title or "").lower()
    for ov in overrides:
        if re.search(ov["pattern"], t, re.IGNORECASE):
            return ov["category"]
    return category


def _v6_group(category: str) -> str:
    """カテゴリ → グループ A/B/C."""
    import config_loader
    v6 = config_loader.get_v6_pricing()
    mapping = v6.get("category_to_group", {})
    return mapping.get(category, "A")  # default A


def _v6_tier(price_usd: float) -> tuple:
    """価格 → (tier_name, upper_usd)."""
    import config_loader
    v6 = config_loader.get_v6_pricing()
    uppers = v6.get("price_tier_uppers", [])
    for i, upper in enumerate(uppers, 1):
        if price_usd <= upper:
            return f"P{i:02d}", upper
    return "P31", uppers[-1] if uppers else 1500


def _v6_policy_shipping_usd(group: str, tier_upper: float) -> float:
    """Policy 送料 = upper × hts × 1.021 × split + 1.5."""
    import config_loader
    v6 = config_loader.get_v6_pricing()
    grp = v6.get("groups", {}).get(group, {})
    return tier_upper * grp.get("hts_rate", 0.18) * 1.021 * grp.get("split", 1.0) + 1.5


def _v6_policy_name(group: str, tier: str) -> str:
    """Policy 名生成 (= "DDP-A-P09" 等)、B→A remap 適用."""
    import config_loader
    v6 = config_loader.get_v6_pricing()
    fmt = v6.get("policy_name_format", "DDP-{group}-{tier}")
    raw = fmt.format(group=group, tier=tier)
    return v6.get("policy_remap", {}).get(raw, raw)


def _compute_listing_price_v6(cost_jpy: float, category: str, country: str = "US",
                                title: str = "") -> dict:
    """V6 closed-form 価格計算 (= V5 spreadsheet row 9 と完全同式).

    Returns dict:
      price          : 出品価格 (item-only USD, $X.98)
      shipping_usd   : Policy 送料 USD
      buyer_total_usd: F + Policy
      shipping_profile_name: eBay Policy 名 (= "DDP-A-P09" 等)
      group, tier    : デバッグ用
      profit_jpy     : 仕入 × IFS (target、V5 P9 と一致)
      profit_rate    : IFS 利益率
    """
    import config_loader
    v6 = config_loader.get_v6_pricing()
    if not v6:
        raise ValueError("v6_pricing 設定が見つかりません")
    # categorize (title override 適用)
    category = _v6_resolve_category(category, title)
    group = _v6_group(category)
    grp_cfg = v6["groups"][group]
    hts = float(grp_cfg["hts_rate"])
    split = float(grp_cfg["split"])
    m = 1 + hts * 1.021

    cache = _profit_load()
    fx = cache["exchange_rate"]
    params = get_category_params(category)
    if params is None:
        raise ValueError(f"Unknown category: {category}")
    shipping_jpy_v = params["shipping_jpy"]
    fvf_eff = _fvf_effective_v5(category, country)
    insertion_jpy = float(v6.get("insertion_usd", 0.4)) * fx
    clearance_jpy = float(v6.get("clearance_jpy", 245))
    mr = fvf_eff + cache["ad_rate"] + cache["payo_fee"]
    profit_rate = _profit_rate_from_cost_v5(cost_jpy)
    P_target_jpy = cost_jpy * profit_rate

    denom = 1 - mr * m
    if denom <= 0:
        raise ValueError(f"net_ratio<=0 (cat={category}, group={group})")
    # G_buyer ¥ = (H + J + ins + clear×mr + P) × m / (1 - mr×m) + clear
    G_buyer_jpy = (cost_jpy + shipping_jpy_v + insertion_jpy + clearance_jpy * mr + P_target_jpy) * m / denom + clearance_jpy
    # C ($) = (G - clear) / m / FX
    C_usd = (G_buyer_jpy - clearance_jpy) / m / fx
    # F ($) = case A: INT(C)+0.98 / case B: INT(C + DDP_from_C×(1-split)) + 0.98
    if split == 1.0:
        F_usd = int(C_usd) + 0.98
    else:
        DDP_from_C = C_usd * hts * 1.021 + clearance_jpy / fx
        F_usd = int(C_usd + DDP_from_C * (1 - split)) + 0.98

    tier, upper = _v6_tier(F_usd)
    D_usd = round(_v6_policy_shipping_usd(group, upper), 2)
    buyer_total = round(F_usd + D_usd, 2)
    policy_name = _v6_policy_name(group, tier)

    return {
        "price": F_usd,
        "shipping_usd": D_usd,
        "buyer_total_usd": buyer_total,
        "shipping_profile_name": policy_name,
        "group": group,
        "tier": tier,
        "tier_upper_usd": upper,
        "profit_jpy": round(P_target_jpy, 0),
        "profit_rate": profit_rate,
        "category_resolved": category,
    }


def compute_listing_price_v6(cost_jpy, median_usd, category, gap_limit_override=None,
                              country="US", title=""):
    """V6 ロジック (= paid shipping + Policy tier).

    互換 API: compute_listing_price と同じ signature (+ country/title 引数).
    """
    calc = _compute_listing_price_v6(cost_jpy, category, country=country, title=title)
    price = calc["price"]
    target_usd = price  # V6 では INT 適用後の値が target
    profit_jpy = calc["profit_jpy"]
    profit_target = calc["profit_rate"]
    # 中央値ゲート
    if not median_usd or median_usd <= 0:
        return {
            "price": price, "target_usd": target_usd, "profit_target": profit_target,
            "profit_jpy": profit_jpy, "gap_pct": None, "gap_limit_pct": None,
            "status": "NO_MEDIAN", "alert_msg": None,
            "shipping_usd": calc["shipping_usd"],
            "buyer_total_usd": calc["buyer_total_usd"],
            "shipping_profile_name": calc["shipping_profile_name"],
            "group": calc["group"], "tier": calc["tier"],
            "category_resolved": calc["category_resolved"],
        }
    _, gap_limit = get_tier_params(price)
    if gap_limit_override is not None:
        gap_limit = gap_limit_override
    gap_pct = (target_usd - median_usd) / median_usd * 100
    gap_limit_pct = gap_limit * 100
    status = "GO" if gap_pct <= gap_limit_pct else "ALERT"
    alert_msg = None if status == "GO" else (
        f"乖離率超過: 当社${price:.2f} vs 中央値${median_usd:.2f} = +{gap_pct:.0f}% "
        f"(上限+{gap_limit_pct:.0f}%) → 見送り推奨"
    )
    return {
        "price": price, "target_usd": target_usd, "profit_target": profit_target,
        "profit_jpy": profit_jpy, "gap_pct": gap_pct, "gap_limit_pct": gap_limit_pct,
        "status": status, "alert_msg": alert_msg,
        "shipping_usd": calc["shipping_usd"],
        "buyer_total_usd": calc["buyer_total_usd"],
        "shipping_profile_name": calc["shipping_profile_name"],
        "group": calc["group"], "tier": calc["tier"],
        "category_resolved": calc["category_resolved"],
    }


def compute_listing_price(cost_jpy, median_usd, category, gap_limit_override=None):
    """出品価格を決定する（全プロジェクト共通SSOT）。

    Args:
      cost_jpy: 仕入価格（円）
      median_usd: eBay市場中央値（USD）。0 or None なら中央値なしモード
      category: 利益計算v2のカテゴリ名（例: "Tシャツ(UT)"）
      gap_limit_override: 乖離許容率の category 別上書き値 (float, 比率).
        None なら従来通り価格帯 tier から取得.
        例: 一番くじ等の collectibles は新発売 median が信頼薄いため override で緩和.
        2026-05-10: 一番くじ専用に gap_limit_override=10.0 (=中央値の 11 倍まで OK)

    Returns:
      dict {
        price:          確定出品価格（USD, $X.98丸め済）
        target_usd:     コストプラス計算結果（丸め前）
        profit_target:  適用された利益率
        profit_jpy:     見込利益（JPY）
        gap_pct:        中央値との乖離率（%）None=中央値なし
        gap_limit_pct:  当該ティアの乖離率上限（%）
        status:         "GO" / "ALERT" / "NO_MEDIAN"
        alert_msg:      見送り推奨時の説明文（statusがALERTの時のみ）
      }
    """
    # V6 / V5 / V4 dispatch (= yaml の enabled フラグで切替)
    import config_loader
    if config_loader.is_v6_pricing_enabled():
        return compute_listing_price_v6(cost_jpy, median_usd, category,
                                         gap_limit_override=gap_limit_override, country="US")
    if config_loader.is_v5_pricing_enabled():
        return compute_listing_price_v5(cost_jpy, median_usd, category,
                                          gap_limit_override=gap_limit_override, country="US")

    # 旧 V4 ロジック (= rollback 用 fallback)
    # 1. 価格決定: 価格自身からティア確定（イテレーション）
    #    まず最高利益率(25%)で仮計算 → 結果のティアで再計算 → 収束
    profit_target, _ = get_tier_params(0)  # 0なら最小ティア(最高利益率)
    for _ in range(5):  # 最大5回反復で収束
        target_usd = _compute_target_usd(cost_jpy, category, profit_target)
        new_profit, _ = get_tier_params(target_usd)
        if abs(new_profit - profit_target) < 0.001:
            break
        profit_target = new_profit
    target_usd = _compute_target_usd(cost_jpy, category, profit_target)
    price = _round_98(target_usd)

    # 利益計算 (DDP実コスト含む、2026-05-18 根本修正)
    cache = _profit_load()
    params = get_category_params(category)
    revenue_jpy = price * cache["exchange_rate"]
    ddp_jpy = _ddp_usd_for_price(price) * cache["exchange_rate"]
    profit_jpy = revenue_jpy * (1 - params["fvf"] - INTL_FEE - cache["ad_rate"] - cache["payo_fee"]) - (cost_jpy + params["shipping_jpy"] + ddp_jpy)

    # 2. 中央値による乖離判定
    if not median_usd or median_usd <= 0:
        return {
            "price": price,
            "target_usd": target_usd,
            "profit_target": profit_target,
            "profit_jpy": profit_jpy,
            "gap_pct": None,
            "gap_limit_pct": None,
            "status": "NO_MEDIAN",
            "alert_msg": None,
        }

    _, gap_limit = get_tier_params(price)
    if gap_limit_override is not None:
        gap_limit = gap_limit_override
    gap_pct = (target_usd - median_usd) / median_usd * 100
    gap_limit_pct = gap_limit * 100

    if gap_pct <= gap_limit_pct:
        status = "GO"
        alert_msg = None
    else:
        status = "ALERT"
        alert_msg = (
            f"乖離率超過: 当社${price:.2f} vs 中央値${median_usd:.2f} = +{gap_pct:.0f}% "
            f"(上限+{gap_limit_pct:.0f}%) → 見送り推奨"
        )

    return {
        "price": price,
        "target_usd": target_usd,
        "profit_target": profit_target,
        "profit_jpy": profit_jpy,
        "gap_pct": gap_pct,
        "gap_limit_pct": gap_limit_pct,
        "status": status,
        "alert_msg": alert_msg,
    }


if __name__ == "__main__":
    # 自己テスト
    cases = [
        (1500, 25, "Tシャツ(UT)", "Tシャツ低価格（GO想定）"),
        (1500, 10, "Tシャツ(UT)", "Tシャツ低価格・市場安すぎ（ALERT想定）"),
        (50000, 800, "TCG(PSA10)", "TCG高額（GO想定）"),
        (1500, 0, "Tシャツ(UT)", "中央値なし（NO_MEDIAN）"),
    ]
    for cost, median, cat, label in cases:
        r = compute_listing_price(cost, median, cat)
        print(f"\n=== {label} ===")
        print(f"  cost ¥{cost} median ${median} cat={cat}")
        print(f"  → 価格 ${r['price']} 利益¥{r['profit_jpy']:.0f} 利益率{r['profit_target']*100:.0f}% status={r['status']}")
        if r['alert_msg']:
            print(f"  ALERT: {r['alert_msg']}")
