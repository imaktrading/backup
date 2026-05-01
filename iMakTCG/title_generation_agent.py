"""title_generation_agent - SEO-aware タイトル生成エージェント (独立モジュール).

設計原則 (修正連鎖を生まないため):
  - 既存モジュール (psa_to_csv.build_title / strip_banned_words / pad_title) を一切修正しない
  - psa_to_csv は build_title() で生成したタイトルに refine_title() を 1 行通すだけで導入
  - 失敗時は元タイトルをそのまま返す (フォールバック耐性)

設計思想:
  - 目的: 売れるタイトル (eBay 検索ヒット + Error 240 回避)
  - PSA Subject は「参考」、絶対不変フィールドは character + card_number のみ
  - その他 Subject 要素 (Pk Set, PERFUME FEMUR 等) は SEO 観点で書換可

Phase 1: EBAY_FORBIDDEN_TERMS の置換 (Error 240 回避) + technique→character 置換
Phase 2: iMakKeywords PDF 上位語スコアリング (検索ボリューム加味)
Phase 3: TOP seller タイトル分析 (sold_data xlsx の頻出語抽出)
最終   : 仮説 (variants) 生成 → 多角スコア → 最良案採用 (card_identification_agent と同形)

使用例:
    from title_generation_agent import refine_title
    title = build_title(game, set_name, card_number, subject)
    title = refine_title(
        title, character=character, franchise=franchise,
        agent_warnings=vision_result.get('agent_warnings'),
    )
"""
from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Optional

# ============================================================================
# 設定
# ============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Phase 2: iMakKeywords PDF (pdftotext 変換済 .txt を参照)
KEYWORDS_TXT_TOYS = DATA_DIR / "toys_hobbies_2026q1.txt"

# Phase 3: TOP seller タイトル (iMakeBayAPI/sold_data/*.xlsx)
SOLD_DATA_DIR = SCRIPT_DIR.parent / "iMakeBayAPI" / "sold_data"

# Phase 2/3 のキャッシュ (パース結果を JSON 化、毎回 PDF/xlsx を読まない)
KEYWORD_POOL_CACHE = DATA_DIR / "title_keyword_pool.json"


# ============================================================================
# eBay 禁止/危険語 (Phase 1)
# ============================================================================
EBAY_FORBIDDEN_TERMS = [
    # bundle 誤判定 (Error 240)
    (r"\bMini[\s\-]?Tin\s+Pk\s+Set\b",   "Mini Tin"),       # P-112 Nami の事故
    (r"\bMini[\s\-]?Tin\s+Pack\s+Set\b", "Mini Tin"),
    (r"\bPk\s+Set\b",                    "Set"),
    (r"\bPack\s+Set\b",                  "Set"),
    (r"\bMulti\s+Pack\b",                ""),
    (r"\bLot\s+of\b",                    ""),
    (r"\bBundle\b",                      ""),
    (r"\bCollection\s+of\b",             "Collection"),
    # 2026-05-01: Pokemon rarity prefix 残存 (Elesa Sparkle 'Fa/' 事故).
    # PSA Subject 'FA/ELESA' 等を smart_titlecase した残骸 'Fa/' を title から除去.
    # card_name_normalizer は C:Card Name/Character は剥がすが、build_title は
    # raw subject 経由のため title 側に残る. 該当 prefix: FA AR SAR SR UR HR MR PR.
    (r"\b(?:FA|AR|SAR|SR|UR|HR|MR|PR)/", ""),
]


# ============================================================================
# 公開 API
# ============================================================================
def refine_title(
    title: str,
    *,
    character: Optional[str] = None,
    card_number: Optional[str] = None,
    franchise: Optional[str] = None,
    agent_warnings: Optional[list] = None,
    target_max: int = 80,
) -> str:
    """既存タイトルに NG フィルタ + SEO スコアリングを適用して返す.

    Args:
        title:          build_title() で生成済タイトル
        character:      agent 補正済キャラ名 (technique→character 差替 + override character 反映)
        card_number:    公式 card# (PSA cert# / Bandai 公式DB 由来) — title 中の余計な #NN 剥がしに使用
        franchise:      "One Piece TCG" / "Pokemon TCG" 等 (キーワードプール選択用)
        agent_warnings: card_identification_agent の警告 (将来用、現状未使用)
        target_max:     80字制限 (eBay Title 上限)

    Returns:
        洗練済タイトル (失敗時は元 title をそのまま返す = フォールバック耐性)
    """
    if not title:
        return title

    try:
        # Phase 1-A: NG 語フィルタ
        v_phase1 = _apply_ng_filter(title)
        # Phase 1-B: card# 以外の '#NN' 剥がし (Bonney "WEEKLY SHONEN JUMP '24-#35" 事故)
        v_phase1 = _strip_non_card_hashes(v_phase1, card_number)
        # Phase 1-D: 連続/近接重複語の dedupe (Anniversary Coll. Collection Card 二重重複)
        v_phase1 = _dedupe_consecutive_words(v_phase1)
        # Phase 1-E: 末尾孤立記号 trim (#35 剥がした残骸 '24-, 末尾ハイフン)
        v_phase1 = _trim_trailing_orphans(v_phase1)
        # Phase 1-C: character がタイトル不在なら末尾追加 (Hancock 'Boa Hancock' 補完)
        v_phase1 = _ensure_character_in_title(v_phase1, character, target_max)
        v_phase1 = _normalize(v_phase1, target_max)

        # Phase 2/3: キーワードプールでスコアリング → 最良案採用
        pool = _load_keyword_pool()
        if not pool:
            # PDF/sold_data 未整備 → Phase 1 だけ返す
            if v_phase1 != title:
                print(f"    🤖 [TitleAgent/P1] {title!r} → {v_phase1!r}")
            return v_phase1

        # 候補生成 (Phase 1 結果 + 高ランク語の追補 variants)
        candidates = _generate_candidates(v_phase1, pool, character, target_max, franchise)
        # スコアリング → 最良採用
        scored = [(c, _score_title(c, pool)) for c in candidates]
        scored.sort(key=lambda x: x[1], reverse=True)
        best, best_score = scored[0]
        best = _normalize(best, target_max)

        if best != title:
            print(
                f"    🤖 [TitleAgent] {title!r} → {best!r} "
                f"(score={best_score:.2f}, candidates={len(candidates)})"
            )
        return best

    except Exception as e:
        print(f"    ⚠️ [TitleAgent] 例外発生、元タイトル採用: {type(e).__name__}: {e}")
        return title


# ============================================================================
# Phase 1: NG フィルタ + technique→character 確保
# ============================================================================
def _apply_ng_filter(title: str) -> str:
    result = title
    for pattern, replacement in EBAY_FORBIDDEN_TERMS:
        result = re.sub(pattern, replacement, result, flags=re.IGNORECASE)
    return result


def _ensure_character_in_title(
    title: str, character: Optional[str], target_max: int = 80
) -> str:
    """character (override.specs.character or agent 補正済) がタイトル不在なら末尾追加.

    技名 Subject (Hancock 'Perfume Femur') や別表記 Subject 経由で
    本キャラ名がタイトルに入らないケースを救済。
    既にタイトル中にあれば何もしない (大半のカードはこのケース)。
    キャラ名追加で 80字超なら、末尾の filler ("Card"/"Holo"/"Foil") 1個を犠牲にして確保。
    それでも収まらなければ追加せず元 title を返す (中途半端な切り落とし禁止)。
    """
    if not character:
        return title
    if character.lower() in title.lower():
        return title
    candidate = f"{title} {character}"
    if len(candidate) <= target_max:
        return candidate
    # 80字超え → 末尾 filler 1個を捨てて再試行
    for filler in (" Card", " Cards", " Holo", " Foil"):
        if title.endswith(filler):
            stripped = title[: -len(filler)]
            candidate2 = f"{stripped} {character}"
            if len(candidate2) <= target_max:
                return candidate2
    # それでもダメなら諦める (元 title 維持、SEO 機会損失だが破損より良い)
    return title


def _strip_non_card_hashes(title: str, card_number: Optional[str]) -> str:
    """title 中の '#XXX' のうち、card_number 以外を全部除去.

    build_title は card# を最初の '#' 位置に置く慣習なので、最初の1つだけ残し、
    2つ目以降の '#XX' (例: 雑誌号数 '#35', カード序列 '#1 of 100') を削除。
    既存 selfcheck の「card# 不一致」誤検出 (Bonney "Weekly Shonen Jump '24-#35" 事故) を回避。
    """
    if not title:
        return title
    seen = False
    def repl(m):
        nonlocal seen
        if seen:
            return ""
        seen = True
        return m.group(0)
    result = re.sub(r"#[A-Za-z0-9\-]+", repl, title)
    return result


def _trim_trailing_orphans(title: str) -> str:
    """title 末尾の孤立記号 (hyphen, comma, apostrophe 等) と filler 語を除去.

    Bonney 事故 (#35 剥がした残骸 '24-) や PRB02 末尾 'Card' filler を整理。
    保守的: 単純 trim、複雑な再構成はしない。
    """
    if not title:
        return title
    # 末尾の hyphen + apostrophe + 空白 + 記号 を除去
    result = re.sub(r"[\s'\-,.]+$", "", title).strip()
    # 末尾の孤立 filler 語 ("Card" / "Cards") は残してよい (UNIQLO 等で使う場合あり)
    # ただし「Card Card」「Coll. Collection」等の重複は除去
    return result


def _dedupe_consecutive_words(title: str) -> str:
    """タイトル中の冗長な連続/近接語を1個に. 例:
        'Anniversary Coll. Collection Card' → 'Anniversary Collection Card'
        'Pokémon Pokemon' → 'Pokémon'
    語幹を比較して同義 → 後出を削除 (前を保護).
    """
    if not title:
        return title
    # 'Coll.' / 'Coll' / 'Collection' を語幹 'collect' で同一視
    # 'Card' / 'Cards' を語幹 'card'
    abbreviation_map = {
        "coll.": "collection",
        "coll":  "collection",
        "anniv.": "anniversary",
        "anniv":  "anniversary",
        "vol.":   "volume",
    }
    tokens = title.split()
    keep_indices = []
    seen_stems = set()
    for i, tok in enumerate(tokens):
        stem = tok.lower().strip(".,;:'\"")
        # 略号 → フル形に正規化して比較
        stem_full = abbreviation_map.get(stem, stem)
        # 同語幹が既出 → スキップ (連続 + 近接、両方カバー)
        if stem_full in seen_stems and len(stem_full) >= 4:  # 機能語 (a/of/and 等) は除外
            continue
        seen_stems.add(stem_full)
        keep_indices.append(i)
    deduped = " ".join(tokens[i] for i in keep_indices)
    return deduped


def _normalize(title: str, target_max: int) -> str:
    """空白正規化 + 80字制限."""
    result = re.sub(r"\s+", " ", title).strip()
    if len(result) > target_max:
        parts = result.split()
        while parts and len(" ".join(parts)) > target_max:
            parts.pop()
        result = " ".join(parts) if parts else result[:target_max]
    return result


# ============================================================================
# Phase 2: iMakKeywords PDF パーサ
# ============================================================================
_KEYWORD_LINE_RE = re.compile(
    r"^\s*(\d+)\s+(?:\d+|NEW)\s+(?:\d+|-|NEW)\s+(.+?)\s*$"
)


def _parse_keywords_pdf_txt(txt_path: Path) -> dict:
    """pdftotext 変換済 .txt から (rank, keyword) を抽出.

    Returns: {keyword_lower: score}, score = max(0, 1.0 - rank/200.0)
    """
    if not txt_path.exists():
        return {}
    pool = {}
    with open(txt_path, encoding="utf-8", errors="replace") as f:
        for line in f:
            m = _KEYWORD_LINE_RE.match(line)
            if not m:
                continue
            rank = int(m.group(1))
            keyword = m.group(2).strip().lower()
            # ノイズ除外: 数字のみ / 1文字 / "Rank" 等のヘッダ語
            if not keyword or keyword.isdigit() or len(keyword) < 2:
                continue
            if keyword in ("rank", "prev rank", "rank diff", "keyword"):
                continue
            score = max(0.0, 1.0 - rank / 200.0)
            # 同一 keyword 複数 rank → 最高 score 採用
            if keyword not in pool or pool[keyword] < score:
                pool[keyword] = score
    return pool


# ============================================================================
# Phase 3: TOP seller タイトル分析 (sold_data xlsx)
# ============================================================================
def _parse_sold_data_xlsx(xlsx_dir: Path, franchise_filter: Optional[str] = None) -> dict:
    """sold_data/*.xlsx から商品名 (タイトル) を読み、頻出 n-gram (1-3語) をスコア化.

    Returns: {term_lower: score}, score = freq / max_freq * 0.5 (PDF より低い weight)
    """
    if not xlsx_dir.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}

    title_freq = Counter()
    for xlsx in xlsx_dir.glob("*.xlsx"):
        # franchise filter (ファイル名にキーワード含むもののみ)
        if franchise_filter and franchise_filter.lower() not in xlsx.name.lower():
            continue
        try:
            wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
            sheet = wb.active
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            try:
                title_col = header.index("商品名")
            except ValueError:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if title_col >= len(row) or not row[title_col]:
                    continue
                title_text = str(row[title_col])
                # 1-2 語の n-gram 抽出 (3語以上は組合せ爆発するので除外)
                tokens = [
                    t.lower().strip(".,;:!?#")
                    for t in re.split(r"[\s\-/]+", title_text)
                    if len(t) >= 2 and not t.isdigit()
                ]
                for tok in tokens:
                    title_freq[tok] += 1
                for i in range(len(tokens) - 1):
                    bi = f"{tokens[i]} {tokens[i+1]}"
                    title_freq[bi] += 1
            wb.close()
        except Exception:
            continue

    if not title_freq:
        return {}
    max_freq = max(title_freq.values())
    pool = {
        term: (freq / max_freq) * 0.5
        for term, freq in title_freq.items()
        if freq >= 3 and len(term) >= 3  # 出現回数 3+ かつ 3文字+
    }
    return pool


# ============================================================================
# キーワードプール統合 + キャッシュ
# ============================================================================
def _load_keyword_pool(force_rebuild: bool = False) -> dict:
    """PDF + sold_data を統合した keyword pool を返す (キャッシュ済優先)."""
    if not force_rebuild and KEYWORD_POOL_CACHE.exists():
        try:
            with open(KEYWORD_POOL_CACHE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    pool_pdf = _parse_keywords_pdf_txt(KEYWORDS_TXT_TOYS)
    # franchise filter なし: TCG カテゴリ全体の sold_data 使う (one piece / pokemon 両方)
    pool_sold = _parse_sold_data_xlsx(SOLD_DATA_DIR)

    # 統合: 同一 term は max score 採用
    pool = dict(pool_pdf)
    for term, score in pool_sold.items():
        if pool.get(term, 0) < score:
            pool[term] = score

    # キャッシュ
    try:
        with open(KEYWORD_POOL_CACHE, "w", encoding="utf-8") as f:
            json.dump(pool, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return pool


def _score_title(title: str, pool: dict) -> float:
    """タイトル中に含まれる pool 語のスコア合計."""
    if not title or not pool:
        return 0.0
    title_lower = title.lower()
    score = 0.0
    for term, term_score in pool.items():
        if term in title_lower:
            score += term_score
    return score


# ============================================================================
# 候補生成 (Phase 2 統合)
# ============================================================================
# Franchise の「総称語」(安全に他カードに追加できる generic 語)
_FRANCHISE_GENERIC_TOKENS = {
    "one piece": {"one piece", "op-tcg", "op tcg"},
    "pokemon":   {"pokemon", "pokémon"},
    "dragon ball": {"dragon ball", "dbs", "dbz"},
    "gundam":    {"gundam"},
}

# Franchise の「キャラ名」(他キャラのカードに追加すると誤情報。必ず本カードのキャラのみ)
_FRANCHISE_CHARACTER_TOKENS = {
    "one piece": {"luffy", "nami", "zoro", "sanji", "robin", "chopper",
                  "hancock", "shanks", "ace", "law", "kid", "rebecca", "boa",
                  "shirahoshi", "yamato", "perona", "vivi", "uta", "kaido",
                  "smoker", "garp", "rayleigh"},
    "pokemon":   {"pikachu", "charizard", "mewtwo", "umbreon", "eevee",
                  "lillie", "elesa", "gengar", "mew", "blastoise", "venusaur",
                  "lugia", "rayquaza", "arceus", "ribombee", "iono", "marnie"},
    "dragon ball": {"goku", "vegeta", "gohan", "piccolo", "frieza",
                    "broly", "trunks", "bulma"},
    "gundam":    {"char", "amuro", "rx-78"},
}

# 後方互換: detect_franchise は generic + character の和集合を使う
_FRANCHISE_OWN_TOKENS = {
    fk: _FRANCHISE_GENERIC_TOKENS.get(fk, set()) | _FRANCHISE_CHARACTER_TOKENS.get(fk, set())
    for fk in _FRANCHISE_GENERIC_TOKENS
}


def _detect_franchise(title: str, franchise_hint: Optional[str]) -> Optional[str]:
    """タイトル + hint から franchise key (small case) を判定."""
    if franchise_hint:
        h = franchise_hint.lower()
        for key in _FRANCHISE_OWN_TOKENS:
            if key in h:
                return key
    title_lower = title.lower()
    for key, tokens in _FRANCHISE_OWN_TOKENS.items():
        if any(tok in title_lower for tok in tokens):
            return key
    return None


def _is_competing_franchise_term(term: str, my_franchise: Optional[str]) -> bool:
    """term が他フランチャイズ専属語なら True (混入禁止)."""
    if not my_franchise:
        return False
    term_lower = term.lower()
    for key, tokens in _FRANCHISE_OWN_TOKENS.items():
        if key == my_franchise:
            continue
        for tok in tokens:
            if tok in term_lower:
                return True
    return False


# Phase 2 で追補可能な「汎用 TCG SEO 語」(BANNED と重複しないもの)
# psa_to_csv の BANNED_TITLE_WORDS (gem mt/mint/graded/l@@k/look/wow/nr) に該当しないもののみ。
# Holo/Foil は Vision で確定時のみ追加すべきだが、Phase 2 では Vision 連携なしの単純語彙のみ扱う。
# 注意: token (単語) 単位で評価するため、複数語の term ("alt art" 等) はここに分割して入れる
#
# 2026-05-01: Title が 50-65 字で頭打ちする問題への対応で大幅拡張.
# 旧 15 単語 → 新 ~45 単語. TOP セラータイトル frequency 高い safe 語を追加.
_UNIVERSAL_SEO_PERMITTED_TOKENS = {
    # 既存
    "holo", "foil", "1st", "first", "edition",
    "tcg", "ccg", "english", "sealed", "alt", "alternate", "art",
    "card", "cards", "rare",
    # 2026-05-01 追加: 地域 (BANNED 解除と同期)
    "japanese", "jp", "jpn",
    # 2026-05-01 追加: 年 (TCG カードの release year は universally 検索される)
    "2019", "2020", "2021", "2022", "2023", "2024", "2025", "2026",
    # 2026-05-01 追加: Pokemon シリーズ era 名 (Pokémon TCG の世代マーカー、TOP セラー必須)
    "sword", "shield", "sun", "moon", "scarlet", "violet", "xy",
    # 2026-05-01 追加: 一般 set descriptor (誤情報リスク低、SEO 寄与高)
    "promo", "promos", "collection", "anniversary", "starter",
    # 2026-05-01 追加: その他 generic
    "japanese", "set",
}


def _is_term_relevant_to_franchise(
    term: str, my_franchise: Optional[str], my_character: Optional[str] = None
) -> bool:
    """term の「全 token」が安全リストに含まれるなら True (タイトル追補可).

    安全リストの内訳:
      - 汎用許可 token (holo/foil/english/card 等)
      - franchise 総称 token (pokemon/one/piece 等)
      - 本カードのキャラ名 token (luffy 等、本カードキャラのみ)

    一つでも未承認 token (例: ascended/heroes/charizard 等) が混入したら False。
    商品名 ('Ascended Heroes Pokemon Center ETB') / 別キャラ ('Charizard 151') 排除。
    """
    tokens = [t for t in re.split(r"\s+", term.lower()) if t]
    if not tokens:
        return False

    safe_tokens = set(_UNIVERSAL_SEO_PERMITTED_TOKENS)

    if my_franchise:
        # franchise 総称 token を分解して safe に追加
        for generic_term in _FRANCHISE_GENERIC_TOKENS.get(my_franchise, set()):
            safe_tokens.update(generic_term.lower().split())

        # 本カードのキャラ名 token のみ safe に追加 (他キャラは除外)
        if my_character:
            my_char_tokens = set(re.split(r"[\s\-/]+", my_character.lower()))
            char_set = _FRANCHISE_CHARACTER_TOKENS.get(my_franchise, set())
            for char_tok in char_set:
                if char_tok in my_char_tokens or any(
                    char_tok in mc for mc in my_char_tokens
                ):
                    safe_tokens.add(char_tok)

    # 全 token が safe に含まれるかチェック
    return all(tok in safe_tokens for tok in tokens)


def _has_token_overlap(term: str, base_lower: str) -> bool:
    """term の構成 token が 1個でも base に既出なら True (重複追加防止)."""
    term_tokens = set(re.split(r"\s+", term.lower()))
    base_tokens = set()
    for raw in re.split(r"[\s\-/]+", base_lower):
        base_tokens.add(raw.strip(".,;:!?#"))
    return bool(term_tokens & base_tokens)


def _generate_candidates(
    base_title: str,
    pool: dict,
    character: Optional[str],
    target_max: int,
    franchise: Optional[str] = None,
) -> list:
    """base_title を起点に、高ランク語を追補した variant 候補を生成.

    保守的方針:
      - base_title が 70字+ なら追補せず (タイトル既に密、改変リスク高い)
      - franchise 関連 OR 汎用許可語のみ追加 (競合 franchise 排除)
      - token レベル重複チェック (substring だけでなく)
    """
    candidates = [base_title]
    if not pool:
        return candidates

    # 既に密なタイトルは追補しない (誤情報追加リスク回避)
    if len(base_title) >= 70:
        return candidates

    my_franchise = _detect_franchise(base_title, franchise)
    base_lower = base_title.lower()
    space_left = target_max - len(base_title) - 1
    if space_left < 5:
        return candidates

    sorted_terms = sorted(pool.items(), key=lambda x: x[1], reverse=True)
    for term, score in sorted_terms[:200]:
        if score < 0.3:
            break
        # franchise relevance check (whitelist + character match)
        if not _is_term_relevant_to_franchise(term, my_franchise, character):
            continue
        # token レベル重複チェック
        if _has_token_overlap(term, base_lower):
            continue
        # length check
        title_cased = " ".join(w.capitalize() for w in term.split())
        if len(title_cased) + 1 > space_left:
            continue
        # 既存候補との部分一致チェック
        if any(title_cased.lower() in c.lower() for c in candidates):
            continue
        candidate = f"{base_title} {title_cased}"
        candidates.append(candidate)
        if len(candidates) >= 4:
            break

    return candidates


# ============================================================================
# CLI (動作確認用)
# ============================================================================
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "rebuild-pool":
        pool = _load_keyword_pool(force_rebuild=True)
        print(f"[OK] keyword pool rebuilt: {len(pool)} terms → {KEYWORD_POOL_CACHE}")
        sys.exit(0)

    # (title, character, card_number, franchise, agent_warnings)
    samples = [
        # P-112 Nami: Pk Set 事故 (Phase 1-A)
        ("PSA 10 One Piece TCG Promo Cards #P-112 Nami Mini-Tin Pk Set Vol.2 Bisai Promo",
         "Nami", "P-112", "One Piece TCG", None),
        # Bonney: Subject 末尾の #35 (雑誌号数) 剥がし (Phase 1-B)
        ("PSA 10 One Piece TCG #OP07-019 Jewelry Bonney Weekly Shonen Jump '24-#35",
         "Jewelry Bonney", "OP07-019", "One Piece TCG", None),
        # Hancock OP07-057: 技名 Subject + override character (Phase 1-C)
        ("PSA 10 One Piece TCG 500 Years in the Future #OP07-057 Perfume Femur Card",
         "Boa Hancock", "OP07-057", "One Piece TCG", None),
        # Pokemon 通常 (回帰確認: 何も変えない)
        ("PSA 10 Pokemon Battle Partners #105 Lillie's Ribombee Art Rare",
         "Lillie's Ribombee", "105", "Pokemon TCG", None),
        # 短いタイトル (誤追補なし回帰)
        ("PSA 10 Pokemon #001 Pikachu", "Pikachu", "001", "Pokemon TCG", None),
    ]
    for title, char, card_num, franchise, warns in samples:
        print(f"--- 入力: {title}")
        out = refine_title(title, character=char, card_number=card_num,
                           franchise=franchise, agent_warnings=warns)
        print(f"    出力: {out} ({len(out)}字)")
        print()
